from __future__ import annotations

import openpyxl

from src.conciliador.exporter import export_no_encontrados_xlsx


def test_export_no_encontrados_xlsx_splits_into_4_sheets_without_empty_side_columns(tmp_path):
    out = tmp_path / "no_encontrados.xlsx"
    result = {
        "no_encontrados": [
            {
                "Tipo no encontrado": "BANCO_SIN_RECIBO",
                "Origen": "BBVA",
                "Fecha movimiento": "2026-02-20",
                "Importe movimiento": 1000.0,
                "Detalle movimiento": "Detalle BBVA",
                "Fila Excel": 10,
            },
            {
                "Tipo no encontrado": "BANCO_SIN_RECIBO",
                "Origen": "MERCADOPAGO",
                "Fecha movimiento": "2026-02-21",
                "Importe movimiento": 2000.0,
                "Detalle movimiento": "Detalle MP",
                "Fila Excel": 20,
            },
            {
                "Tipo no encontrado": "BANCO_SIN_RECIBO",
                "Origen": "GALICIA",
                "Fecha movimiento": "2026-02-22",
                "Importe movimiento": 3000.0,
                "Detalle movimiento": "Detalle Galicia",
                "Fila Excel": 30,
            },
            {
                "Tipo no encontrado": "RECIBO_SIN_BANCO",
                "Empresa": "ALARCON",
                "Nro recibo": "68744",
                "Nro cliente": "30424",
                "Cliente": "Cliente Demo",
                "Medio de pago": "TRANSFERENCIA",
                "Fecha recibo": "2026-02-23",
                "Importe recibo": 575511.55,
                "Vendedor": "206 - Andres Dominguez",
                "Ítem en recibo": "",
            },
        ]
    }

    export_no_encontrados_xlsx(result, str(out))

    wb = openpyxl.load_workbook(str(out), data_only=False)
    assert wb.sheetnames == ["BBVA", "Mercado Pago", "Galicia", "Recibos sin banco"]

    ws_bbva = wb["BBVA"]
    ws_mp = wb["Mercado Pago"]
    ws_gal = wb["Galicia"]
    ws_rec = wb["Recibos sin banco"]

    headers_bbva = [c.value for c in ws_bbva[1]]
    headers_mp = [c.value for c in ws_mp[1]]
    headers_gal = [c.value for c in ws_gal[1]]
    headers_rec = [c.value for c in ws_rec[1]]

    for headers in (headers_bbva, headers_mp, headers_gal):
        assert "Origen" in headers
        assert "Fecha movimiento" in headers
        assert "Importe movimiento" in headers
        assert "Nro recibo" not in headers
        assert "Nro cliente" not in headers
        assert "Importe recibo" not in headers

    assert "Nro recibo" in headers_rec
    assert "Nro cliente" in headers_rec
    assert "Cliente" in headers_rec
    assert "Importe recibo" in headers_rec
    assert "Origen" not in headers_rec
    assert "Fecha movimiento" not in headers_rec
    assert "Importe movimiento" not in headers_rec
    cliente_col = headers_rec.index("Cliente") + 1
    assert ws_rec.cell(2, cliente_col).value == "Cliente Demo"


def test_export_no_encontrados_xlsx_formats_money_and_widens_columns(tmp_path):
    out = tmp_path / "no_encontrados_money.xlsx"
    result = {
        "no_encontrados": [
            {
                "Tipo no encontrado": "BANCO_SIN_RECIBO",
                "Origen": "BBVA",
                "Fecha movimiento": "2026-02-20",
                "Importe movimiento": "85500,00",
                "Detalle movimiento": "Transferencia de terceros con detalle largo para validar ancho",
                "Fila Excel": 1334,
            }
        ]
    }

    export_no_encontrados_xlsx(result, str(out))

    wb = openpyxl.load_workbook(str(out), data_only=False)
    ws = wb["BBVA"]
    headers = [c.value for c in ws[1]]
    importe_col = headers.index("Importe movimiento") + 1
    detalle_col = headers.index("Detalle movimiento") + 1
    importe_cell = ws.cell(2, importe_col)
    importe_col_letter = openpyxl.utils.get_column_letter(importe_col)
    detalle_col_letter = openpyxl.utils.get_column_letter(detalle_col)

    assert isinstance(importe_cell.value, (int, float))
    assert importe_cell.number_format == "#,##0.##"
    assert (ws.column_dimensions[importe_col_letter].width or 0) >= 18
    assert (ws.column_dimensions[detalle_col_letter].width or 0) >= 30


def test_export_no_encontrados_xlsx_formats_mp_dates_as_dd_mm_yyyy(tmp_path):
    out = tmp_path / "no_encontrados_mp_dates.xlsx"
    result = {
        "no_encontrados": [
            {
                "Tipo no encontrado": "BANCO_SIN_RECIBO",
                "Origen": "MERCADOPAGO",
                "Fecha movimiento": "2026-03-09",
                "Importe movimiento": 123.0,
                "Detalle movimiento": "Detalle MP",
                "Fila Excel": 20,
            }
        ]
    }

    export_no_encontrados_xlsx(result, str(out))

    wb = openpyxl.load_workbook(str(out), data_only=False)
    ws = wb["Mercado Pago"]
    headers = [c.value for c in ws[1]]
    fecha_col = headers.index("Fecha movimiento") + 1
    assert ws.cell(2, fecha_col).value == "09/03/2026"
