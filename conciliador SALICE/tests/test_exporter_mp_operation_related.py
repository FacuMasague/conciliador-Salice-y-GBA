from __future__ import annotations

import openpyxl
import zipfile
import xml.etree.ElementTree as ET

from src.conciliador.exporter import export_filled_bank_excel


def _build_workbook(path: str) -> None:
    wb = openpyxl.Workbook()
    ws0 = wb.active
    ws0.title = "SALICE BBVA"
    ws0.append(["x"])
    ws0.append(["x"])
    ws0.append(["", "A", "Número Documento", "Oficina", "Importe", "ok", "cliente", "recibo"])

    ws1 = wb.create_sheet(" ALARCON BBVA")
    ws1.append(["x"])
    ws1.append(["x"])
    ws1.append(["", "A", "Número Documento", "Oficina", "Importe", "ok", "cliente", "recibo"])

    ws2 = wb.create_sheet("SALICE GALICIA (ALARCON)")
    ws2.append(["Fecha", "Concepto", "Razon social", "CUIT", "Importe", "ok", "Cliente", "Recibo"])

    ws3 = wb.create_sheet("MercadoPago ")
    ws3.append(
        [
            "Fecha de Pago",
            "Tipo de Operación",
            "Número de Movimiento",
            "Operación Relacionada",
            "Unnamed: 4",
            "Recibio?",
            "Cliente",
            "Recibo",
            "NÚMERO DE IDENTIFICACIÓN DEL PAGADOR",
        ]
    )
    # valor en notación científica simulando origen problemático
    ws3.append(["20/02/2026 00:00:00", "Cobro", "", "1,4382E+11", 1000.0, "", "", "", "20999999888"])

    wb.save(path)


def test_export_mp_operacion_relacionada_is_text_not_scientific(tmp_path):
    original = tmp_path / "original.xlsx"
    out = tmp_path / "out.xlsx"
    _build_workbook(str(original))

    result = {
        "validados": [
            {
                "Origen": "MERCADOPAGO",
                "Fila Excel": 2,
                "Nro cliente": "123",
                "Nro recibo": "456",
                "Vendedor": "206 - Andres Dominguez",
                "Detalle movimiento": "143820000000",
                "Ranking": 1,
            }
        ]
    }
    export_filled_bank_excel(str(original), result, str(out))

    wb = openpyxl.load_workbook(str(out), data_only=False)
    ws = wb["MercadoPago "]

    op_cell = ws.cell(2, 4)
    cuit_cell = ws.cell(2, 9)
    headers = [cell.value for cell in ws[1]]
    vendedor_col = headers.index("vendedor/fletero") + 1
    assert op_cell.data_type == "s"
    assert "E+" not in str(op_cell.value or "").upper()
    assert str(op_cell.value) == "143820000000"
    assert (cuit_cell.value or "") == ""
    assert ws.cell(2, vendedor_col).value == "206 - Andres Dominguez"


def test_export_vendor_does_not_reuse_unnamed_column_with_historical_data(tmp_path):
    original = tmp_path / "original_galicia.xlsx"
    out = tmp_path / "out_galicia.xlsx"
    _build_workbook(str(original))

    wb = openpyxl.load_workbook(str(original))
    ws = wb["SALICE GALICIA (ALARCON)"]
    ws.cell(2, 9, "dato histórico")
    wb.save(str(original))

    result = {
        "validados": [
            {
                "Origen": "GALICIA",
                "Fila Excel": 2,
                "Nro cliente": "123",
                "Nro recibo": "456",
                "Vendedor": "203 - Edgardo Larrea",
                "Ranking": 1,
            }
        ]
    }
    export_filled_bank_excel(str(original), result, str(out))

    wb2 = openpyxl.load_workbook(str(out), data_only=False)
    ws2 = wb2["SALICE GALICIA (ALARCON)"]
    assert ws2.cell(2, 9).value == "dato histórico"
    headers = [ws2.cell(1, c).value for c in range(1, ws2.max_column + 1)]
    vendedor_col = headers.index("vendedor/fletero") + 1
    assert vendedor_col != 9
    assert ws2.cell(2, vendedor_col).value == "203 - Edgardo Larrea"


def test_export_dudosos_compacts_and_keeps_only_dudoso_rows(tmp_path):
    original = tmp_path / "original_compact.xlsx"
    out = tmp_path / "out_compact.xlsx"
    _build_workbook(str(original))

    wb = openpyxl.load_workbook(str(original))
    ws = wb["MercadoPago "]
    ws.append(["21/02/2026 00:00:00", "Cobro", "", "145000000001", 1100.0, "", "", "", "20111111111"])
    ws.append(["22/02/2026 00:00:00", "Cobro", "", "145000000002", 1200.0, "", "", "", "20222222222"])
    ws.append(["23/02/2026 00:00:00", "Cobro", "", "145000000003", 1300.0, "", "", "", "20333333333"])
    wb.save(str(original))

    result = {
        "dudosos": [
            {"Origen": "MERCADOPAGO", "Fila Excel": 4, "Nro cliente": "1", "Nro recibo": "10", "Detalle movimiento": "145000000002", "Ranking": 1},
            {"Origen": "MERCADOPAGO", "Fila Excel": 5, "Nro cliente": "2", "Nro recibo": "20", "Detalle movimiento": "145000000003", "Ranking": 1},
        ]
    }

    export_filled_bank_excel(
        str(original),
        result,
        str(out),
        row_source="dudosos",
        only_ranking_1=False,
        write_ok_marker=False,
        compact_only_source_rows=True,
    )

    wb2 = openpyxl.load_workbook(str(out), data_only=False)
    ws2 = wb2["MercadoPago "]

    # Solo quedan las dos filas dudosas, compactadas arriba.
    assert ws2.cell(2, 4).value == "145000000002"
    assert ws2.cell(3, 4).value == "145000000003"
    assert (ws2.cell(4, 4).value or "") == ""


def test_export_mp_clears_saved_filters_and_forces_cliente_recibo_general(tmp_path):
    original = tmp_path / "original_filters.xlsx"
    out = tmp_path / "out_filters.xlsx"
    _build_workbook(str(original))

    wb = openpyxl.load_workbook(str(original))
    ws = wb["MercadoPago "]
    ws.auto_filter.ref = "A1:I2"
    ws.auto_filter.add_filter_column(4, ["1,4382E+11"])
    ws.row_dimensions[2].hidden = True
    ws.cell(2, 7).number_format = "_($* #,##0.00_)"
    ws.cell(2, 8).number_format = "_($* #,##0.00_)"
    wb.save(str(original))

    result = {
        "validados": [
            {
                "Origen": "MERCADOPAGO",
                "Fila Excel": 2,
                "Nro cliente": "123",
                "Nro recibo": "456",
                "Detalle movimiento": "143820000000",
                "Ranking": 1,
            }
        ]
    }
    export_filled_bank_excel(str(original), result, str(out))

    wb2 = openpyxl.load_workbook(str(out), data_only=False)
    ws2 = wb2["MercadoPago "]
    assert ws2.cell(2, 7).number_format == "General"
    assert ws2.cell(2, 8).number_format == "General"
    assert not bool(ws2.row_dimensions[2].hidden)
    assert len(list(ws2.auto_filter.filterColumn)) == 0

    # Verificación XML directa: existe autoFilter pero sin criterios activos.
    with zipfile.ZipFile(str(out), "r") as z:
        wb_xml = ET.fromstring(z.read("xl/workbook.xml"))
        rels_xml = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        ns_rel = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        ns_pkgrel = "http://schemas.openxmlformats.org/package/2006/relationships"
        rid = None
        for sh in wb_xml.find(f"{{{ns_main}}}sheets").findall(f"{{{ns_main}}}sheet"):
            if sh.attrib.get("name") == "MercadoPago ":
                rid = sh.attrib.get(f"{{{ns_rel}}}id")
                break
        assert rid
        target = None
        for rel in rels_xml.findall(f"{{{ns_pkgrel}}}Relationship"):
            if rel.attrib.get("Id") == rid:
                target = rel.attrib.get("Target")
                break
        assert target
        target = target.lstrip("/")
        if not target.startswith("xl/"):
            target = "xl/" + target
        sheet_xml = ET.fromstring(z.read(target))
        af = sheet_xml.find(f"{{{ns_main}}}autoFilter")
        assert af is not None
        assert len(list(af)) == 0
