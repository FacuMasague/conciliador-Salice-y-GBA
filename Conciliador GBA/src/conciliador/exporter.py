from __future__ import annotations

import copy
import csv
import datetime as dt
import io
import os
import shutil
import tempfile
import zipfile
from typing import Dict, List

import openpyxl
from openpyxl.utils.cell import coordinate_to_tuple


AR_NUMBER_FORMAT = '#.##0,00'
AR_NUMBER_FORMAT_TRIM = '#,##0.##'


def _coerce_export_date(v: object) -> object:
    """Normalize date-like values to dd/mm/yyyy to avoid Excel US-locale flips."""
    if v is None:
        return v
    if isinstance(v, dt.datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, dt.date):
        return v.strftime("%d/%m/%Y")
    s = str(v).strip()
    if not s:
        return v
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(s[:19], fmt).strftime("%d/%m/%Y")
        except Exception:
            pass
    return v


def _stringify_long_id(value: object) -> str:
    """Return long operation/document IDs as plain text, avoiding scientific notation."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, "f").rstrip("0").rstrip(".")

    s = str(value).strip()
    if not s:
        return ""
    s_compact = s.replace(" ", "")
    if "e" in s_compact.lower():
        try:
            f = float(s_compact.replace(",", "."))
            if abs(f - round(f)) < 1e-6:
                return str(int(round(f)))
            return format(f, "f").rstrip("0").rstrip(".")
        except Exception:
            return s
    if s_compact.endswith(".0") and s_compact[:-2].isdigit():
        return s_compact[:-2]
    return s


def _strip_calc_chain(xlsx_path: str) -> None:
    """Remove calcChain from an .xlsx to avoid Excel "repair" prompts.

    Some workbooks include an outdated xl/calcChain.xml. After modifying and saving
    with openpyxl, Excel may warn that it found problems and wants to repair the file.
    Removing calcChain (and its references) is safe: Excel will rebuild it.

    Usa parsing XML real para Content_Types.xml en vez de string replace, lo que lo
    hace robusto ante distintos formatos de indentado y orden de atributos.
    """
    import re
    import xml.etree.ElementTree as ET

    try:
        with zipfile.ZipFile(xlsx_path, 'r') as zin:
            names = set(zin.namelist())
            if 'xl/calcChain.xml' not in names:
                return
            files: dict[str, bytes] = {n: zin.read(n) for n in zin.namelist()}

        files.pop('xl/calcChain.xml', None)

        # --- Content_Types.xml: eliminar el Override de calcChain con XML parsing ---
        ct = files.get('[Content_Types].xml')
        if ct:
            try:
                # Registrar namespace para preservar el prefijo en el output
                ET.register_namespace('', 'http://schemas.openxmlformats.org/package/2006/content-types')
                root = ET.fromstring(ct)
                to_remove = [
                    el for el in root
                    if el.get('PartName') == '/xl/calcChain.xml'
                ]
                for el in to_remove:
                    root.remove(el)
                files['[Content_Types].xml'] = ET.tostring(root, encoding='UTF-8', xml_declaration=True)
            except Exception:
                # Fallback: string replace (menos robusto pero nunca falla)
                s = ct.decode('utf-8', errors='ignore')
                s = s.replace(
                    '<Override PartName="/xl/calcChain.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.calcChain+xml"/>',
                    '',
                )
                files['[Content_Types].xml'] = s.encode('utf-8')

        # --- workbook.xml.rels: eliminar la Relationship de calcChain ---
        rels_name = 'xl/_rels/workbook.xml.rels'
        rels = files.get(rels_name)
        if rels:
            s = rels.decode('utf-8', errors='ignore')
            s = re.sub(r'<Relationship[^>]*Target="calcChain\.xml"[^>]*/>', '', s)
            files[rels_name] = s.encode('utf-8')

        tmp_path = xlsx_path + '.tmp'
        with zipfile.ZipFile(tmp_path, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
            for name, data in files.items():
                zout.writestr(name, data)
        os.replace(tmp_path, xlsx_path)
    except Exception:
        return


def _format_es_ar(n: float) -> str:
    """Format number in es-AR style: 1.234.567,89."""
    try:
        s = f"{float(n):,.2f}"
    except Exception:
        return str(n)
    # Python uses ',' as thousands and '.' as decimal by default
    return s.replace(',', 'X').replace('.', ',').replace('X', '.')


def _safe_sheet_name(name: str) -> str:
    # Excel sheet names: max 31 chars, no []:*?/\\
    bad = set('[]:*?/\\')
    cleaned = ''.join('_' if c in bad else c for c in name)
    return cleaned[:31] or 'Sheet'


def _style_receipt_bank_sections(ws, cols: List[str], rows_count: int) -> None:
    """Separa visualmente datos del recibo y del ingreso sin columna artificial."""
    if not cols:
        return
    receipt_headers = {
        "Empresa", "Nro recibo", "Nro cliente", "Cliente", "Vendedor/Repartidor",
        "Medio de pago", "Fecha recibo", "Importe recibo", "CUIT recibo",
    }
    bank_headers = {
        "Origen", "Fecha movimiento", "Importe movimiento", "Detalle movimiento",
        "Fila Excel", "CUIT ingreso",
    }
    receipt_fill = openpyxl.styles.PatternFill("solid", fgColor="DCEAF7")
    bank_fill = openpyxl.styles.PatternFill("solid", fgColor="E2F0E8")
    separator = openpyxl.styles.Side(style="medium", color="5B6B7A")
    for col_idx, name in enumerate(cols, start=1):
        header = ws.cell(1, col_idx)
        if name in receipt_headers:
            header.fill = receipt_fill
        elif name in bank_headers:
            header.fill = bank_fill
        if name == "Origen":
            for row_idx in range(1, rows_count + 2):
                cell = ws.cell(row_idx, col_idx)
                cell.border = copy.copy(cell.border)
                cell.border = openpyxl.styles.Border(
                    left=separator,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom,
                    diagonal=cell.border.diagonal,
                    diagonal_direction=cell.border.diagonal_direction,
                    diagonalUp=cell.border.diagonalUp,
                    diagonalDown=cell.border.diagonalDown,
                    outline=cell.border.outline,
                    vertical=cell.border.vertical,
                    horizontal=cell.border.horizontal,
                )


def export_xlsx(result: Dict[str, List[dict]], out_path: str) -> str:
    """Write an .xlsx with 3 sheets: Validados, Dudosos, No encontrados."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    def add_sheet(title: str, rows: List[dict], *, column_order: List[str] | None = None, money_cols: List[str] | None = None):
        ws = wb.create_sheet(_safe_sheet_name(title))
        if not rows:
            ws.append(['(sin filas)'])
            return
        if column_order:
            cols = [c for c in column_order if any(c in r for r in rows)]
            # append any extra keys at end (shouldn't happen in practice)
            extras = []
            seen = set(cols)
            for r in rows:
                for k in r.keys():
                    if str(k).startswith("__"):
                        continue
                    if k == "Divisor":
                        continue
                    if k not in seen:
                        seen.add(k)
                        extras.append(k)
            cols += extras
        else:
            cols = []
            seen = set()
            for r in rows:
                for k in r.keys():
                    if k == "Divisor":
                        continue
                    if k not in seen:
                        seen.add(k)
                        cols.append(k)
        ws.append(cols)
        for r in rows:
            vals = []
            for c in cols:
                value = r.get(c, None)
                if c in {"Fecha movimiento", "Fecha recibo"}:
                    value = _coerce_export_date(value)
                vals.append(value)
            ws.append(vals)

        # Apply number formatting for money columns if present
        if money_cols:
            for col_idx, col_name in enumerate(cols, start=1):
                if col_name not in money_cols:
                    continue
                # Skip header row -> start at row 2
                for row_idx in range(2, 2 + len(rows)):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    # only format numeric cells
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = AR_NUMBER_FORMAT
        _style_receipt_bank_sections(ws, cols, len(rows))

    # V2.0:
    # - Validados: sin columna Motivo
    # - Dudosos: Motivo por código
    common_cols = [
        'Tipo fila',
        'Ranking',
        'Nro recibo',
        'Nro cliente',
        'Cliente',
        'Vendedor/Repartidor',
        'Medio de pago',
        'Fecha recibo',
        'Importe recibo',
        'Origen',
        'Fecha movimiento',
        'Importe movimiento',
        'Detalle movimiento',
        'Fila Excel',
        'Dif días',
        'Dif importe',
        'Peso',
        'Motivo',
            ]
    ne_cols = [
        'Tipo no encontrado',
        'Nro recibo',
        'Nro cliente',
        'Cliente',
        'Vendedor/Repartidor',
        'Medio de pago',
        'Fecha recibo',
        'Importe recibo',
        'Peso',
        'Origen',
        'Fecha movimiento',
        'Importe movimiento',
        'Detalle movimiento',
        'Fila Excel',
    ]

    val_cols = [c for c in common_cols if c != 'Motivo']
    add_sheet('Validados', result.get('validados', []), column_order=val_cols, money_cols=['Importe recibo', 'Importe movimiento', 'Dif importe', 'Peso'])
    add_sheet('Dudosos', result.get('dudosos', []), column_order=common_cols, money_cols=['Importe recibo', 'Importe movimiento', 'Dif importe', 'Peso'])
    add_sheet('No encontrados', result.get('no_encontrados', []), column_order=ne_cols, money_cols=['Importe recibo', 'Importe movimiento', 'Peso'])

    # Optional meta sheet
    meta = result.get('meta')
    if isinstance(meta, dict) and meta:
        ws = wb.create_sheet(_safe_sheet_name('Meta'))
        ws.append(['key', 'value'])
        for k, v in meta.items():
            ws.append([k, str(v)])

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    # Help Excel avoid "repair" prompts by forcing full recalculation
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.calcMode = 'auto'
    except Exception:
        pass
    wb.save(out_path)
    _strip_calc_chain(out_path)
    return out_path


def export_no_encontrados_xlsx(result: Dict[str, List[dict]], out_path: str) -> str:
    """Write an .xlsx for no_encontrados split by operational review bucket."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    rows = result.get("no_encontrados", []) or []
    rec_rows = [r for r in rows if str(r.get("Tipo no encontrado", "")).upper() == "RECIBO_SIN_BANCO"]
    bank_rows = [r for r in rows if str(r.get("Tipo no encontrado", "")).upper() == "BANCO_SIN_RECIBO"]

    def _origin_bucket(origen: object) -> str:
        o = str(origen or "").strip().upper()
        if o == "BBVA":
            return "BBVA"
        if o == "GALICIA":
            return "Galicia"
        if o == "MERCADOPAGO":
            return "Mercado Pago"
        return ""

    bank_buckets: dict[str, List[dict]] = {"BBVA": [], "Mercado Pago": []}
    for r in bank_rows:
        b = _origin_bucket(r.get("Origen"))
        if b == "Mercado Pago":
            bank_buckets["Mercado Pago"].append(r)
        elif b in {"BBVA", "Galicia"}:
            bank_buckets["BBVA"].append(r)

    def _non_empty(v: object) -> bool:
        if v is None:
            return False
        if isinstance(v, str):
            return v.strip() != ""
        return True

    def _coerce_money(v: object) -> object:
        """Try to convert money-like strings into numbers for stable Excel rendering."""
        if v is None or isinstance(v, (int, float)):
            return round(float(v), 2) if isinstance(v, (int, float)) else v
        if not isinstance(v, str):
            return v
        s = v.strip()
        if not s:
            return v
        # Handles variants like "85.500,00", "85500,00", "85500.00", "$ 85.500,00"
        s = s.replace("$", "").replace(" ", "")
        has_dot = "." in s
        has_comma = "," in s
        try:
            if has_dot and has_comma:
                s = s.replace(".", "").replace(",", ".")
            elif has_comma:
                s = s.replace(",", ".")
            return float(s)
        except Exception:
            return v

    def _format_es_ar_trim(n: float) -> str:
        s = _format_es_ar(float(n))
        if "," in s:
            whole, dec = s.split(",", 1)
            dec = dec.rstrip("0")
            return whole if not dec else f"{whole},{dec}"
        return s

    def _estimate_text_len(value: object, col_name: str) -> int:
        if value is None:
            return 0
        if isinstance(value, (int, float)):
            if col_name in {"Importe movimiento", "Importe recibo", "Peso"}:
                return len(_format_es_ar_trim(float(value)))
            return len(str(value))
        return len(str(value))

    def _autosize_sheet(ws, cols: List[str], rows_count: int) -> None:
        min_width_by_col = {
            "Tipo no encontrado": 22,
            "Origen": 14,
            "Fecha movimiento": 16,
            "Fecha recibo": 14,
            "Importe movimiento": 18,
            "Importe recibo": 18,
            "Detalle movimiento": 48,
            "Vendedor/Repartidor": 34,
            "Fila Excel": 12,
            "Nro recibo": 12,
            "Nro cliente": 12,
            "Medio de pago": 18,
            "CUIT ingreso": 16,
            "CUIT recibo": 16,
            "Peso": 14,
        }
        for col_idx, col_name in enumerate(cols, start=1):
            max_len = len(str(col_name))
            for row_idx in range(2, 2 + rows_count):
                val = ws.cell(row=row_idx, column=col_idx).value
                max_len = max(max_len, _estimate_text_len(val, col_name))
            width = max(max_len + 2, min_width_by_col.get(col_name, 12))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(width, 60)

        ws.row_dimensions[1].height = 24
        for row_idx in range(2, 2 + rows_count):
            ws.row_dimensions[row_idx].height = 20

    def _style_sheet(ws, rows_count: int, cols_count: int) -> None:
        if rows_count <= 0 or cols_count <= 0:
            return
        header_fill = openpyxl.styles.PatternFill("solid", fgColor="E8EEF5")
        header_font = openpyxl.styles.Font(bold=True, color="22313F")
        border = openpyxl.styles.Border(
            bottom=openpyxl.styles.Side(style="thin", color="D5DEE8")
        )
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2, max_row=1 + rows_count, max_col=cols_count):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(vertical="center", wrap_text=False)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(cols_count)}{1 + rows_count}"

    def _add_sheet(title: str, sheet_rows: List[dict], preferred_cols: List[str], money_cols: List[str]) -> None:
        ws = wb.create_sheet(_safe_sheet_name(title))
        if not sheet_rows:
            ws.append(["(sin filas)"])
            return

        cols = [c for c in preferred_cols if any(_non_empty(r.get(c)) for r in sheet_rows)]
        extras = []
        seen = set(cols)
        for r in sheet_rows:
            for k in r.keys():
                if str(k).startswith("__"):
                    continue
                if k == "Divisor":
                    continue
                if k in seen:
                    continue
                if any(_non_empty(x.get(k)) for x in sheet_rows):
                    seen.add(k)
                    extras.append(k)
        cols += extras

        ws.append(cols)
        for r in sheet_rows:
            row_values = []
            for c in cols:
                value = r.get(c, None)
                if c in {"Fecha movimiento", "Fecha recibo"}:
                    value = _coerce_export_date(value)
                if c in money_cols:
                    value = _coerce_money(value)
                row_values.append(value)
            ws.append(row_values)

        for col_idx, col_name in enumerate(cols, start=1):
            if col_name not in money_cols:
                continue
            for row_idx in range(2, 2 + len(sheet_rows)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = AR_NUMBER_FORMAT_TRIM

        _autosize_sheet(ws, cols, len(sheet_rows))
        _style_sheet(ws, len(sheet_rows), len(cols))

    bank_cols = [
        "Tipo no encontrado",
        "Origen",
        "Fecha movimiento",
        "Importe movimiento",
        "Detalle movimiento",
        "Fila Excel",
        "CUIT ingreso",
    ]
    rec_cols = [
        "Tipo no encontrado",
        "Empresa",
        "Nro recibo",
        "Nro cliente",
        "Cliente",
        "Vendedor/Repartidor",
        "Medio de pago",
        "Fecha recibo",
        "Importe recibo",
        "CUIT recibo",
        "Peso",
    ]

    def _rec_sort_key(row: dict) -> tuple:
        fecha = str(row.get("Fecha recibo", "") or "")
        try:
            fecha_key = dt.datetime.strptime(fecha[:10], "%Y-%m-%d").date()
        except Exception:
            try:
                fecha_key = dt.datetime.strptime(fecha[:10], "%d/%m/%Y").date()
            except Exception:
                fecha_key = dt.date.max
        nro = str(row.get("Nro recibo", "") or "")
        digits = "".join(ch for ch in nro if ch.isdigit())
        nro_key = int(digits) if digits else 10**18
        return (fecha_key, nro_key, str(row.get("Nro cliente", "") or ""))

    def _bank_sort_key(row: dict) -> tuple:
        fecha = str(row.get("Fecha movimiento", "") or "")
        try:
            fecha_key = dt.datetime.strptime(fecha[:10], "%Y-%m-%d").date()
        except Exception:
            try:
                fecha_key = dt.datetime.strptime(fecha[:10], "%d/%m/%Y").date()
            except Exception:
                fecha_key = dt.date.max
        try:
            amount_key = float(str(row.get("Importe movimiento") or "0").replace(",", "."))
        except Exception:
            amount_key = 0.0
        return (fecha_key, str(row.get("Fila Excel", "") or ""), amount_key)

    rec_rows = sorted(rec_rows, key=_rec_sort_key)
    bank_buckets["Mercado Pago"] = sorted(bank_buckets["Mercado Pago"], key=_bank_sort_key)
    bank_buckets["BBVA"] = sorted(bank_buckets["BBVA"], key=_bank_sort_key)

    _add_sheet("Recibos no encontrados", rec_rows, rec_cols, ["Importe recibo", "Peso"])
    _add_sheet("Ingresos MP no encontrados", bank_buckets["Mercado Pago"], bank_cols, ["Importe movimiento", "Peso"])
    _add_sheet("Ingresos BBVA no encontrados", bank_buckets["BBVA"], bank_cols, ["Importe movimiento", "Peso"])

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass
    wb.save(out_path)
    _strip_calc_chain(out_path)
    return out_path


def export_dudosos_xlsx(result: Dict[str, List[dict]], out_path: str) -> str:
    """Exporta dudosos activos y borrados, separados por origen operativo."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    rows: List[dict] = []
    for row in result.get("dudosos", []) or []:
        r = dict(row)
        rows.append(r)
    for row in result.get("dudosos_borrados", []) or []:
        r = dict(row)
        rows.append(r)

    def _bucket(row: dict) -> str:
        origen = str(row.get("Origen", "") or "").strip().upper()
        if origen == "MERCADOPAGO":
            return "Mercado Pago"
        return "BBVA"

    buckets: dict[str, List[dict]] = {"BBVA": [], "Mercado Pago": []}
    for row in rows:
        buckets[_bucket(row)].append(row)

    def _non_empty(v: object) -> bool:
        return v is not None and str(v).strip() != ""

    def _coerce_money(v: object) -> object:
        if v is None or isinstance(v, (int, float)):
            return round(float(v), 2) if isinstance(v, (int, float)) else v
        if not isinstance(v, str):
            return v
        s = v.strip().replace("$", "").replace(" ", "")
        if not s:
            return v
        try:
            if "." in s and "," in s:
                s = s.replace(".", "").replace(",", ".")
            elif "," in s:
                s = s.replace(",", ".")
            return float(s)
        except Exception:
            return v

    preferred_cols = [
        "Tipo fila",
        "Nro recibo",
        "Nro cliente",
        "Cliente",
        "Vendedor/Repartidor",
        "Medio de pago",
        "Fecha recibo",
        "Importe recibo",
        "Origen",
        "Fecha movimiento",
        "Importe movimiento",
        "Detalle movimiento",
        "Fila Excel",
        "Dif días",
        "Dif importe",
        "Peso",
        "Motivo",
    ]
    money_cols = {"Importe recibo", "Importe movimiento", "Dif importe", "Peso"}

    def _style_sheet(ws, rows_count: int, cols_count: int) -> None:
        if rows_count <= 0 or cols_count <= 0:
            return
        header_fill = openpyxl.styles.PatternFill("solid", fgColor="E8EEF5")
        header_font = openpyxl.styles.Font(bold=True, color="22313F")
        border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style="thin", color="D5DEE8"))
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2, max_row=1 + rows_count, max_col=cols_count):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(vertical="center", wrap_text=False)
        ws.row_dimensions[1].height = 24
        for row_idx in range(2, 2 + rows_count):
            ws.row_dimensions[row_idx].height = 20
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(cols_count)}{1 + rows_count}"

    def _add_sheet(title: str, sheet_rows: List[dict]) -> None:
        ws = wb.create_sheet(_safe_sheet_name(title))
        if not sheet_rows:
            ws.append(["(sin filas)"])
            return
        force_cols: set[str] = set()
        hidden_cols = {"Estado dudoso", "Ranking", "Divisor"}
        cols = [c for c in preferred_cols if c in force_cols or any(_non_empty(r.get(c)) for r in sheet_rows)]
        seen = set(cols)
        for row in sheet_rows:
            for key in row.keys():
                if str(key).startswith("__") or key in seen or key in hidden_cols:
                    continue
                if any(_non_empty(r.get(key)) for r in sheet_rows):
                    cols.append(key)
                    seen.add(key)
        ws.append(cols)
        for row in sheet_rows:
            values = []
            for col in cols:
                value = row.get(col, "")
                if col in {"Fecha movimiento", "Fecha recibo"}:
                    value = _coerce_export_date(value)
                if col in money_cols:
                    value = _coerce_money(value)
                values.append(value)
            ws.append(values)
        min_width_by_col = {
            "Tipo fila": 13,
            "Nro recibo": 12,
            "Nro cliente": 13,
            "Cliente": 34,
            "Vendedor/Repartidor": 34,
            "Medio de pago": 18,
            "Fecha recibo": 14,
            "Importe recibo": 18,
            "Origen": 14,
            "Fecha movimiento": 16,
            "Importe movimiento": 18,
            "Detalle movimiento": 48,
            "Fila Excel": 12,
            "Dif días": 12,
            "Dif importe": 14,
            "Peso": 14,
            "Motivo": 32,
        }
        for col_idx, col in enumerate(cols, start=1):
            width = max(len(str(col)) + 2, min_width_by_col.get(col, 12))
            for row_idx in range(2, 2 + len(sheet_rows)):
                value = ws.cell(row_idx, col_idx).value
                width = max(width, min(len(str(value or "")) + 2, 60))
                if col in money_cols and isinstance(value, (int, float)):
                    ws.cell(row_idx, col_idx).number_format = AR_NUMBER_FORMAT_TRIM
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(width, 60)
        _style_sheet(ws, len(sheet_rows), len(cols))
        _style_receipt_bank_sections(ws, cols, len(sheet_rows))

    _add_sheet("Dudosos Mercado Pago", buckets["Mercado Pago"])
    _add_sheet("Dudosos BBVA", buckets["BBVA"])

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass
    wb.save(out_path)
    _strip_calc_chain(out_path)
    return out_path


def export_filled_generic_excel(
    original_excel_path: str,
    result: Dict[str, List[dict]],
    out_path: str,
    *,
    allowed_origins: set[str] | None = None,
    record_key: str | None = None,
    row_source: str = "validados",
    only_ranking_1: bool = True,
) -> str:
    wb = openpyxl.load_workbook(original_excel_path)

    def _norm(value: object) -> str:
        return str(value or "").strip().lower()

    def _find_header_row(ws) -> int:
        for r in range(1, min(ws.max_row or 1, 20) + 1):
            vals = {_norm(ws.cell(r, c).value) for c in range(1, min(ws.max_column or 1, 40) + 1)}
            if "fecha de pago" in vals or ("fecha" in vals and ("concepto" in vals or "credito" in vals or "crédito" in vals)):
                return r
        return 1

    def _find_col(ws, header_row: int, names: list[str]) -> int | None:
        wanted = {_norm(n) for n in names}
        for c in range(1, min(ws.max_column or 1, 120) + 1):
            if _norm(ws.cell(header_row, c).value) in wanted:
                return c
        return None

    def _append_header_col(ws, header_row: int, label: str) -> int:
        col = int(ws.max_column or 0) + 1
        cell = ws.cell(header_row, col, label)
        if col > 1:
            src = ws.cell(header_row, col - 1)
            if src.has_style:
                cell.font = copy.copy(src.font)
                cell.fill = copy.copy(src.fill)
                cell.border = copy.copy(src.border)
                cell.alignment = copy.copy(src.alignment)
                cell.number_format = src.number_format
        widths = {
            "ok": 8,
            "cliente": 12,
            "cliente nombre": 30,
            "recibo": 12,
            "fecha recibo": 14,
            "medio de pago": 20,
            "importe recibo": 16,
            "vendedor/fletero": 32,
        }
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = widths.get(label, 16)
        return col

    def _normalize_mp_operation_ids(ws) -> None:
        header_row = _find_header_row(ws)
        headers = {
            _norm(ws.cell(header_row, c).value): c
            for c in range(1, min(ws.max_column or 1, 120) + 1)
        }
        if "fecha de pago" not in headers:
            return
        id_cols = [
            headers.get("tipo de operacion"),
            headers.get("tipo de operación"),
            headers.get("id de operacion en mercado pago"),
            headers.get("id de operación en mercado pago"),
            headers.get("numero de operacion de mercado pago (operation_id)"),
            headers.get("número de operación de mercado pago (operation_id)"),
            headers.get("numero de operacion de mercado pago"),
            headers.get("número de operación de mercado pago"),
        ]
        for col in [c for c in id_cols if c is not None]:
            for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
                cell = ws.cell(row_idx, int(col))
                cell.value = _stringify_long_id(cell.value)
                cell.number_format = "@"

    def _ensure_cols(ws) -> dict[str, int]:
        header_row = _find_header_row(ws)
        cols = {
            "header_row": header_row,
            "ok": _find_col(ws, header_row, ["ok"]),
            "cliente": _find_col(ws, header_row, ["cliente", "nro cliente", "nro_cliente"]),
            "cliente_nombre": _find_col(ws, header_row, ["cliente nombre", "nombre cliente", "cliente_nombre"]),
            "recibo": _find_col(ws, header_row, ["recibo", "nro recibo", "nro_recibo"]),
            "fecha_recibo": _find_col(ws, header_row, ["fecha recibo", "fecha_recibo"]),
            "medio_pago": _find_col(ws, header_row, ["medio de pago", "medio_pago"]),
            "importe_recibo": _find_col(ws, header_row, ["importe recibo", "importe_recibo"]),
            "vendedor_fletero": _find_col(ws, header_row, ["vendedor/fletero", "vendedor repartidor", "vendedor/repartidor", "fletero", "vendedor"]),
        }
        if cols["ok"] is None:
            cols["ok"] = _append_header_col(ws, header_row, "ok")
        if cols["cliente"] is None:
            cols["cliente"] = _append_header_col(ws, header_row, "cliente")
        if cols["cliente_nombre"] is None:
            cols["cliente_nombre"] = _append_header_col(ws, header_row, "cliente nombre")
        if cols["recibo"] is None:
            cols["recibo"] = _append_header_col(ws, header_row, "recibo")
        if cols["fecha_recibo"] is None:
            cols["fecha_recibo"] = _append_header_col(ws, header_row, "fecha recibo")
        if cols["medio_pago"] is None:
            cols["medio_pago"] = _append_header_col(ws, header_row, "medio de pago")
        if cols["importe_recibo"] is None:
            cols["importe_recibo"] = _append_header_col(ws, header_row, "importe recibo")
        if cols["vendedor_fletero"] is None:
            cols["vendedor_fletero"] = _append_header_col(ws, header_row, "vendedor/fletero")
        widths_by_key = {
            "ok": 8,
            "cliente": 12,
            "cliente_nombre": 30,
            "recibo": 12,
            "fecha_recibo": 14,
            "medio_pago": 20,
            "importe_recibo": 16,
            "vendedor_fletero": 32,
        }
        for key, width in widths_by_key.items():
            col_idx = cols.get(key)
            if col_idx is not None:
                ws.column_dimensions[openpyxl.utils.get_column_letter(int(col_idx))].width = width
        return {k: int(v) for k, v in cols.items() if v is not None}

    def _pick_sheet(row: dict) -> str | None:
        explicit = str(row.get("__sheet_name") or "").strip()
        if explicit and explicit in wb.sheetnames:
            return explicit
        origen = str(row.get("Origen") or "").strip().upper()
        if origen == "MERCADOPAGO":
            fecha = str(row.get("Fecha movimiento") or "")
            try:
                month = dt.date.fromisoformat(fecha).month
            except Exception:
                month = None
            month_names = {
                1: "ENERO",
                2: "FEBRERO",
                3: "MARZO",
                4: "ABRIL",
                5: "MAYO",
                6: "JUNIO",
                7: "JULIO",
                8: "AGOSTO",
                9: "SEPTIEMBRE",
                10: "OCTUBRE",
                11: "NOVIEMBRE",
                12: "DICIEMBRE",
            }
            wanted = month_names.get(month or 0)
            if wanted and wanted in wb.sheetnames:
                return wanted
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row = _find_header_row(ws)
            vals = {_norm(ws.cell(header_row, c).value) for c in range(1, min(ws.max_column or 1, 40) + 1)}
            if origen == "MERCADOPAGO" and "fecha de pago" in vals:
                return sheet_name
            if origen == "BBVA" and "concepto" in vals and ("credito" in vals or "crédito" in vals):
                return sheet_name
            if origen == "GALICIA" and "fecha" in vals and ("importe" in vals or "creditos" in vals or "créditos" in vals):
                return sheet_name
        return wb.sheetnames[0] if wb.sheetnames else None

    rows = []
    for row in (result.get(row_source) or []):
        if only_ranking_1:
            try:
                if int(row.get("Ranking") or 0) != 1:
                    continue
            except Exception:
                continue
        origen = str(row.get("Origen") or "").strip().upper()
        if allowed_origins and origen not in allowed_origins:
            continue
        row_record_key = str(row.get("__record_key") or "").strip()
        if record_key and row_record_key and row_record_key != record_key:
            continue
        rows.append(row)

    touched = False
    sheet_cols_cache: dict[str, dict[str, int]] = {}

    for row in rows:
        sheet_name = _pick_sheet(row)
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        cols = sheet_cols_cache.setdefault(sheet_name, _ensure_cols(ws))
        fila_excel = int(row.get("Fila Excel") or 0)
        if fila_excel <= cols["header_row"]:
            continue
        ws.cell(fila_excel, cols["ok"], "ok")
        ws.cell(fila_excel, cols["cliente"], str(row.get("Nro cliente") or "").strip())
        ws.cell(fila_excel, cols["cliente_nombre"], str(row.get("Cliente") or "").strip())
        ws.cell(fila_excel, cols["recibo"], str(row.get("Nro recibo") or "").strip())
        ws.cell(fila_excel, cols["fecha_recibo"], _coerce_export_date(row.get("Fecha recibo")))
        ws.cell(fila_excel, cols["medio_pago"], str(row.get("Medio de pago") or "").strip())
        ws.cell(fila_excel, cols["vendedor_fletero"], str(row.get("Vendedor/Repartidor") or "").strip())
        importe_cell = ws.cell(fila_excel, cols["importe_recibo"], row.get("Importe recibo"))
        if isinstance(importe_cell.value, (int, float)):
            importe_cell.number_format = AR_NUMBER_FORMAT_TRIM
        touched = True

    for sheet_name in wb.sheetnames:
        _normalize_mp_operation_ids(wb[sheet_name])

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass
    wb.save(out_path)
    if touched:
        _strip_calc_chain(out_path)
    return out_path


def export_combined_records_excel(
    records: list[dict],
    result: Dict[str, List[dict]],
    out_path: str,
) -> str:
    """Exporta un único Excel conciliado con una hoja BBVA y una hoja Mercado Pago.

    Cada record se completa primero con `export_filled_generic_excel` para reutilizar
    la lógica de escritura existente. Luego se copian las hojas relevantes a un libro
    nuevo, preservando el formato visual de cada record original.
    """

    def _norm(value: object) -> str:
        return str(value or "").strip().lower()

    def _sheet_headers(ws) -> set[str]:
        max_rows = min(int(ws.max_row or 1), 20)
        max_cols = min(int(ws.max_column or 1), 80)
        for r in range(1, max_rows + 1):
            headers = {_norm(ws.cell(r, c).value) for c in range(1, max_cols + 1)}
            if headers:
                yield_headers = headers
                if (
                    "fecha de pago" in yield_headers
                    or ("fecha" in yield_headers and "concepto" in yield_headers)
                    or ("fecha" in yield_headers and ("importe" in yield_headers or "creditos" in yield_headers))
                ):
                    return yield_headers
        return set()

    def _rows_for_record(record_key: str, origins: set[str]) -> list[dict]:
        rows: list[dict] = []
        for row in result.get("validados") or []:
            try:
                if int(row.get("Ranking") or 0) != 1:
                    continue
            except Exception:
                continue
            origen = str(row.get("Origen") or "").strip().upper()
            if origins and origen not in origins:
                continue
            row_record_key = str(row.get("__record_key") or "").strip()
            if record_key and row_record_key and row_record_key != record_key:
                continue
            rows.append(row)
        return rows

    def _sheet_from_ingestion_meta(wb: openpyxl.Workbook, group: str) -> str | None:
        meta = result.get("meta") if isinstance(result.get("meta"), dict) else {}
        summary = meta.get("raw_ingestion_summary") if isinstance(meta, dict) else None
        if not isinstance(summary, dict):
            return None
        bank_key = "MERCADOPAGO" if group == "mp" else "BBVA"
        data = summary.get(bank_key)
        if not isinstance(data, dict):
            return None
        sheet_raw = str(data.get("sheet") or "").strip()
        if not sheet_raw:
            return None
        # Si un archivo crudo abarca más de una hoja mensual, la última hoja
        # listada es la más reciente en la práctica del ingestador.
        for sheet_name in reversed([part.strip() for part in sheet_raw.split(",") if part.strip()]):
            if sheet_name in wb.sheetnames:
                return sheet_name
        return None

    def _parse_export_sheet_date(value: object) -> dt.date | None:
        if value is None:
            return None
        if isinstance(value, dt.datetime):
            return value.date()
        if isinstance(value, dt.date):
            return value
        s = str(value).strip()
        if not s:
            return None
        try:
            return dt.datetime.fromisoformat(s.replace("Z", "+00:00")).date()
        except Exception:
            pass
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d-%m-%Y"):
            try:
                return dt.datetime.strptime(s[:19], fmt).date()
            except Exception:
                pass
        return None

    def _latest_mp_sheet(wb: openpyxl.Workbook) -> str | None:
        best: tuple[dt.date, str] | None = None
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = _sheet_headers(ws)
            if "fecha de pago" not in headers:
                continue
            fecha_col = None
            max_cols = min(int(ws.max_column or 1), 80)
            for c in range(1, max_cols + 1):
                if _norm(ws.cell(1, c).value) == "fecha de pago":
                    fecha_col = c
                    break
            if fecha_col is None:
                continue
            sheet_max: dt.date | None = None
            max_rows = int(ws.max_row or 1)
            for r in range(2, max_rows + 1):
                d = _parse_export_sheet_date(ws.cell(r, fecha_col).value)
                if d is not None and (sheet_max is None or d > sheet_max):
                    sheet_max = d
            if sheet_max is not None and (best is None or sheet_max > best[0]):
                best = (sheet_max, sheet_name)
        return best[1] if best else None

    def _pick_source_sheet(wb: openpyxl.Workbook, group: str, rows: list[dict]) -> str | None:
        meta_sheet = _sheet_from_ingestion_meta(wb, group)
        if meta_sheet:
            return meta_sheet

        preferred: list[str] = []
        for row in rows:
            sheet_name = str(row.get("__sheet_name") or "").strip()
            if sheet_name and sheet_name in wb.sheetnames and sheet_name not in preferred:
                preferred.append(sheet_name)
        for sheet_name in preferred:
            ws = wb[sheet_name]
            headers = _sheet_headers(ws)
            if group == "mp" and "fecha de pago" in headers:
                return sheet_name
            if group == "bank" and "fecha" in headers and "concepto" in headers and (
                "credito" in headers or "crédito" in headers
            ):
                return sheet_name
            if group == "bank" and "fecha" in headers and (
                "importe" in headers or "creditos" in headers or "créditos" in headers
            ) and (
                "razon social" in headers or "razón social" in headers or "cuit" in headers
            ):
                return sheet_name
        if group == "mp":
            latest = _latest_mp_sheet(wb)
            if latest:
                return latest
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = _sheet_headers(ws)
            if group == "mp" and "fecha de pago" in headers:
                return sheet_name
            if group == "bank" and "fecha" in headers and "concepto" in headers and (
                "credito" in headers or "crédito" in headers
            ):
                return sheet_name
            if group == "bank" and "fecha" in headers and (
                "importe" in headers or "creditos" in headers or "créditos" in headers
            ) and (
                "razon social" in headers or "razón social" in headers or "cuit" in headers
            ):
                return sheet_name
        return None

    def _copy_sheet(src_ws, dst_wb: openpyxl.Workbook, title: str) -> None:
        clean_title = title[:31]
        if clean_title in dst_wb.sheetnames:
            base = clean_title[:28]
            i = 2
            while f"{base} {i}" in dst_wb.sheetnames:
                i += 1
            clean_title = f"{base} {i}"
        dst_ws = dst_wb.create_sheet(clean_title)

        for row in src_ws.iter_rows():
            for src_cell in row:
                dst_cell = dst_ws.cell(src_cell.row, src_cell.column, src_cell.value)
                if src_cell.has_style:
                    dst_cell.font = copy.copy(src_cell.font)
                    dst_cell.fill = copy.copy(src_cell.fill)
                    dst_cell.border = copy.copy(src_cell.border)
                    dst_cell.alignment = copy.copy(src_cell.alignment)
                    dst_cell.protection = copy.copy(src_cell.protection)
                if src_cell.hyperlink:
                    dst_cell._hyperlink = copy.copy(src_cell.hyperlink)
                if src_cell.comment:
                    dst_cell.comment = copy.copy(src_cell.comment)
                dst_cell.number_format = src_cell.number_format

        for key, dim in src_ws.column_dimensions.items():
            dst_dim = dst_ws.column_dimensions[key]
            dst_dim.width = dim.width
            dst_dim.hidden = dim.hidden
            dst_dim.outlineLevel = dim.outlineLevel
            dst_dim.collapsed = dim.collapsed

        for key, dim in src_ws.row_dimensions.items():
            dst_dim = dst_ws.row_dimensions[key]
            dst_dim.height = dim.height
            dst_dim.hidden = dim.hidden
            dst_dim.outlineLevel = dim.outlineLevel
            dst_dim.collapsed = dim.collapsed

        for merged in src_ws.merged_cells.ranges:
            dst_ws.merge_cells(str(merged))

        def _safe_freeze_panes(value: object) -> str | None:
            if not value:
                return "A2"
            try:
                row, col = coordinate_to_tuple(str(value))
            except Exception:
                return "A2"
            if row <= 5 and col <= 3:
                return str(value)
            return "A2"

        # Algunos records quedan guardados con paneles congelados en filas muy bajas
        # (ej. A6165). Copiar eso bloquea el desplazamiento normal en Excel.
        dst_ws.freeze_panes = _safe_freeze_panes(src_ws.freeze_panes)
        dst_ws.auto_filter.ref = src_ws.auto_filter.ref
        dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
        if dst_ws.sheet_view.selection:
            selection = dst_ws.sheet_view.selection[0]
            selection.activeCell = "A1"
            selection.sqref = "A1"
            selection.pane = "bottomLeft" if dst_ws.freeze_panes else None
        dst_ws.sheet_format = copy.copy(src_ws.sheet_format)
        dst_ws.sheet_properties = copy.copy(src_ws.sheet_properties)
        dst_ws.page_margins = copy.copy(src_ws.page_margins)
        dst_ws.page_setup = copy.copy(src_ws.page_setup)
        dst_ws.print_options = copy.copy(src_ws.print_options)

    tmp_dir = tempfile.mkdtemp(prefix="conciliador_combined_export_")
    try:
        filled_books: list[tuple[dict, str]] = []
        for idx, rec in enumerate(records):
            origins = set(str(o).strip().upper() for o in (rec.get("origins") or []) if str(o).strip())
            record_key = str(rec.get("key") or "")
            filled_path = os.path.join(tmp_dir, f"record_{idx + 1}_{record_key or 'default'}_filled.xlsx")
            export_filled_generic_excel(
                str(rec.get("working_excel_path") or ""),
                result,
                filled_path,
                allowed_origins=origins or None,
                record_key=record_key,
            )
            filled_books.append((rec, filled_path))

        out_wb = openpyxl.Workbook()
        default_sheet = out_wb.active
        out_wb.remove(default_sheet)

        copied_bank = False
        copied_mp = False
        for rec, filled_path in filled_books:
            origins = set(str(o).strip().upper() for o in (rec.get("origins") or []) if str(o).strip())
            record_key = str(rec.get("key") or "")
            wb = openpyxl.load_workbook(filled_path)
            try:
                rows = _rows_for_record(record_key, origins)
                if not copied_bank and {"BBVA", "GALICIA"} & origins:
                    sheet_name = _pick_source_sheet(wb, "bank", rows)
                    if sheet_name:
                        _copy_sheet(wb[sheet_name], out_wb, "BBVA")
                        copied_bank = True
                if not copied_mp and "MERCADOPAGO" in origins:
                    sheet_name = _pick_source_sheet(wb, "mp", rows)
                    if sheet_name:
                        _copy_sheet(wb[sheet_name], out_wb, "Mercado Pago")
                        copied_mp = True
            finally:
                try:
                    wb.close()
                except Exception:
                    pass

        if not out_wb.sheetnames:
            raise ValueError("No pude encontrar hojas compatibles para exportar records conciliados.")

        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        try:
            out_wb.calculation.fullCalcOnLoad = True
            out_wb.calculation.calcMode = "auto"
        except Exception:
            pass
        out_wb.save(out_path)
        _strip_calc_chain(out_path)
        return out_path
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def export_filled_bank_excel(
    original_excel_path: str,
    result: Dict[str, List[dict]],
    out_path: str,
    *,
    default_empresa: str | None = None,
    row_source: str = "validados",
    only_ranking_1: bool = True,
    write_cliente_nombre_col: bool = False,
    clear_existing_assignments: bool = False,
    write_ok_marker: bool = True,
    compact_only_source_rows: bool = False,
) -> str:
    """Devuelve el MISMO Excel de ingresos subido por el usuario, pero completado con VALIDADOS.

    Reglas:
      - Completa/sobrescribe columnas (ok/cliente/recibo) para cada validado (Ranking=1).
      - Cliente y Recibo se escriben como NÚMERO cuando sea posible (formato General).
      - Se intenta mantener el libro idéntico al original (gráficos, hojas, formatos).
        Para eso, NO re-guardamos el workbook completo con openpyxl (puede romper libros complejos).
        En su lugar, hacemos un "patch" mínimo sobre los XML de las hojas afectadas dentro del .xlsx.

    Nota:
      - default_empresa se usa cuando el modo es "simple" (1 PDF) y el resultado no incluye Empresa.
        En ese caso, se asume esa empresa para todos los validados (solo para decidir cruces BBVA).
    """
    import os
    import shutil
    import unicodedata
    import zipfile
    import re

    # IMPORTANT:
    # Prefer lxml (if available) to preserve namespace prefixes and reduce Excel "repair" prompts
    # on workbooks that contain drawings/named ranges. If lxml isn't installed, fall back to the
    # stdlib XML parser (this may be more likely to trigger Excel repair on some workbooks, but
    # it will still produce a valid file and keeps installation simple).
    try:
        import lxml.etree as ET  # type: ignore
    except ModuleNotFoundError:
        import xml.etree.ElementTree as ET
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    shutil.copy2(original_excel_path, out_path)

    # --- Helpers de normalización ---
    def _norm(s) -> str:
        if s is None:
            return ""
        s = str(s).strip().lower()
        s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
        return s

    OK_KEYS = {"ok", "recibio", "recibio?", "recibio ?", "recibió", "recibió?", "recibi", "recibido", "recibido?"}
    CLIENTE_KEYS = {"cliente", "nro cliente", "nro_cliente", "nro. cliente", "nrocliente"}
    RECIBO_KEYS = {"recibo", "nro recibo", "nro_recibo", "nro. recibo", "nrorecibo"}
    MP_OPER_KEYS = {
        "operacion relacionada",
        "operación relacionada",
        "id de operacion en mercado pago",
        "id de operación en mercado pago",
    }
    MP_CUIT_KEYS = {
        "numero de identificacion del pagador",
        "número de identificación del pagador",
    }

    # --- Leer el libro SOLO para detectar columnas/valores (sin guardarlo) ---
    wb = load_workbook(original_excel_path, read_only=True, data_only=False)

    # Mapeo de origen -> hojas candidatas
    bbva_sheets = ["SALICE BBVA", " ALARCON BBVA"]
    galicia_sheets = ["SALICE GALICIA (ALARCON)"]
    mp_sheets = ["MercadoPago "]

    def _sheet_candidates(origen: str) -> list[str]:
        o = (origen or "").strip().upper()
        if o == "BBVA":
            return [s for s in bbva_sheets if s in wb.sheetnames]
        if o == "GALICIA":
            return [s for s in galicia_sheets if s in wb.sheetnames]
        if o == "MERCADOPAGO":
            return [s for s in mp_sheets if s in wb.sheetnames]
        return []

    def _find_header_and_cols(ws) -> tuple[int, int, int, int] | None:
        """Busca fila de headers y devuelve (header_row, ok_col, cliente_col, recibo_col)."""
        max_scan_rows = min(40, ws.max_row or 40)
        max_scan_cols = min(60, ws.max_column or 60)
        for r in range(1, max_scan_rows + 1):
            ok_col = cli_col = rec_col = None
            for c in range(1, max_scan_cols + 1):
                v = ws.cell(r, c).value
                key = _norm(v)
                if key in OK_KEYS and ok_col is None:
                    ok_col = c
                if key in CLIENTE_KEYS and cli_col is None:
                    cli_col = c
                if key in RECIBO_KEYS and rec_col is None:
                    rec_col = c
            if ok_col and cli_col and rec_col:
                return (r, ok_col, cli_col, rec_col)
        return None

    def _find_col_in_header(ws, header_row: int, keyset: set[str]) -> int | None:
        wanted = {_norm(k) for k in keyset}
        max_scan_cols = min(120, ws.max_column or 120)
        for c in range(1, max_scan_cols + 1):
            if _norm(ws.cell(header_row, c).value) in wanted:
                return c
        return None

    # Cache por hoja
    cols_by_sheet: dict[str, tuple[int,int,int,int]] = {}
    compact_row_map_by_sheet: dict[str, dict[int, int]] = {}
    mp_oper_col_by_sheet: dict[str, int | None] = {}
    mp_cuit_col_by_sheet: dict[str, int | None] = {}

    def _get_cols(sheet_name: str):
        if sheet_name in cols_by_sheet:
            return cols_by_sheet[sheet_name]
        ws = wb[sheet_name]
        info = _find_header_and_cols(ws)
        if info is None:
            raise ValueError(f"No pude encontrar columnas ok/cliente/recibo en la hoja '{sheet_name}'.")
        cols_by_sheet[sheet_name] = info
        return info

    def _parse_intish(x):
        if x is None:
            return None
        s = str(x).strip()
        if s == "":
            return None
        # Quitar separadores típicos
        s2 = s.replace(".", "").replace(",", "").replace(" ", "")
        if s2.isdigit():
            try:
                return int(s2)
            except Exception:
                return None
        return None

    def _choose_bbva_sheet(row: dict) -> str | None:
        """Intenta decidir entre SALICE BBVA y  ALARCON BBVA usando el contenido de la fila."""
        cands = _sheet_candidates("BBVA")
        if not cands:
            return None
        fila = int(row.get("Fila Excel") or 0)
        # Si solo hay una, listo
        if len(cands) == 1:
            return cands[0]

        # Heurística: comparar fecha e importe contra la fila de la hoja
        fecha_str = str(row.get("Fecha movimiento") or "").strip()
        imp = row.get("Importe movimiento")
        try:
            imp = float(imp)
        except Exception:
            imp = None

        for s in cands:
            ws = wb[s]
            if fila <= 0 or fila > (ws.max_row or 0):
                continue
            # detectar columna fecha e importe en BBVA: suelen ser col 1 y col 5 según el layout visto
            v_fecha = ws.cell(fila, 1).value
            v_imp = ws.cell(fila, 5).value
            # comparar fecha "YYYY-MM-DD"
            ok_fecha = False
            try:
                if hasattr(v_fecha, "date"):
                    ok_fecha = (v_fecha.date().isoformat() == fecha_str)
                else:
                    ok_fecha = (str(v_fecha).strip() == fecha_str)
            except Exception:
                ok_fecha = False
            ok_imp = False
            try:
                if imp is not None:
                    ok_imp = abs(float(v_imp) - imp) < 0.01
            except Exception:
                ok_imp = False
            if ok_fecha or ok_imp:
                return s

        # Fallback: si hay default_empresa, usamos eso para sesgar
        if (default_empresa or "").upper().strip() == "SALICE":
            return "SALICE BBVA" if "SALICE BBVA" in cands else cands[0]
        if (default_empresa or "").upper().strip() == "ALARCON":
            return " ALARCON BBVA" if " ALARCON BBVA" in cands else cands[0]
        return cands[0]

    # --- Construir lista de escrituras ---
    writes: dict[str, list[tuple[int,int,str|int|None,str]]] = {}
    # writes[sheet] = [(row, col, value, type)], type in {"inlineStr","n"}
    general_targets: dict[str, set[tuple[int, int]]] = {}

    source_rows = result.get(row_source) or []
    principal_by_case: dict[str, dict] = {}
    for r in source_rows:
        cid = str(r.get("__case_id", "") or "")
        if cid and str(r.get("Tipo fila", "")) == "PRINCIPAL":
            principal_by_case[cid] = r

    def _effective_row(r: dict) -> dict:
        cid = str(r.get("__case_id", "") or "")
        p = principal_by_case.get(cid)
        if not p:
            return r
        out = dict(r)
        for k in ("Nro recibo", "Nro cliente", "Cliente", "Medio de pago", "Fecha recibo", "Importe recibo"):
            v = out.get(k, None)
            if v is None or str(v).strip() == "":
                out[k] = p.get(k, "")
        return out

    cliente_nombre_col_by_sheet: dict[str, int] = {}
    selected_rows_by_sheet: dict[str, set[int]] = {}

    def _get_cliente_nombre_col(sheet: str, header_row: int, rec_col: int) -> int | None:
        if not write_cliente_nombre_col:
            return None
        if sheet in cliente_nombre_col_by_sheet:
            return cliente_nombre_col_by_sheet[sheet]

        ws_hdr = wb[sheet]
        rec_col_candidate = rec_col + 1
        max_scan_cols = min(120, ws_hdr.max_column or 120)

        # Si ya existe una columna de "cliente nombre", la usamos.
        for c in range(1, max_scan_cols + 1):
            hv = _norm(ws_hdr.cell(header_row, c).value)
            if hv in {"cliente nombre", "nombre cliente", "cliente_nombre"}:
                cliente_nombre_col_by_sheet[sheet] = c
                return c

        # Preferimos la columna inmediata a la derecha de recibo si está libre.
        hv = _norm(ws_hdr.cell(header_row, rec_col_candidate).value)
        if hv == "" or hv in {"cliente nombre", "nombre cliente", "cliente_nombre"}:
            cliente_nombre_col_by_sheet[sheet] = rec_col_candidate
            writes.setdefault(sheet, []).append((header_row, rec_col_candidate, "cliente nombre", "inlineStr"))
            return rec_col_candidate

        # Buscar primera libre hacia la derecha.
        for c in range(rec_col_candidate + 1, max_scan_cols + 20):
            hv2 = _norm(ws_hdr.cell(header_row, c).value)
            if hv2 == "":
                cliente_nombre_col_by_sheet[sheet] = c
                writes.setdefault(sheet, []).append((header_row, c, "cliente nombre", "inlineStr"))
                return c
        return None

    if clear_existing_assignments:
        candidate_sheets = [s for s in (bbva_sheets + galicia_sheets + mp_sheets) if s in wb.sheetnames]
        for sheet in candidate_sheets:
            try:
                header_row, ok_col, cli_col, rec_col = _get_cols(sheet)
            except Exception:
                continue
            ws = wb[sheet]
            max_r = ws.max_row or header_row
            cliente_col = _get_cliente_nombre_col(sheet, header_row, rec_col)
            for rr in range(header_row + 1, max_r + 1):
                writes.setdefault(sheet, []).append((rr, ok_col, "", "inlineStr"))
                writes[sheet].append((rr, cli_col, "", "inlineStr"))
                writes[sheet].append((rr, rec_col, "", "inlineStr"))
                if cliente_col is not None:
                    writes[sheet].append((rr, cliente_col, "", "inlineStr"))

    for raw in source_rows:
        r = _effective_row(raw)
        if only_ranking_1:
            try:
                if int(r.get("Ranking") or 0) != 1:
                    continue
            except Exception:
                continue

        origen = str(r.get("Origen") or "").strip().upper()
        fila_excel = int(r.get("Fila Excel") or 0)
        if fila_excel <= 0:
            continue

        # elegir hoja
        if origen == "BBVA":
            sheet = _choose_bbva_sheet(r)
        else:
            cands = _sheet_candidates(origen)
            sheet = cands[0] if cands else None
        if not sheet or sheet not in wb.sheetnames:
            continue

        header_row, ok_col, cli_col, rec_col = _get_cols(sheet)
        if fila_excel <= header_row:
            continue  # no tocar headers
        writes.setdefault(sheet, [])
        if compact_only_source_rows:
            selected_rows_by_sheet.setdefault(sheet, set()).add(int(fila_excel))

        cliente_nombre_col = _get_cliente_nombre_col(sheet, header_row, rec_col)
        ws_hdr = wb[sheet]
        if sheet not in mp_oper_col_by_sheet:
            mp_oper_col_by_sheet[sheet] = _find_col_in_header(ws_hdr, header_row, MP_OPER_KEYS)
        if sheet not in mp_cuit_col_by_sheet:
            mp_cuit_col_by_sheet[sheet] = _find_col_in_header(ws_hdr, header_row, MP_CUIT_KEYS)

        cli_num = _parse_intish(r.get("Nro cliente"))
        rec_num = _parse_intish(r.get("Nro recibo"))

        # OK como texto (opcional para export de dudosos)
        if write_ok_marker:
            writes.setdefault(sheet, []).append((fila_excel, ok_col, "ok", "inlineStr"))
        # Cliente/recibo deben forzarse a estilo General, incluso si salen como texto.
        general_targets.setdefault(sheet, set()).add((fila_excel, cli_col))
        general_targets.setdefault(sheet, set()).add((fila_excel, rec_col))
        # MercadoPago:
        # - Cliente/recibo como texto para evitar "########" por ancho/estilo en Excel.
        # - Operacion Relacionada se normaliza aparte como texto plano.
        if origen == "MERCADOPAGO":
            writes[sheet].append((fila_excel, cli_col, str(r.get("Nro cliente") or "").strip(), "inlineStr"))
            writes[sheet].append((fila_excel, rec_col, str(r.get("Nro recibo") or "").strip(), "inlineStr"))
        else:
            if cli_num is not None:
                writes[sheet].append((fila_excel, cli_col, cli_num, "n"))
                general_targets.setdefault(sheet, set()).add((fila_excel, cli_col))
            else:
                writes[sheet].append((fila_excel, cli_col, str(r.get("Nro cliente") or "").strip(), "inlineStr"))
            if rec_num is not None:
                writes[sheet].append((fila_excel, rec_col, rec_num, "n"))
                general_targets.setdefault(sheet, set()).add((fila_excel, rec_col))
            else:
                writes[sheet].append((fila_excel, rec_col, str(r.get("Nro recibo") or "").strip(), "inlineStr"))
        if write_cliente_nombre_col and cliente_nombre_col is not None:
            writes[sheet].append((fila_excel, cliente_nombre_col, str(r.get("Cliente") or "").strip(), "inlineStr"))

    if compact_only_source_rows:
        candidate_sheets = [s for s in (bbva_sheets + galicia_sheets + mp_sheets) if s in wb.sheetnames]
        for sheet in candidate_sheets:
            try:
                header_row, _, _, _ = _get_cols(sheet)
            except Exception:
                continue
            selected = sorted(x for x in selected_rows_by_sheet.get(sheet, set()) if int(x) > int(header_row))
            compact_row_map_by_sheet[sheet] = {old: (header_row + i + 1) for i, old in enumerate(selected)}

        # Re-mapeamos writes existentes a las filas compactadas.
        for sheet, ops in list(writes.items()):
            row_map = compact_row_map_by_sheet.get(sheet)
            if not row_map:
                continue
            remapped: list[tuple[int, int, str | int | None, str]] = []
            for (rnum, cnum, val, vtype) in ops:
                if rnum in row_map:
                    remapped.append((row_map[rnum], cnum, val, vtype))
                elif rnum <= (_get_cols(sheet)[0]):
                    remapped.append((rnum, cnum, val, vtype))
            writes[sheet] = remapped

        # Re-mapeamos también celdas objetivo de estilo General.
        for sheet, cells in list(general_targets.items()):
            row_map = compact_row_map_by_sheet.get(sheet)
            if not row_map:
                continue
            remapped_cells: set[tuple[int, int]] = set()
            header_row = _get_cols(sheet)[0]
            for (rnum, cnum) in cells:
                if rnum in row_map:
                    remapped_cells.add((row_map[rnum], cnum))
                elif rnum <= header_row:
                    remapped_cells.add((rnum, cnum))
            general_targets[sheet] = remapped_cells

    # V4.2.0: en export al cliente, ocultar la columna CUIT de Mercado Pago.
    for sheet in [s for s in mp_sheets if s in wb.sheetnames]:
        try:
            header_row, _, _, _ = _get_cols(sheet)
        except Exception:
            continue
        ws_mp = wb[sheet]
        op_col = mp_oper_col_by_sheet.get(sheet)
        if op_col is None:
            op_col = _find_col_in_header(ws_mp, header_row, MP_OPER_KEYS)
            mp_oper_col_by_sheet[sheet] = op_col
        cuit_col = mp_cuit_col_by_sheet.get(sheet)
        if cuit_col is None:
            cuit_col = _find_col_in_header(ws_mp, header_row, MP_CUIT_KEYS)
            mp_cuit_col_by_sheet[sheet] = cuit_col
        max_r = ws_mp.max_row or header_row
        row_map = compact_row_map_by_sheet.get(sheet, {}) if compact_only_source_rows else {}
        for rr in range(header_row, max_r + 1):
            mapped_rr = rr
            if compact_only_source_rows:
                if rr <= header_row:
                    mapped_rr = rr
                else:
                    mapped_rr = row_map.get(rr)
                    if mapped_rr is None:
                        continue
            if op_col and rr > header_row:
                raw_val = ws_mp.cell(rr, op_col).value
                writes.setdefault(sheet, []).append((mapped_rr, op_col, _stringify_long_id(raw_val), "inlineStr"))
            if cuit_col:
                writes.setdefault(sheet, []).append((mapped_rr, cuit_col, "", "inlineStr"))

    if not writes:
        return out_path

    # --- Resolver paths internos de hojas dentro del xlsx ---
    NS = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pkgrel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }

    def _get_sheet_paths(xlsx_path: str) -> dict[str, str]:
        with zipfile.ZipFile(xlsx_path, "r") as z:
            wb_xml = z.read("xl/workbook.xml")
            rels_xml = z.read("xl/_rels/workbook.xml.rels")
        wb_root = ET.fromstring(wb_xml)
        rels_root = ET.fromstring(rels_xml)

        rid_to_target: dict[str, str] = {}
        for rel_el in rels_root.findall("{%s}Relationship" % NS["pkgrel"]):
            rid = rel_el.attrib.get("Id")
            target = rel_el.attrib.get("Target")
            if rid and target:
                rid_to_target[rid] = target

        out: dict[str, str] = {}
        sheets_el = wb_root.find("{%s}sheets" % NS["main"])
        if sheets_el is None:
            return out
        for sh_el in sheets_el.findall("{%s}sheet" % NS["main"]):
            name = sh_el.attrib.get("name")
            rid = sh_el.attrib.get("{%s}id" % NS["rel"])
            if not name or not rid:
                continue
            target = rid_to_target.get(rid)
            if not target:
                continue
            # target suele ser "worksheets/sheetX.xml"
            target = target.lstrip("/")
            if not target.startswith("xl/"):
                target = "xl/" + target
            out[name] = target
        return out

    sheet_path_map = _get_sheet_paths(out_path)

    # --- Utilidades para editar XML de hoja ---
    def _col_to_index(col_letters: str) -> int:
        # A->1, B->2...
        n = 0
        for ch in col_letters:
            n = n * 26 + (ord(ch.upper()) - ord("A") + 1)
        return n

    def _cell_ref(col: int, row: int) -> str:
        return f"{get_column_letter(col)}{row}"

    def _ensure_cell(row_el, ref: str):
        """Devuelve el elemento <c> con r=ref. Si no existe, lo crea y lo inserta ordenado."""
        main_ns = NS["main"]
        tag_c = "{%s}c" % main_ns
        # buscar
        for c in row_el.findall(tag_c):
            if c.attrib.get("r") == ref:
                return c
        # crear
        new_c = ET.Element(tag_c, {"r": ref})
        # insertar ordenado
        def _ref_key(rstr: str) -> int:
            m = re.match(r"([A-Z]+)(\d+)", rstr)
            if not m:
                return 10**9
            return _col_to_index(m.group(1))
        inserted = False
        ref_k = _ref_key(ref)
        children = list(row_el.findall(tag_c))
        for i, c in enumerate(children):
            cr = c.attrib.get("r") or ""
            if _ref_key(cr) > ref_k:
                row_el.insert(list(row_el).index(c), new_c)
                inserted = True
                break
        if not inserted:
            row_el.append(new_c)
        return new_c

    def _set_cell_value(c_el, value, vtype: str):
        """Set cell value using inlineStr or numeric."""
        main_ns = NS["main"]
        # limpiar hijos (no tocar atributos no relacionados, salvo type/style cuando corresponda)
        for ch in list(c_el):
            c_el.remove(ch)
        if vtype == "n":
            # número
            c_el.attrib.pop("t", None)
            c_el.attrib.pop("s", None)
            v = ET.SubElement(c_el, "{%s}v" % main_ns)
            v.text = str(int(value)) if isinstance(value, int) else str(value)
        else:
            # inline string
            c_el.attrib.pop("s", None)
            c_el.attrib["t"] = "inlineStr"
            is_el = ET.SubElement(c_el, "{%s}is" % main_ns)
            t_el = ET.SubElement(is_el, "{%s}t" % main_ns)
            t_el.text = "" if value is None else str(value)

    # --- Reescribir zip reemplazando solo las hojas tocadas ---
    with zipfile.ZipFile(out_path, "r") as zin:
        files = {name: zin.read(name) for name in zin.namelist()}

    # Detectar un style id que realmente sea "General" (numFmtId=0) en ESTE workbook.
    general_style_idx: int | None = None
    try:
        styles_xml = files.get("xl/styles.xml")
        if styles_xml:
            styles_root = ET.fromstring(styles_xml)
            cellxfs = styles_root.find(".//{%s}cellXfs" % NS["main"])
            if cellxfs is not None:
                xfs = cellxfs.findall("{%s}xf" % NS["main"])
                preferred: int | None = None
                fallback: int | None = None
                for i, xf in enumerate(xfs):
                    if xf.get("numFmtId") != "0":
                        continue
                    if fallback is None:
                        fallback = i
                    # Preferimos uno "base" (xfId=0, font/fill/border en 0) para evitar efectos visuales raros.
                    if (
                        xf.get("xfId") == "0"
                        and xf.get("fontId") == "0"
                        and xf.get("fillId") == "0"
                        and xf.get("borderId") == "0"
                    ):
                        preferred = i
                        break
                general_style_idx = preferred if preferred is not None else fallback
    except Exception:
        general_style_idx = None

    for sheet_name, ops in writes.items():
        sheet_path = sheet_path_map.get(sheet_name)
        if not sheet_path or sheet_path not in files:
            continue
        data = files[sheet_path]
        root = ET.fromstring(data)

        sheetData = root.find("{%s}sheetData" % NS["main"])
        if sheetData is None:
            continue

        # Limpiar estado de filtros guardado en la hoja para evitar apertura filtrada por defecto.
        af = root.find("{%s}autoFilter" % NS["main"])
        if af is not None:
            for ch in list(af):
                af.remove(ch)
        sort_state = root.find("{%s}sortState" % NS["main"])
        if sort_state is not None:
            root.remove(sort_state)
        for row_el in sheetData.findall("{%s}row" % NS["main"]):
            row_el.attrib.pop("hidden", None)

        if compact_only_source_rows:
            row_map = compact_row_map_by_sheet.get(sheet_name, {})
            try:
                header_row, _, _, _ = _get_cols(sheet_name)
            except Exception:
                header_row = 1
            if row_map is not None:
                tag_row = "{%s}row" % NS["main"]
                tag_c = "{%s}c" % NS["main"]
                existing_rows = list(sheetData.findall(tag_row))
                by_r: dict[int, object] = {}
                for r_el in existing_rows:
                    rr = r_el.attrib.get("r")
                    if rr and rr.isdigit():
                        by_r[int(rr)] = r_el

                # Limpiar filas de datos (>header) y reinsertar solo las seleccionadas compactadas.
                for r_el in existing_rows:
                    rr = r_el.attrib.get("r")
                    if rr and rr.isdigit() and int(rr) > int(header_row):
                        sheetData.remove(r_el)

                for old_r, new_r in sorted(row_map.items(), key=lambda kv: kv[1]):
                    src_row = by_r.get(int(old_r))
                    if src_row is None:
                        continue
                    new_row = ET.fromstring(ET.tostring(src_row))
                    new_row.attrib["r"] = str(int(new_r))
                    for c_el in new_row.findall(tag_c):
                        cref = c_el.attrib.get("r", "")
                        m = re.match(r"([A-Z]+)(\d+)", cref)
                        if not m:
                            continue
                        c_el.attrib["r"] = f"{m.group(1)}{int(new_r)}"
                    sheetData.append(new_row)

        # indexar filas
        rows_by_r = {}
        for row_el in sheetData.findall("{%s}row" % NS["main"]):
            rr = row_el.attrib.get("r")
            if rr and rr.isdigit():
                rows_by_r[int(rr)] = row_el

        for (rnum, cnum, val, vtype) in ops:
            row_el = rows_by_r.get(int(rnum))
            if row_el is None:
                continue
            ref = _cell_ref(int(cnum), int(rnum))
            c_el = _ensure_cell(row_el, ref)
            _set_cell_value(c_el, val, vtype)

        # Forzar estilo General para celdas numéricas controladas por export
        # (cliente/recibo y, en MP, Operación Relacionada).
        if general_style_idx is not None:
            for (rnum, cnum) in sorted(general_targets.get(sheet_name, set())):
                row_el = rows_by_r.get(int(rnum))
                if row_el is None:
                    continue
                ref = _cell_ref(int(cnum), int(rnum))
                c_el = _ensure_cell(row_el, ref)
                c_el.attrib["s"] = str(general_style_idx)

        # serializar
        # Preserve a valid XML header similar to the original (standalone="yes").
        files[sheet_path] = ET.tostring(root, encoding="UTF-8", xml_declaration=True, standalone="yes")

    tmp = out_path + ".tmp"
    with zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for name, data in files.items():
            zout.writestr(name, data)
    os.replace(tmp, out_path)
    return out_path

def export_zip_csv(result: Dict[str, List[dict]], out_path: str) -> str:
    """Write a .zip containing 3 CSV files + meta.json."""
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with zipfile.ZipFile(out_path, 'w', compression=zipfile.ZIP_DEFLATED) as z:
        for key, fname in [
            ('validados', 'validados.csv'),
            ('dudosos', 'dudosos.csv'),
            ('no_encontrados', 'no_encontrados.csv'),
        ]:
            rows = result.get(key, []) or []
            content = _rows_to_csv_bytes(rows)
            z.writestr(fname, content)

        # meta
        meta = result.get('meta')
        if isinstance(meta, dict):
            import json

            z.writestr('meta.json', json.dumps(meta, ensure_ascii=False, indent=2))
    return out_path


def _rows_to_csv_bytes(rows: List[dict]) -> bytes:
    if not rows:
        return b''
    cols = []
    seen = set()
    for r in rows:
        for k in r.keys():
            if k not in seen:
                seen.add(k)
                cols.append(k)
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=cols, extrasaction='ignore')
    w.writeheader()
    for r in rows:
        # For readability, format numeric money fields in es-AR style when possible
        out = {}
        for c in cols:
            v = r.get(c, '')
            if c in {'Importe recibo', 'Importe movimiento', 'Dif importe', 'Peso'} and isinstance(v, (int, float)):
                out[c] = _format_es_ar(v)
            else:
                out[c] = v
        w.writerow(out)
    return buf.getvalue().encode('utf-8')
