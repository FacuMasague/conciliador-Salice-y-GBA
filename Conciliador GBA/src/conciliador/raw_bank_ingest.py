from __future__ import annotations

import datetime as dt
import os
import re
import shutil
import unicodedata
import zipfile
from collections import Counter
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Tuple

import openpyxl
import pandas as pd
import xml.etree.ElementTree as ET


RAW_RECENT_FILES_MAX_GAP_DAYS = 10


@dataclass(frozen=True)
class RawBankTxn:
    bank: str  # BBVA | GALICIA | MERCADOPAGO
    fecha: dt.date
    importe: float
    detalle: str
    cuit: Optional[str]
    razon_social: Optional[str]
    source_filename: str
    source_sheet: Optional[str] = None
    concept: Optional[str] = None
    numero_documento: Optional[str] = None
    oficina: Optional[str] = None
    secondary_detail: Optional[str] = None
    hora: Optional[dt.time] = None


def _norm(s: object) -> str:
    if s is None:
        return ""
    txt = str(s).strip().lower()
    txt = "".join(c for c in unicodedata.normalize("NFKD", txt) if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", txt)


def _norm_for_sig(s: object) -> str:
    return re.sub(r"\s+", " ", _norm(s))


def _clean_cell_text(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return ""
    return s


def _is_bbva_historical_sheet_name(sheet_name: object) -> bool:
    n = _norm(sheet_name)
    return "histor" in n


def _parse_amount(value: object) -> Optional[float]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    s = s.replace("$", "").replace(" ", "")
    if re.match(r"^-?\d{1,3}(?:\.\d{3})*,\d{1,2}$", s):
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", "")
    try:
        return float(s)
    except Exception:
        return None


def _parse_date(value: object) -> Optional[dt.date]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value

    # Excel serial date (very common in .xls/.xlsx internals)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            serial = float(value)
            if 1 <= serial <= 100000:
                base = dt.datetime(1899, 12, 30)
                return (base + dt.timedelta(days=serial)).date()
        except Exception:
            pass

    s = str(value).strip()
    if not s:
        return None

    # IMPORTANT: parse explicit formats first (dd-mm-yyyy / dd/mm/yyyy),
    # otherwise pandas can interpret ambiguous values as mm-dd-yyyy.
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(s[:10], fmt).date()
        except Exception:
            pass

    # Numeric text can still be an Excel serial.
    if re.fullmatch(r"\d+(?:\.\d+)?", s):
        try:
            serial = float(s)
            if 1 <= serial <= 100000:
                base = dt.datetime(1899, 12, 30)
                return (base + dt.timedelta(days=serial)).date()
        except Exception:
            pass

    try:
        d = pd.to_datetime(s, errors="coerce", dayfirst=True)
        if pd.notna(d):
            return d.date()
    except Exception:
        pass
    return None


def _parse_datetime(value: object) -> tuple[Optional[dt.date], Optional[dt.time]]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None, None
    if isinstance(value, dt.datetime):
        return value.date(), value.time().replace(microsecond=0)
    if isinstance(value, dt.date):
        return value, None

    s = str(value).strip()
    if not s:
        return None, None

    # ISO datetimes from providers such as Mercado Pago must be parsed before
    # generic day-first handling, otherwise values like 2026-02-01 can be
    # misread as 2026-01-02.
    try:
        iso = dt.datetime.fromisoformat(s.replace("Z", "+00:00"))
        return iso.date(), iso.timetz().replace(microsecond=0, tzinfo=None)
    except Exception:
        pass

    for fmt in ("%d/%m/%Y %H:%M:%S", "%d-%m-%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            d = dt.datetime.strptime(s, fmt)
            return d.date(), d.time()
        except Exception:
            pass

    try:
        d = pd.to_datetime(s, errors="coerce", dayfirst=True)
        if pd.notna(d):
            return d.date(), d.time().replace(microsecond=0)
    except Exception:
        pass

    d = _parse_date(s)
    return d, None


def _pick_col(columns: Iterable[object], candidates: List[str]) -> Optional[str]:
    want = {_norm(c) for c in candidates}
    for c in columns:
        if _norm(c) in want:
            return str(c)
    return None


def _cuit_checksum_ok(cuit11: str) -> bool:
    if not re.fullmatch(r"\d{11}", cuit11 or ""):
        return False
    weights = [5, 4, 3, 2, 7, 6, 5, 4, 3, 2]
    acc = sum(int(cuit11[i]) * weights[i] for i in range(10))
    mod = 11 - (acc % 11)
    if mod == 11:
        check = 0
    elif mod == 10:
        check = 9
    else:
        check = mod
    return int(cuit11[-1]) == check


def _extract_cuit(text: object) -> Optional[str]:
    if text is None:
        return None
    s = str(text)
    if not s.strip():
        return None

    explicit = re.findall(r"\b(\d{2})[-\s]?(\d{8})[-\s]?(\d)\b", s)
    for g1, g2, g3 in explicit:
        cand = f"{g1}{g2}{g3}"
        if _cuit_checksum_ok(cand):
            return cand

    digits = "".join(ch for ch in s if ch.isdigit())
    if len(digits) >= 11:
        for i in range(0, len(digits) - 10):
            cand = digits[i : i + 11]
            if _cuit_checksum_ok(cand):
                return cand

    m = re.search(r"\b(\d{11})\b", s)
    return m.group(1) if m else None


def _plain_id(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, "f").rstrip("0").rstrip(".")
    s = str(value).strip()
    if not s:
        return ""
    s_compact = s.replace(" ", "")
    if "e" in s_compact.lower():
        try:
            f = float(s_compact.replace(",", "."))
            if abs(f - round(f)) < 1e-6:
                return str(int(round(f)))
            return format(f, "f").rstrip("0").rstrip(".")
        except Exception:
            return s
    if s_compact.endswith(".0") and s_compact[:-2].isdigit():
        return s_compact[:-2]
    return s


def _read_preview(path: str, nrows: int = 16) -> pd.DataFrame:
    try:
        return pd.read_excel(path, sheet_name=0, header=None, nrows=nrows, dtype=str)
    except ImportError as e:
        if path.lower().endswith(".xls"):
            raise ValueError(
                "No se puede leer .xls porque falta dependencia 'xlrd'. Instalá xlrd>=2.0.1."
            ) from e
        raise


def _xlsx_first_sheet_headers(path: str) -> List[str]:
    """Lee headers de la primera hoja via XML (sin openpyxl), útil para xlsx con estilos rotos."""
    ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns_rel = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ns_pkgrel = "http://schemas.openxmlformats.org/package/2006/relationships"
    with zipfile.ZipFile(path, "r") as z:
        wb_root = ET.fromstring(z.read("xl/workbook.xml"))
        rels_root = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rid_to_target: dict[str, str] = {}
        for rel_el in rels_root.findall(f"{{{ns_pkgrel}}}Relationship"):
            rid = rel_el.attrib.get("Id")
            target = rel_el.attrib.get("Target")
            if rid and target:
                rid_to_target[rid] = target
        sheets_el = wb_root.find(f"{{{ns_main}}}sheets")
        if sheets_el is None:
            return []
        first_sheet = sheets_el.find(f"{{{ns_main}}}sheet")
        if first_sheet is None:
            return []
        rid = first_sheet.attrib.get(f"{{{ns_rel}}}id")
        if not rid:
            return []
        target = rid_to_target.get(rid, "").lstrip("/")
        if not target:
            return []
        if not target.startswith("xl/"):
            target = "xl/" + target
        sheet_root = ET.fromstring(z.read(target))
        sst: dict[int, str] = {}
        if "xl/sharedStrings.xml" in z.namelist():
            sroot = ET.fromstring(z.read("xl/sharedStrings.xml"))
            items: List[str] = []
            for si in sroot.findall(f"{{{ns_main}}}si"):
                txt = ""
                t = si.find(f"{{{ns_main}}}t")
                if t is not None and t.text:
                    txt = t.text
                else:
                    for r in si.findall(f"{{{ns_main}}}r"):
                        tt = r.find(f"{{{ns_main}}}t")
                        if tt is not None and tt.text:
                            txt += tt.text
                items.append(txt)
            sst = {i: v for i, v in enumerate(items)}

        row1 = sheet_root.find(f".//{{{ns_main}}}sheetData/{{{ns_main}}}row[@r='1']")
        if row1 is None:
            return []

        out: List[str] = []
        for c in row1.findall(f"{{{ns_main}}}c"):
            t = c.attrib.get("t")
            if t == "inlineStr":
                te = c.find(f"{{{ns_main}}}is/{{{ns_main}}}t")
                out.append((te.text if te is not None and te.text else "").strip())
                continue
            v = c.find(f"{{{ns_main}}}v")
            raw = v.text if (v is not None and v.text is not None) else ""
            if t == "s":
                try:
                    out.append(sst.get(int(raw), "").strip())
                except Exception:
                    out.append("")
            else:
                out.append(str(raw).strip())
        return out


def _parse_mp_raw_xml(path: str) -> List[RawBankTxn]:
    """Fallback de parseo MP vía XML para xlsx que openpyxl no puede abrir."""
    ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns_rel = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ns_pkgrel = "http://schemas.openxmlformats.org/package/2006/relationships"

    def _col_to_idx(ref: str) -> int:
        m = re.match(r"([A-Z]+)\d+", ref or "")
        if not m:
            return -1
        n = 0
        for ch in m.group(1):
            n = n * 26 + (ord(ch) - ord("A") + 1)
        return n - 1

    with zipfile.ZipFile(path, "r") as z:
        wb_root = ET.fromstring(z.read("xl/workbook.xml"))
        rels_root = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rid_to_target: dict[str, str] = {}
        for rel_el in rels_root.findall(f"{{{ns_pkgrel}}}Relationship"):
            rid = rel_el.attrib.get("Id")
            target = rel_el.attrib.get("Target")
            if rid and target:
                rid_to_target[rid] = target
        sheets_el = wb_root.find(f"{{{ns_main}}}sheets")
        if sheets_el is None:
            raise ValueError(f"Mercado Pago: no pude leer workbook en '{os.path.basename(path)}'.")
        first_sheet = sheets_el.find(f"{{{ns_main}}}sheet")
        if first_sheet is None:
            raise ValueError(f"Mercado Pago: no encontré hojas en '{os.path.basename(path)}'.")
        rid = first_sheet.attrib.get(f"{{{ns_rel}}}id")
        target = rid_to_target.get(rid or "", "").lstrip("/")
        if not target:
            raise ValueError(f"Mercado Pago: no pude resolver hoja en '{os.path.basename(path)}'.")
        if not target.startswith("xl/"):
            target = "xl/" + target

        sst: dict[int, str] = {}
        if "xl/sharedStrings.xml" in z.namelist():
            sroot = ET.fromstring(z.read("xl/sharedStrings.xml"))
            items: List[str] = []
            for si in sroot.findall(f"{{{ns_main}}}si"):
                txt = ""
                t = si.find(f"{{{ns_main}}}t")
                if t is not None and t.text:
                    txt = t.text
                else:
                    for r in si.findall(f"{{{ns_main}}}r"):
                        tt = r.find(f"{{{ns_main}}}t")
                        if tt is not None and tt.text:
                            txt += tt.text
                items.append(txt)
            sst = {i: v for i, v in enumerate(items)}

        sheet_root = ET.fromstring(z.read(target))
        rows = sheet_root.findall(f".//{{{ns_main}}}sheetData/{{{ns_main}}}row")
        if not rows:
            return []

        parsed_rows: List[List[str]] = []
        max_idx = -1
        for row in rows:
            cells = {}
            for c in row.findall(f"{{{ns_main}}}c"):
                idx = _col_to_idx(c.attrib.get("r", ""))
                if idx < 0:
                    continue
                t = c.attrib.get("t")
                val = ""
                if t == "inlineStr":
                    te = c.find(f"{{{ns_main}}}is/{{{ns_main}}}t")
                    val = te.text if te is not None and te.text else ""
                else:
                    v = c.find(f"{{{ns_main}}}v")
                    raw = v.text if (v is not None and v.text is not None) else ""
                    if t == "s":
                        try:
                            val = sst.get(int(raw), "")
                        except Exception:
                            val = ""
                    else:
                        val = str(raw)
                cells[idx] = val
                max_idx = max(max_idx, idx)
            parsed_rows.append([cells.get(i, "") for i in range(max_idx + 1)])

    if not parsed_rows:
        return []
    headers = [str(x or "").strip() for x in parsed_rows[0]]
    data_rows = parsed_rows[1:]
    out: List[RawBankTxn] = []

    def _get_idx(candidates: List[str]) -> Optional[int]:
        want = {_norm(c) for c in candidates}
        for i, h in enumerate(headers):
            if _norm(h) in want:
                return i
        return None

    detalle_idx = _get_idx(
        [
            "ID DE OPERACIÓN EN MERCADO PAGO",
            "ID DE OPERACION EN MERCADO PAGO",
            "Número de operación de Mercado Pago (operation_id)",
            "Numero de operacion de Mercado Pago (operation_id)",
            "Número de operación de Mercado Pago",
            "Numero de operacion de Mercado Pago",
        ]
    )
    importe_idx = _get_idx(["VALOR DE LA COMPRA", "Valor del producto (transaction_amount)", "Valor del producto"])
    fecha_idx = _get_idx(["FECHA DE ORIGEN", "Fecha de compra (date_created)", "Fecha de compra"])
    cuit_idx = _get_idx(
        [
            "NÚMERO DE IDENTIFICACIÓN DEL PAGADOR",
            "NUMERO DE IDENTIFICACION DEL PAGADOR",
            "Documento de la contraparte (buyer_document)",
            "Documento de la contraparte",
        ]
    )
    pagador_idx = _get_idx(["PAGADOR", "Nombre de la contraparte (counterpart_name)", "Nombre de la contraparte"])
    medio_idx = _get_idx(["MEDIO DE PAGO", "Tipo de Operación", "TIPO DE MEDIO DE PAGO"])

    if detalle_idx is None or importe_idx is None or fecha_idx is None:
        raise ValueError(f"Mercado Pago: faltan columnas esperadas en '{os.path.basename(path)}'.")

    for row in data_rows:
        fecha, hora = _parse_datetime(row[fecha_idx] if fecha_idx < len(row) else None)
        importe = _parse_amount(row[importe_idx] if importe_idx < len(row) else None)
        if not fecha or importe is None or importe <= 0:
            continue
        detalle = str(row[detalle_idx] if detalle_idx < len(row) else "").strip()
        if not detalle:
            continue
        cuit_raw = row[cuit_idx] if (cuit_idx is not None and cuit_idx < len(row)) else None
        pagador_raw = row[pagador_idx] if (pagador_idx is not None and pagador_idx < len(row)) else None
        out.append(
            RawBankTxn(
                bank="MERCADOPAGO",
                fecha=fecha,
                importe=float(importe),
                detalle=detalle,
                cuit=_extract_cuit(cuit_raw),
                razon_social=str(pagador_raw or "").strip() or None,
                source_filename=os.path.basename(path),
                secondary_detail=str(row[medio_idx] if (medio_idx is not None and medio_idx < len(row)) else "").strip() or None,
                hora=hora,
            )
        )
    return out


def detect_raw_bank_kind(path: str) -> str:
    try:
        preview = _read_preview(path, nrows=20)
    except Exception:
        preview = pd.DataFrame([_xlsx_first_sheet_headers(path)])

    # Mercado Pago: encabezado en fila 1.
    if len(preview) >= 1:
        first = {_norm(x) for x in preview.iloc[0].tolist()}
        if (
            "id de operacion en mercado pago" in first
            and "valor de la compra" in first
            and "fecha de origen" in first
        ):
            return "MERCADOPAGO"
        if (
            "numero de identificacion del pagador" in first
            and "id de operacion en mercado pago" in first
        ):
            return "MERCADOPAGO"
        if (
            "numero de operacion de mercado pago (operation_id)" in first
            and "valor del producto (transaction_amount)" in first
            and "fecha de compra (date_created)" in first
        ):
            return "MERCADOPAGO"

    # Galicia: encabezado en fila 1.
    if len(preview) >= 1:
        first = {_norm(x) for x in preview.iloc[0].tolist()}
        if "descripcion" in first and ("creditos" in first or "creditos" in " ".join(first)):
            return "GALICIA"
        if "leyendas adicionales 1" in first and "leyendas adicionales 2" in first:
            return "GALICIA"

    # BBVA: encabezado en fila 7 aprox.
    for i in range(min(len(preview), 20)):
        row = {_norm(x) for x in preview.iloc[i].tolist()}
        if "fecha" in row and "concepto" in row and ("credito" in row or "creditos" in row):
            return "BBVA"

    raise ValueError(f"No pude detectar el tipo de banco para '{os.path.basename(path)}'.")


def _parse_bbva_raw(path: str) -> List[RawBankTxn]:
    try:
        xls = pd.ExcelFile(path)
        sheet_names = list(xls.sheet_names)
    except Exception:
        sheet_names = [0]

    out: List[RawBankTxn] = []
    seen: set[tuple[str, float, str]] = set()

    for sheet in sheet_names:
        try:
            preview = pd.read_excel(path, sheet_name=sheet, header=None, nrows=40, dtype=str)
        except Exception:
            continue

        header_row = None
        for i in range(min(len(preview), 40)):
            row = {_norm(x) for x in preview.iloc[i].tolist()}
            if "fecha" in row and "concepto" in row and ("credito" in row or "creditos" in row):
                header_row = i
                break
        if header_row is None:
            continue

        try:
            df = pd.read_excel(path, sheet_name=sheet, header=header_row, dtype=str)
        except Exception:
            continue
        fecha_col = _pick_col(df.columns, ["Fecha"])
        detalle_col = _pick_col(df.columns, ["Concepto", "Descripcion", "Descripción"])
        importe_col = _pick_col(df.columns, ["Credito", "Crédito", "Creditos", "Créditos"])
        numero_doc_col = _pick_col(df.columns, ["Numero Documento", "Número Documento"])
        oficina_col = _pick_col(df.columns, ["Oficina"])
        detalle_extra_col = _pick_col(df.columns, ["Detalle"])
        if not fecha_col or not detalle_col or not importe_col:
            continue

        for _, row in df.iterrows():
            fecha = _parse_date(row.get(fecha_col))
            importe = _parse_amount(row.get(importe_col))
            if not fecha or importe is None or importe <= 0:
                continue
            concepto = _clean_cell_text(row.get(detalle_col))
            detalle_extra = _clean_cell_text(row.get(detalle_extra_col)) if detalle_extra_col else ""
            numero_doc = _clean_cell_text(row.get(numero_doc_col)) if numero_doc_col else ""
            detalle = " | ".join(x for x in [concepto, numero_doc, detalle_extra] if str(x or "").strip())
            if not detalle:
                continue
            sig = (fecha.isoformat(), round(float(importe), 2), _norm_for_sig(detalle))
            if sig in seen:
                continue
            seen.add(sig)
            out.append(
                RawBankTxn(
                    bank="BBVA",
                    fecha=fecha,
                    importe=float(importe),
                    detalle=detalle,
                    cuit=_extract_cuit(row.get(numero_doc_col)) if numero_doc_col else _extract_cuit(detalle),
                    razon_social=None,
                    source_filename=os.path.basename(path),
                    source_sheet=str(sheet),
                    concept=concepto or None,
                    numero_documento=numero_doc or None,
                    oficina=(_clean_cell_text(row.get(oficina_col)) or None) if oficina_col else None,
                    secondary_detail=detalle_extra or None,
                )
            )

    if not out:
        raise ValueError(f"BBVA: no encontré movimientos válidos en '{os.path.basename(path)}'.")
    return out


def _parse_mp_raw(path: str) -> List[RawBankTxn]:
    try:
        df = pd.read_excel(path, sheet_name=0, header=0, dtype=str)
    except Exception:
        return _parse_mp_raw_xml(path)
    detalle_col = _pick_col(
        df.columns,
        [
            "ID DE OPERACIÓN EN MERCADO PAGO",
            "ID DE OPERACION EN MERCADO PAGO",
            "Número de operación de Mercado Pago (operation_id)",
            "Numero de operacion de Mercado Pago (operation_id)",
            "Número de operación de Mercado Pago",
            "Numero de operacion de Mercado Pago",
        ],
    )
    importe_col = _pick_col(df.columns, ["VALOR DE LA COMPRA", "Valor del producto (transaction_amount)", "Valor del producto"])
    fecha_col = _pick_col(df.columns, ["FECHA DE ORIGEN", "Fecha de compra (date_created)", "Fecha de compra"])
    cuit_col = _pick_col(
        df.columns,
        [
            "NÚMERO DE IDENTIFICACIÓN DEL PAGADOR",
            "NUMERO DE IDENTIFICACION DEL PAGADOR",
            "Documento de la contraparte (buyer_document)",
            "Documento de la contraparte",
        ],
    )
    pagador_col = _pick_col(df.columns, ["PAGADOR", "Nombre de la contraparte (counterpart_name)", "Nombre de la contraparte"])
    medio_col = _pick_col(df.columns, ["MEDIO DE PAGO", "Tipo de Operación", "TIPO DE MEDIO DE PAGO"])

    if not detalle_col or not importe_col or not fecha_col:
        raise ValueError(f"Mercado Pago: faltan columnas esperadas en '{os.path.basename(path)}'.")

    out: List[RawBankTxn] = []
    for _, row in df.iterrows():
        fecha, hora = _parse_datetime(row.get(fecha_col))
        importe = _parse_amount(row.get(importe_col))
        if not fecha or importe is None or importe <= 0:
            continue
        detalle = str(row.get(detalle_col) or "").strip()
        if not detalle:
            continue
        out.append(
            RawBankTxn(
                bank="MERCADOPAGO",
                fecha=fecha,
                importe=float(importe),
                detalle=detalle,
                cuit=_extract_cuit(row.get(cuit_col)) if cuit_col else None,
                razon_social=str(row.get(pagador_col) or "").strip() or None,
                source_filename=os.path.basename(path),
                secondary_detail=(str(row.get(medio_col) or "").strip() or None) if medio_col else None,
                hora=hora,
            )
        )
    return out


def _parse_galicia_raw(path: str) -> List[RawBankTxn]:
    df = pd.read_excel(path, sheet_name=0, header=0, dtype=str)
    fecha_col = _pick_col(df.columns, ["Fecha"])
    detalle_col = _pick_col(df.columns, ["Descripción", "Descripcion"])
    importe_col = _pick_col(df.columns, ["Créditos", "Creditos"])
    razon_col = _pick_col(df.columns, ["Leyendas Adicionales 1", "Razon social", "Razón social"])
    cuit_col = _pick_col(df.columns, ["Leyendas Adicionales 2", "CUIT"])

    if not fecha_col or not detalle_col or not importe_col:
        raise ValueError(f"Galicia: faltan columnas esperadas en '{os.path.basename(path)}'.")

    out: List[RawBankTxn] = []
    for _, row in df.iterrows():
        fecha = _parse_date(row.get(fecha_col))
        importe = _parse_amount(row.get(importe_col))
        if not fecha or importe is None or importe <= 0:
            continue
        detalle = str(row.get(detalle_col) or "").strip()
        if not detalle:
            continue
        out.append(
            RawBankTxn(
                bank="GALICIA",
                fecha=fecha,
                importe=float(importe),
                detalle=detalle,
                cuit=_extract_cuit(row.get(cuit_col)) if cuit_col else None,
                razon_social=str(row.get(razon_col) or "").strip() or None,
                source_filename=os.path.basename(path),
            )
        )
    return out


def parse_raw_bank_file(path: str) -> tuple[str, List[RawBankTxn]]:
    kind = detect_raw_bank_kind(path)
    if kind == "BBVA":
        return kind, _parse_bbva_raw(path)
    if kind == "GALICIA":
        return kind, _parse_galicia_raw(path)
    if kind == "MERCADOPAGO":
        return kind, _parse_mp_raw(path)
    raise ValueError(f"Banco no soportado: {kind}")


def _find_header_row(ws: openpyxl.worksheet.worksheet.Worksheet, required: List[str], max_rows: int = 40) -> Optional[int]:
    req = {_norm(x) for x in required}
    for r in range(1, min(max_rows, ws.max_row or max_rows) + 1):
        row_vals = {_norm(ws.cell(r, c).value) for c in range(1, min(80, ws.max_column or 80) + 1)}
        if req.issubset(row_vals):
            return r
    return None


def _find_col_idx(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    candidates: List[str],
    max_cols: int = 80,
) -> Optional[int]:
    want = {_norm(c) for c in candidates}
    for c in range(1, min(max_cols, ws.max_column or max_cols) + 1):
        if _norm(ws.cell(header_row, c).value) in want:
            return c
    return None


def _sig_bbva(fecha: dt.date, importe: float, detalle: str) -> tuple[str, float, str]:
    return (fecha.isoformat(), round(float(importe), 2), _norm_for_sig(detalle))


def _sig_mp(fecha: dt.date, importe: float, detalle: str) -> tuple[str, float, str]:
    return (fecha.isoformat(), round(float(importe), 2), _norm_for_sig(detalle))


def _sig_galicia(fecha: dt.date, importe: float, detalle: str, cuit: Optional[str]) -> tuple[str, float, str, str]:
    return (fecha.isoformat(), round(float(importe), 2), _norm_for_sig(detalle), str(cuit or ""))


def _next_append_row(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    *,
    header_row: int,
    key_cols: List[int],
) -> int:
    """Devuelve la primera fila libre real (ignorando filas vacías con solo estilo)."""
    max_r = ws.max_row or header_row
    last_with_data = header_row
    for r in range(header_row + 1, max_r + 1):
        has_data = False
        for c in key_cols:
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).strip()
            if s and s.lower() != "nan":
                has_data = True
                break
        if has_data:
            last_with_data = r
    return last_with_data + 1


def _existing_counter_bbva(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int, amount_col: int) -> Counter:
    counter: Counter = Counter()
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        fecha = _parse_date(ws.cell(r, 1).value)
        importe = _parse_amount(ws.cell(r, amount_col).value)
        detalle = str(ws.cell(r, 2).value or "").strip()
        if not fecha or importe is None or not detalle:
            continue
        counter[_sig_bbva(fecha, float(importe), detalle)] += 1
    return counter


def _existing_counter_mp(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    fecha_col: int,
    detalle_col: int,
    importe_col: int,
) -> Counter:
    counter: Counter = Counter()
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        fecha, _t = _parse_datetime(ws.cell(r, fecha_col).value)
        importe = _parse_amount(ws.cell(r, importe_col).value)
        detalle = str(ws.cell(r, detalle_col).value or "").strip()
        if not fecha or importe is None or not detalle:
            continue
        counter[_sig_mp(fecha, float(importe), detalle)] += 1
    return counter


def _existing_counter_galicia(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    fecha_col: int,
    detalle_col: int,
    importe_col: int,
    cuit_col: int,
) -> Counter:
    counter: Counter = Counter()
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        fecha = _parse_date(ws.cell(r, fecha_col).value)
        importe = _parse_amount(ws.cell(r, importe_col).value)
        detalle = str(ws.cell(r, detalle_col).value or "").strip()
        cuit = _extract_cuit(ws.cell(r, cuit_col).value)
        if not fecha or importe is None or not detalle:
            continue
        counter[_sig_galicia(fecha, float(importe), detalle, cuit)] += 1
    return counter


def _pick_bbva_sheet(wb: openpyxl.Workbook, source_name: str) -> Optional[str]:
    by_norm = {_norm(s): s for s in wb.sheetnames}
    gba_sheet = by_norm.get(_norm("Movimientos gba"))
    salice_bbva = by_norm.get(_norm("SALICE BBVA"))
    alarcon_bbva = by_norm.get(_norm(" ALARCON BBVA"))

    if gba_sheet:
        return gba_sheet
    if salice_bbva and not alarcon_bbva:
        return salice_bbva
    if alarcon_bbva and not salice_bbva:
        return alarcon_bbva
    if salice_bbva and alarcon_bbva:
        n = _norm(source_name)
        if "alarcon" in n:
            return alarcon_bbva
        if "salice" in n:
            return salice_bbva
        return salice_bbva
    return None


def _pick_galicia_sheet(wb: openpyxl.Workbook) -> Optional[str]:
    if "SALICE GALICIA (ALARCON)" in wb.sheetnames:
        return "SALICE GALICIA (ALARCON)"
    for s in wb.sheetnames:
        if "GALICIA" in s.upper():
            return s
    return None


_MONTH_SHEETS_ES = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}


def _pick_mp_sheet(wb: openpyxl.Workbook, fecha: dt.date | None = None) -> Optional[str]:
    if fecha is not None:
        wanted = _MONTH_SHEETS_ES.get(int(fecha.month))
        if wanted and wanted in wb.sheetnames:
            return wanted
    if "MercadoPago " in wb.sheetnames:
        return "MercadoPago "
    for s in wb.sheetnames:
        if "MERCADOPAGO" in _norm(s):
            return s
    if fecha is not None:
        for s in wb.sheetnames:
            if _norm(s) == _norm(_MONTH_SHEETS_ES.get(int(fecha.month), "")):
                return s
    return None


def _sheet_name_to_path_from_xlsx(xlsx_path: str) -> dict[str, str]:
    ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns_rel = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ns_pkgrel = "http://schemas.openxmlformats.org/package/2006/relationships"
    with zipfile.ZipFile(xlsx_path, "r") as z:
        wb_xml = z.read("xl/workbook.xml")
        rels_xml = z.read("xl/_rels/workbook.xml.rels")

    wb_root = ET.fromstring(wb_xml)
    rels_root = ET.fromstring(rels_xml)

    rid_to_target: dict[str, str] = {}
    for rel_el in rels_root.findall(f"{{{ns_pkgrel}}}Relationship"):
        rid = rel_el.attrib.get("Id")
        target = rel_el.attrib.get("Target")
        if rid and target:
            rid_to_target[rid] = target

    out: dict[str, str] = {}
    sheets_el = wb_root.find(f"{{{ns_main}}}sheets")
    if sheets_el is None:
        return out
    for sh_el in sheets_el.findall(f"{{{ns_main}}}sheet"):
        name = sh_el.attrib.get("name")
        rid = sh_el.attrib.get(f"{{{ns_rel}}}id")
        if not name or not rid:
            continue
        target = rid_to_target.get(rid)
        if not target:
            continue
        target = target.lstrip("/")
        if not target.startswith("xl/"):
            target = "xl/" + target
        out[name] = target
    return out


def _preserve_original_package(
    *,
    original_xlsx_path: str,
    modified_xlsx_path: str,
    out_xlsx_path: str,
    touched_sheet_names: set[str],
) -> None:
    # Si no hubo hojas modificadas, devolvemos copia bit-a-bit del original.
    if not touched_sheet_names:
        shutil.copy2(original_xlsx_path, out_xlsx_path)
        return

    original_map = _sheet_name_to_path_from_xlsx(original_xlsx_path)
    modified_map = _sheet_name_to_path_from_xlsx(modified_xlsx_path)

    touched_paths: set[str] = set()
    for name in touched_sheet_names:
        p_mod = modified_map.get(name)
        p_org = original_map.get(name)
        if p_mod:
            touched_paths.add(p_mod)
        if p_org:
            touched_paths.add(p_org)

    with zipfile.ZipFile(original_xlsx_path, "r") as z_org:
        original_files = {n: z_org.read(n) for n in z_org.namelist()}
    with zipfile.ZipFile(modified_xlsx_path, "r") as z_mod:
        modified_files = {n: z_mod.read(n) for n in z_mod.namelist()}

    merged = dict(original_files)
    for p in touched_paths:
        if p in modified_files:
            merged[p] = modified_files[p]
    # Las hojas tocadas pueden referenciar style ids emitidos por openpyxl.
    # Para mantener consistencia interna (y evitar "recuperar contenido"),
    # alineamos styles.xml con el origen de esas hojas.
    if "xl/styles.xml" in modified_files:
        merged["xl/styles.xml"] = modified_files["xl/styles.xml"]

    tmp = out_xlsx_path + ".tmp"
    with zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for name, data in merged.items():
            zout.writestr(name, data)
    os.replace(tmp, out_xlsx_path)


def _ensure_header_col(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    *,
    header_row: int,
    candidates: List[str],
    label: str,
) -> int:
    col = _find_col_idx(ws, header_row, candidates)
    if col is not None:
        return col
    col = int(ws.max_column or 0) + 1
    ws.cell(header_row, col, label)
    return col


def _append_bbva(ws: openpyxl.worksheet.worksheet.Worksheet, txns: List[RawBankTxn]) -> tuple[int, int]:
    header_row = (
        _find_header_row(ws, ["Fecha", "Concepto", "Crédito"])
        or _find_header_row(ws, ["Fecha", "Concepto", "Credito"])
        or _find_header_row(ws, ["Número Documento", "Importe"])
        or _find_header_row(ws, ["Numero Documento", "Importe"])
    )
    if header_row is None:
        raise ValueError(f"No pude detectar encabezado BBVA en hoja '{ws.title}'.")

    fecha_col = _find_col_idx(ws, header_row, ["Fecha"]) or 1
    fecha_valor_col = _find_col_idx(ws, header_row, ["Fecha Valor"])
    concepto_col = _find_col_idx(ws, header_row, ["Concepto"]) or 2
    numero_doc_col = _find_col_idx(ws, header_row, ["Número Documento", "Numero Documento"])
    oficina_col = _find_col_idx(ws, header_row, ["Oficina"])
    credito_col = _find_col_idx(ws, header_row, ["Crédito", "Credito"])
    importe_named_col = _find_col_idx(ws, header_row, ["Importe"])
    amount_col = credito_col or importe_named_col or 5
    detalle_extra_col = _find_col_idx(ws, header_row, ["Detalle"])
    gba_layout = bool(fecha_valor_col or detalle_extra_col or (credito_col is not None and importe_named_col is None))

    ok_col = _ensure_header_col(ws, header_row=header_row, candidates=["ok", "ok?", "recibio", "recibio?", "acreditado?"], label="ok")
    cliente_col = _ensure_header_col(ws, header_row=header_row, candidates=["cliente", "nro cliente", "nro_cliente"], label="cliente")
    recibo_col = _ensure_header_col(ws, header_row=header_row, candidates=["recibo", "nro recibo", "nro_recibo"], label="recibo")

    existing: Counter = Counter()
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        fecha = _parse_date(ws.cell(r, fecha_col).value)
        importe = _parse_amount(ws.cell(r, amount_col).value)
        if not fecha or importe is None or importe <= 0:
            continue
        concepto = str(ws.cell(r, concepto_col).value or "").strip()
        numero_doc = str(ws.cell(r, numero_doc_col).value or "").strip() if numero_doc_col else ""
        detalle_extra = str(ws.cell(r, detalle_extra_col).value or "").strip() if detalle_extra_col else ""
        if gba_layout:
            sig_text = " | ".join(x for x in [concepto, numero_doc, detalle_extra] if x)
        else:
            sig_text = concepto
        if not sig_text:
            continue
        existing[_sig_bbva(fecha, float(importe), sig_text)] += 1

    txns_sorted = sorted(txns, key=lambda t: (t.fecha, round(float(t.importe), 2), _norm_for_sig(t.detalle)))
    append_row = _next_append_row(ws, header_row=header_row, key_cols=[fecha_col, concepto_col, amount_col])
    skipped = 0
    added = 0
    for t in txns_sorted:
        sig = _sig_bbva(t.fecha, t.importe, t.detalle)
        if existing[sig] > 0:
            existing[sig] -= 1
            skipped += 1
            continue
        ws.cell(append_row, fecha_col, t.fecha.strftime("%d-%m-%Y"))
        if fecha_valor_col:
            ws.cell(append_row, fecha_valor_col, t.fecha.strftime("%d-%m-%Y"))
        if concepto_col:
            ws.cell(append_row, concepto_col, t.concept or t.detalle)
        if numero_doc_col:
            ws.cell(append_row, numero_doc_col, t.numero_documento or "")
        if oficina_col:
            ws.cell(append_row, oficina_col, t.oficina or "")
        ws.cell(append_row, amount_col, float(t.importe))
        if detalle_extra_col:
            ws.cell(append_row, detalle_extra_col, t.secondary_detail or "")
        elif not gba_layout:
            ws.cell(append_row, 2, t.detalle)
        ws.cell(append_row, ok_col, None)
        ws.cell(append_row, cliente_col, None)
        ws.cell(append_row, recibo_col, None)
        append_row += 1
        added += 1
    return added, skipped


def _append_mp(ws: openpyxl.worksheet.worksheet.Worksheet, txns: List[RawBankTxn]) -> tuple[int, int]:
    header_row = _find_header_row(ws, ["Fecha de Pago"])
    if header_row is None:
        raise ValueError(f"No pude detectar encabezado Mercado Pago en hoja '{ws.title}'.")

    fecha_col = _find_col_idx(ws, header_row, ["Fecha de Pago"]) or 1
    tipo_col = _find_col_idx(
        ws,
        header_row,
        [
            "Tipo de Operación",
            "Tipo de Operacion",
            "ID DE OPERACIÓN EN MERCADO PAGO",
            "ID DE OPERACION EN MERCADO PAGO",
        ],
    )
    detalle_col = _find_col_idx(
        ws,
        header_row,
        ["Operación Relacionada", "Operacion Relacionada", "ID DE OPERACIÓN EN MERCADO PAGO", "ID DE OPERACION EN MERCADO PAGO"],
    ) or 4
    importe_col = _find_col_idx(ws, header_row, ["Unnamed: 4", "Importe", "VALOR DE LA COMPRA"]) or 5
    gba_layout = _find_col_idx(ws, header_row, ["Control Logistica"]) is not None or _find_col_idx(ws, header_row, ["Importe"]) is not None
    cuit_col = _find_col_idx(
        ws,
        header_row,
        [
            "NÚMERO DE IDENTIFICACIÓN DEL PAGADOR",
            "NUMERO DE IDENTIFICACION DEL PAGADOR",
            "Número de identificación del pagador",
        ],
    )
    ok_col = _ensure_header_col(ws, header_row=header_row, candidates=["ok", "ok?", "recibio", "recibio?", "control logistica"], label="ok")
    cliente_col = _ensure_header_col(ws, header_row=header_row, candidates=["cliente", "nro cliente", "nro_cliente"], label="cliente")
    recibo_col = _ensure_header_col(ws, header_row=header_row, candidates=["recibo", "nro recibo", "nro_recibo"], label="recibo")

    if cuit_col is None:
        cuit_col = int(ws.max_column) + 1
        ws.cell(header_row, cuit_col, "NÚMERO DE IDENTIFICACIÓN DEL PAGADOR")

    existing: Counter = Counter()
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        fecha, _t = _parse_datetime(ws.cell(r, fecha_col).value)
        importe = _parse_amount(ws.cell(r, importe_col).value)
        detail_key = _plain_id(ws.cell(r, tipo_col).value) if (gba_layout and tipo_col) else str(ws.cell(r, detalle_col).value or "").strip()
        if not fecha or importe is None or importe <= 0 or not detail_key:
            continue
        existing[_sig_mp(fecha, float(importe), detail_key)] += 1

    txns_sorted = sorted(txns, key=lambda t: (t.fecha, round(float(t.importe), 2), _norm_for_sig(t.detalle)))
    append_row = _next_append_row(ws, header_row=header_row, key_cols=[fecha_col, tipo_col or detalle_col, importe_col])
    skipped = 0
    added = 0
    for t in txns_sorted:
        sig = _sig_mp(t.fecha, t.importe, t.detalle)
        if existing[sig] > 0:
            existing[sig] -= 1
            skipped += 1
            continue
        if t.hora is not None:
            ts = dt.datetime.combine(t.fecha, t.hora).isoformat(timespec="seconds")
            ws.cell(append_row, fecha_col, ts)
        else:
            ws.cell(append_row, fecha_col, t.fecha.strftime("%d/%m/%Y 00:00:00"))
        if tipo_col:
            ws.cell(append_row, tipo_col, t.detalle if gba_layout else "Cobro")
        ws.cell(append_row, detalle_col, (t.secondary_detail or "") if gba_layout else t.detalle)
        ws.cell(append_row, importe_col, float(t.importe))
        ws.cell(append_row, cuit_col, t.cuit or "")
        ws.cell(append_row, ok_col, None)
        ws.cell(append_row, cliente_col, None)
        ws.cell(append_row, recibo_col, None)
        append_row += 1
        added += 1
    return added, skipped


def _append_galicia(ws: openpyxl.worksheet.worksheet.Worksheet, txns: List[RawBankTxn]) -> tuple[int, int]:
    header_row = _find_header_row(ws, ["Fecha", "Importe"])
    if header_row is None:
        header_row = _find_header_row(ws, ["Fecha", "Créditos"]) or _find_header_row(ws, ["Fecha", "Creditos"])
    if header_row is None:
        raise ValueError(f"No pude detectar encabezado Galicia en hoja '{ws.title}'.")

    fecha_col = _find_col_idx(ws, header_row, ["Fecha"]) or 1
    detalle_col = _find_col_idx(ws, header_row, ["Concepto", "Descripción", "Descripcion"]) or 2
    razon_col = _find_col_idx(ws, header_row, ["Razon social", "Razón social", "Leyendas Adicionales 1"]) or 3
    cuit_col = _find_col_idx(ws, header_row, ["CUIT", "Leyendas Adicionales 2"]) or 4
    importe_col = _find_col_idx(ws, header_row, ["Importe", "Créditos", "Creditos"]) or 5
    ok_col = _find_col_idx(ws, header_row, ["ok", "ok?", "recibio", "recibio?"]) or 6
    cliente_col = _find_col_idx(ws, header_row, ["cliente", "nro cliente", "nro_cliente"]) or 7
    recibo_col = _find_col_idx(ws, header_row, ["recibo", "nro recibo", "nro_recibo"]) or 8

    existing = _existing_counter_galicia(ws, header_row, fecha_col, detalle_col, importe_col, cuit_col)
    txns_sorted = sorted(txns, key=lambda t: (t.fecha, round(float(t.importe), 2), _norm_for_sig(t.detalle)))
    append_row = _next_append_row(ws, header_row=header_row, key_cols=[fecha_col, detalle_col, importe_col])
    skipped = 0
    added = 0
    for t in txns_sorted:
        sig = _sig_galicia(t.fecha, t.importe, t.detalle, t.cuit)
        if existing[sig] > 0:
            existing[sig] -= 1
            skipped += 1
            continue
        ws.cell(append_row, fecha_col, t.fecha.isoformat())
        ws.cell(append_row, detalle_col, t.detalle)
        ws.cell(append_row, razon_col, t.razon_social or "")
        ws.cell(append_row, cuit_col, t.cuit or "")
        ws.cell(append_row, importe_col, float(t.importe))
        ws.cell(append_row, ok_col, None)
        ws.cell(append_row, cliente_col, None)
        ws.cell(append_row, recibo_col, None)
        append_row += 1
        added += 1
    return added, skipped


def build_runtime_workbook_from_raw(
    *,
    record_excel_path: str,
    raw_bank_paths: List[str],
    out_excel_path: str,
) -> Dict[str, object]:
    if not raw_bank_paths:
        raise ValueError("Tenés que subir al menos 1 archivo crudo bancario.")

    shutil.copy2(record_excel_path, out_excel_path)

    detected_rows: Dict[str, List[RawBankTxn]] = {"BBVA": [], "GALICIA": [], "MERCADOPAGO": []}
    files_meta: List[dict] = []
    raw_max_date: dt.date | None = None
    raw_min_date_all: dt.date | None = None
    for p in raw_bank_paths:
        bank, rows = parse_raw_bank_file(p)
        detected_rows.setdefault(bank, []).extend(rows)
        file_min_date: dt.date | None = None
        file_max_date: dt.date | None = None
        file_receipts_end_max_date: dt.date | None = None
        rows_eligible_for_end = 0
        for row in rows:
            if file_min_date is None or row.fecha < file_min_date:
                file_min_date = row.fecha
            if file_max_date is None or row.fecha > file_max_date:
                file_max_date = row.fecha
            if raw_min_date_all is None or row.fecha < raw_min_date_all:
                raw_min_date_all = row.fecha
            eligible_for_cutoff = True
            if row.bank == "BBVA":
                eligible_for_cutoff = _is_bbva_historical_sheet_name(row.source_sheet)
            if eligible_for_cutoff and (raw_max_date is None or row.fecha > raw_max_date):
                raw_max_date = row.fecha
            if eligible_for_cutoff:
                rows_eligible_for_end += 1
                if file_receipts_end_max_date is None or row.fecha > file_receipts_end_max_date:
                    file_receipts_end_max_date = row.fecha
        files_meta.append(
            {
                "filename": os.path.basename(p),
                "bank_detected": bank,
                "rows_parsed": len(rows),
                "rows_eligible_for_receipts_end_date": rows_eligible_for_end,
                "file_min_date": (file_min_date.isoformat() if file_min_date else None),
                "file_max_date": (file_max_date.isoformat() if file_max_date else None),
                "file_receipts_end_max_date": (
                    file_receipts_end_max_date.isoformat() if file_receipts_end_max_date else None
                ),
            }
        )

    raw_min_date: dt.date | None = raw_min_date_all
    recent_anchor_date = raw_max_date
    stale_files_for_receipts_start: list[str] = []
    if recent_anchor_date is not None:
        recent_file_mins: list[dt.date] = []
        for meta in files_meta:
            file_max_raw = meta.get("file_max_date")
            file_min_raw = meta.get("file_min_date")
            if not file_max_raw or not file_min_raw:
                continue
            try:
                file_max_date = dt.date.fromisoformat(str(file_max_raw))
                file_min_date = dt.date.fromisoformat(str(file_min_raw))
            except Exception:
                continue
            gap_days = (recent_anchor_date - file_max_date).days
            meta["days_behind_recent_anchor"] = int(gap_days)
            if gap_days <= RAW_RECENT_FILES_MAX_GAP_DAYS:
                recent_file_mins.append(file_min_date)
                meta["used_for_receipts_start_date"] = True
            else:
                stale_files_for_receipts_start.append(str(meta.get("filename") or ""))
                meta["used_for_receipts_start_date"] = False
        if recent_file_mins:
            raw_min_date = min(recent_file_mins)

    wb = openpyxl.load_workbook(out_excel_path)
    touched_sheet_names: set[str] = set()
    summary: Dict[str, dict] = {
        "BBVA": {"input": len(detected_rows.get("BBVA", [])), "appended": 0, "duplicates_skipped": 0, "sheet": None},
        "GALICIA": {"input": len(detected_rows.get("GALICIA", [])), "appended": 0, "duplicates_skipped": 0, "sheet": None},
        "MERCADOPAGO": {"input": len(detected_rows.get("MERCADOPAGO", [])), "appended": 0, "duplicates_skipped": 0, "sheet": None},
    }

    # BBVA puede tener dos hojas. Procesamos por archivo para decidir destino.
    bbva_sheet_default = _pick_bbva_sheet(wb, "")
    if detected_rows.get("BBVA"):
        if bbva_sheet_default is None:
            found = ", ".join(wb.sheetnames) if wb.sheetnames else "(sin hojas)"
            raise ValueError(
                "El consolidado bancario no tiene hoja BBVA compatible para esta corrida. "
                "Se esperaba 'Movimientos gba' (GBA) o las hojas legacy de BBVA. "
                f"Hojas encontradas: {found}"
            )
        by_source: Dict[str, List[RawBankTxn]] = {}
        for t in detected_rows["BBVA"]:
            by_source.setdefault(t.source_filename, []).append(t)
        for source_name, rows in by_source.items():
            sheet_name = _pick_bbva_sheet(wb, source_name) or bbva_sheet_default
            ws = wb[sheet_name]
            added, skipped = _append_bbva(ws, rows)
            summary["BBVA"]["appended"] += int(added)
            summary["BBVA"]["duplicates_skipped"] += int(skipped)
            summary["BBVA"]["sheet"] = sheet_name
            if added > 0:
                touched_sheet_names.add(sheet_name)

    if detected_rows.get("GALICIA"):
        galicia_sheet = _pick_galicia_sheet(wb)
        if galicia_sheet is None:
            raise ValueError("El consolidado no tiene hoja Galicia compatible.")
        ws = wb[galicia_sheet]
        added, skipped = _append_galicia(ws, detected_rows["GALICIA"])
        summary["GALICIA"]["appended"] = int(added)
        summary["GALICIA"]["duplicates_skipped"] = int(skipped)
        summary["GALICIA"]["sheet"] = galicia_sheet
        if added > 0:
            touched_sheet_names.add(galicia_sheet)

    if detected_rows.get("MERCADOPAGO"):
        by_sheet: Dict[str, List[RawBankTxn]] = {}
        for t in detected_rows["MERCADOPAGO"]:
            mp_sheet = _pick_mp_sheet(wb, t.fecha)
            if mp_sheet is None:
                raise ValueError("El consolidado no tiene hoja MercadoPago compatible para el mes del extracto.")
            by_sheet.setdefault(mp_sheet, []).append(t)
        touched_mp_sheets: list[str] = []
        for mp_sheet, rows in by_sheet.items():
            ws = wb[mp_sheet]
            added, skipped = _append_mp(ws, rows)
            summary["MERCADOPAGO"]["appended"] += int(added)
            summary["MERCADOPAGO"]["duplicates_skipped"] += int(skipped)
            touched_mp_sheets.append(mp_sheet)
            if added > 0:
                touched_sheet_names.add(mp_sheet)
        summary["MERCADOPAGO"]["sheet"] = ", ".join(touched_mp_sheets)

    modified_tmp = out_excel_path + ".modified_tmp.xlsx"
    wb.save(modified_tmp)
    _preserve_original_package(
        original_xlsx_path=record_excel_path,
        modified_xlsx_path=modified_tmp,
        out_xlsx_path=out_excel_path,
        touched_sheet_names=touched_sheet_names,
    )
    try:
        os.remove(modified_tmp)
    except Exception:
        pass

    return {
        "record_excel_filename": os.path.basename(record_excel_path),
        "raw_bank_files": files_meta,
        "raw_ingestion_summary": summary,
        "raw_total_input_rows": sum(v["input"] for v in summary.values()),
        "raw_total_appended_rows": sum(v["appended"] for v in summary.values()),
        "raw_min_date_all": (raw_min_date_all.isoformat() if raw_min_date_all else None),
        "raw_min_date": (raw_min_date.isoformat() if raw_min_date else None),
        "raw_max_date": (raw_max_date.isoformat() if raw_max_date else None),
        "raw_recent_files_max_gap_days": RAW_RECENT_FILES_MAX_GAP_DAYS,
        "raw_stale_files_ignored_for_receipts_start_date": stale_files_for_receipts_start,
    }
