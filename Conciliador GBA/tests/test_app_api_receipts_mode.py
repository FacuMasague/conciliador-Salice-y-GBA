from __future__ import annotations

from io import BytesIO
import re
import zipfile

import openpyxl
from fastapi.testclient import TestClient

import app as app_module


def _minimal_xlsx_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["a"])
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _strip_sheet_dimension(path) -> None:
    src = path.read_bytes()
    out = BytesIO()
    with zipfile.ZipFile(BytesIO(src), "r") as zin, zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename.startswith("xl/worksheets/sheet") and info.filename.endswith(".xml"):
                text = data.decode("utf-8")
                text = re.sub(r"<dimension[^>]*/>", "", text, count=1)
                data = text.encode("utf-8")
            zout.writestr(info, data)
    path.write_bytes(out.getvalue())


def test_detect_record_kinds_handles_xlsx_without_sheet_dimension(tmp_path):
    path = tmp_path / "bbva_no_dimension.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos gba"
    ws.append(["Fecha", "Fecha Valor", "Concepto", "Número Documento", "Oficina", "Crédito", "Débito", "Detalle"])
    ws.append(["2026-01-02", "2026-01-02", "TRANSF.BANEL 20301020304", "", "733 - N/A", 1000.0, "", "CTE 20301020304"])
    wb.save(path)
    _strip_sheet_dimension(path)

    assert app_module._detect_record_kinds(str(path)) == {"bank"}


def test_compare_api_mode_accepts_no_pdfs(monkeypatch):
    client = TestClient(app_module.app)

    def _fake_compare(*args, **kwargs):
        assert kwargs.get("receipts_source") == "api"
        return {"validados": [], "dudosos": [], "no_encontrados": [], "meta": {"ok": True}}

    monkeypatch.setattr(app_module, "compare_excel_pdfs", _fake_compare)

    xbytes = _minimal_xlsx_bytes()
    files = {
        "excel": ("legacy.xlsx", xbytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    }

    r = client.post("/compare?receipts_source=api&api_receipts_days=15", files=files)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["meta"]["app_version"] == app_module.APP_VERSION


def test_compare_pdf_mode_requires_pdf(monkeypatch):
    client = TestClient(app_module.app)

    xbytes = _minimal_xlsx_bytes()
    files = {
        "excel": ("legacy.xlsx", xbytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    }

    r = client.post("/compare?receipts_source=pdf", files=files)
    assert r.status_code == 400
    assert "al menos 1 PDF" in r.text


def test_compare_api_mode_passes_manual_receipts_date_range(monkeypatch):
    client = TestClient(app_module.app)
    observed = {"start": None, "end": None}

    def _fake_compare(*args, **kwargs):
        observed["start"] = kwargs.get("api_start_date_override")
        observed["end"] = kwargs.get("api_end_date_override")
        return {"validados": [], "dudosos": [], "no_encontrados": [], "meta": {"ok": True}}

    monkeypatch.setattr(app_module, "compare_excel_pdfs", _fake_compare)

    xbytes = _minimal_xlsx_bytes()
    files = {
        "excel": ("legacy.xlsx", xbytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    }

    r = client.post(
        "/compare?receipts_source=api&api_start_date=2026-03-01&api_end_date=2026-03-07",
        files=files,
    )
    assert r.status_code == 200, r.text
    assert observed["start"] == "2026-03-01"
    assert observed["end"] == "2026-03-07"


def test_export_api_mode_returns_combined_workbook_when_split_records_are_prepared(monkeypatch, tmp_path):
    client = TestClient(app_module.app)

    bank_path = tmp_path / "bank.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos gba"
    ws.append(["Fecha", "Fecha Valor", "Concepto", "Número Documento", "Oficina", "Crédito", "Débito", "Detalle", "Acreditado?"])
    ws.append(["20-03-2026", "20-03-2026", "DEPOSITO AUT BUZON/01/07:02", "", "331 - PANAMERICANA", 29000.0, "", "CTE 000000495002", ""])
    wb.save(bank_path)

    mp_path = tmp_path / "mp.xlsx"
    wb_mp = openpyxl.Workbook()
    ws_mp = wb_mp.active
    ws_mp.title = "MARZO"
    ws_mp.append(["Fecha de Pago", "Tipo de Operación", "Operación Relacionada", "Importe", "Control Logistica"])
    ws_mp.append(["2026-03-21T12:24:12.000-04:00", "151349770684", "Dinero disponible", "158792.00", ""])
    wb_mp.save(mp_path)

    async def _fake_prepare(**kwargs):
        return {
            "working_excel_paths": [str(bank_path), str(mp_path)],
            "records": [
                {"key": "bank", "working_excel_path": str(bank_path), "base_excel_filename": "Movimientos bancarios 2026.xlsx", "origins": ["BBVA"], "export_mode": "generic"},
                {"key": "mp", "working_excel_path": str(mp_path), "base_excel_filename": "Mercado Pago 2026.xlsx", "origins": ["MERCADOPAGO"], "export_mode": "generic"},
            ],
            "excel_record_map": {str(bank_path): "bank", str(mp_path): "mp"},
            "base_excel_filename": None,
            "input_mode": "v5_split_records",
            "raw_bank_filenames": [],
            "raw_ingestion_meta": {"raw_max_date": "2026-03-21"},
        }

    def _fake_compare(*args, **kwargs):
        return {
            "validados": [
                {
                    "Ranking": 1,
                    "Origen": "BBVA",
                    "Fila Excel": 2,
                    "Nro cliente": "33119",
                    "Cliente": "Cliente Banco",
                    "Nro recibo": "272641",
                    "Fecha recibo": "2026-03-20",
                    "Medio de pago": "TRANSFERENCIA",
                    "Importe recibo": 29000.0,
                    "__sheet_name": "Movimientos gba",
                    "__record_key": "bank",
                },
                {
                    "Ranking": 1,
                    "Origen": "MERCADOPAGO",
                    "Fila Excel": 2,
                    "Nro cliente": "33120",
                    "Cliente": "Cliente MP",
                    "Nro recibo": "272642",
                    "Fecha recibo": "2026-03-21",
                    "Medio de pago": "MERCADOPAGO",
                    "Importe recibo": 158792.0,
                    "__sheet_name": "MARZO",
                    "__record_key": "mp",
                },
            ],
            "dudosos": [],
            "no_encontrados": [],
            "meta": {},
        }

    monkeypatch.setattr(app_module, "_prepare_excel_for_run", _fake_prepare)
    monkeypatch.setattr(app_module, "compare_excel_pdfs", _fake_compare)

    r = client.post("/export?receipts_source=api&format=xlsx")
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    wb = openpyxl.load_workbook(BytesIO(r.content), read_only=True)
    assert wb.sheetnames == ["BBVA", "Mercado Pago"]
    bbva_headers = [c.value for c in next(wb["BBVA"].iter_rows(min_row=1, max_row=1))]
    mp_headers = [c.value for c in next(wb["Mercado Pago"].iter_rows(min_row=1, max_row=1))]
    assert "recibo" in [str(h).strip().lower() for h in bbva_headers]
    assert "recibo" in [str(h).strip().lower() for h in mp_headers]
