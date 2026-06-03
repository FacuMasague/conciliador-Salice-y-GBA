from __future__ import annotations

import datetime as dt

import openpyxl
import pandas as pd

from src.conciliador.raw_bank_ingest import (
    build_runtime_workbook_from_raw,
    detect_raw_bank_kind,
    parse_raw_bank_file,
)


def _build_record_workbook(path: str) -> None:
    wb = openpyxl.Workbook()
    ws0 = wb.active
    ws0.title = "SALICE BBVA"
    ws0.append(["BBVA Frances"])
    ws0.append(["Cuentas"])
    ws0.append(["", "A", "Número Documento", "Oficina", "Importe", "ok", "cliente", "recibo"])
    ws0.append(["2026-02-01", "TRANSF.BANEL 20301020304", "136", "733 - N/A", 1000.0, "", "", ""])

    ws1 = wb.create_sheet(" ALARCON BBVA")
    ws1.append(["ALARCON"])
    ws1.append(["CTA"])
    ws1.append(["", "A", "Número Documento", "Oficina", "Importe", "ok", "cliente", "recibo"])

    ws2 = wb.create_sheet("SALICE GALICIA (ALARCON)")
    ws2.append(["Fecha", "Concepto", "Razon social", "CUIT", "Importe", "ok", "Cliente", "Recibo"])
    ws2.append([dt.datetime(2026, 2, 1), "Transferencia De Terceros", "CLIENTE A", "20301020304", 2000.0, "", "", ""])

    ws3 = wb.create_sheet("MercadoPago ")
    ws3.append(["Fecha de Pago", "Tipo de Operación", "Número de Movimiento", "Operación Relacionada", "Unnamed: 4", "Recibio?", "Cliente", "Recibo"])
    ws3.append([dt.datetime(2026, 2, 1, 10, 0, 0), "Cobro", "", "123456", 3000.0, "", "", ""])

    wb.save(path)


def test_detect_and_parse_raw_bank_files(tmp_path):
    mp_path = tmp_path / "mp.xlsx"
    df_mp = pd.DataFrame(
        {
            "NÚMERO DE IDENTIFICACIÓN": ["default"],
            "ID DE OPERACIÓN EN MERCADO PAGO": ["999001"],
            "VALOR DE LA COMPRA": ["3500.00"],
            "MONEDA": ["ARS"],
            "FECHA DE ORIGEN": ["2026-02-03T18:31:06.000-04:00"],
            "FECHA DE APROBACIÓN": ["2026-02-03T18:31:07.000-04:00"],
            "MONTO NETO DE LA OPERACIÓN": ["3400.00"],
            "TIPO DE IDENTIFICACIÓN DEL PAGADOR": ["CUIT"],
            "NÚMERO DE IDENTIFICACIÓN DEL PAGADOR": ["20301020304"],
            "PAGADOR": ["Comercio X"],
        }
    )
    df_mp.to_excel(mp_path, index=False)

    gal_path = tmp_path / "galicia.xlsx"
    df_gal = pd.DataFrame(
        {
            "Fecha": ["2026-02-03"],
            "Descripción": ["Transferencia De Terceros"],
            "Origen": ["A0D5"],
            "Créditos": ["4500.50"],
            "Número de Terminal": [""],
            "Observaciones Cliente": [""],
            "Número de Comprobante": [""],
            "Leyendas Adicionales 1": ["CLIENTE Y"],
            "Leyendas Adicionales 2": ["20301020304"],
            "Leyendas Adicionales 3": [""],
            "Leyendas Adicionales 4": [""],
            "Tipo de Movimiento": ["Imputado"],
        }
    )
    df_gal.to_excel(gal_path, index=False)

    bbva_path = tmp_path / "bbva.xlsx"
    with pd.ExcelWriter(bbva_path, engine="openpyxl") as w:
        pd.DataFrame([["x"] for _ in range(6)]).to_excel(w, index=False, header=False)
        pd.DataFrame(
            {
                "Fecha": ["2026-02-03"],
                "Concepto": ["TRANSF.BANEL 20301020304"],
                "Col3": [""],
                "Col4": [""],
                "Col5": [""],
                "Col6": [""],
                "Credito": ["5500.00"],
            }
        ).to_excel(w, index=False, startrow=6)

    assert detect_raw_bank_kind(str(mp_path)) == "MERCADOPAGO"
    assert detect_raw_bank_kind(str(gal_path)) == "GALICIA"
    assert detect_raw_bank_kind(str(bbva_path)) == "BBVA"

    _, mp_rows = parse_raw_bank_file(str(mp_path))
    _, gal_rows = parse_raw_bank_file(str(gal_path))
    _, bbva_rows = parse_raw_bank_file(str(bbva_path))

    assert len(mp_rows) == 1
    assert len(gal_rows) == 1
    assert len(bbva_rows) == 1
    assert bbva_rows[0].cuit == "20301020304"


def test_detect_and_parse_mp_new_layout_uses_counterpart_document_as_cuit(tmp_path):
    mp_path = tmp_path / "mp_new_layout.xlsx"
    df_mp = pd.DataFrame(
        {
            "Fecha de compra (date_created)": ["22/02/2026 09:53:15", "22/02/2026 10:15:00"],
            "Nombre de la contraparte (counterpart_name)": ["Cliente A", "Cliente B"],
            "Número de operación de Mercado Pago (operation_id)": ["146597529109", "146597529110"],
            "Estado de la operación (status)": ["approved", "approved"],
            "Detalle del estado de la operación (status_detail)": ["accredited", "accredited"],
            "Valor del producto (transaction_amount)": ["827553.00", "1000.00"],
            "Monto recibido (net_received_amount)": ["798009.36", "980.00"],
            "Monto devuelto (amount_refunded)": ["0.00", "0.00"],
            "Documento de la contraparte (buyer_document)": ["20-26756539-3", ""],
        }
    )
    df_mp.to_excel(mp_path, index=False)

    assert detect_raw_bank_kind(str(mp_path)) == "MERCADOPAGO"
    _, rows = parse_raw_bank_file(str(mp_path))
    assert len(rows) == 2
    assert rows[0].cuit == "20267565393"
    assert rows[1].cuit is None


def test_bbva_dd_mm_yyyy_date_is_parsed_as_day_first(tmp_path):
    bbva_path = tmp_path / "bbva_ddmmyyyy.xlsx"
    with pd.ExcelWriter(bbva_path, engine="openpyxl") as w:
        pd.DataFrame([["x"] for _ in range(6)]).to_excel(w, index=False, header=False)
        pd.DataFrame(
            {
                "Fecha": ["02-01-2026"],
                "Concepto": ["TRANSF.BANEL 20301020304"],
                "Col3": [""],
                "Col4": [""],
                "Col5": [""],
                "Col6": [""],
                "Credito": ["500000.00"],
            }
        ).to_excel(w, index=False, startrow=6)

    _, rows = parse_raw_bank_file(str(bbva_path))
    assert len(rows) == 1
    assert rows[0].fecha.isoformat() == "2026-01-02"


def test_bbva_reads_multiple_sheets_and_dedupes(tmp_path):
    bbva_path = tmp_path / "bbva_multisheet.xlsx"
    with pd.ExcelWriter(bbva_path, engine="openpyxl") as w:
        # Sheet 1: "día"
        pd.DataFrame([["x"] for _ in range(6)]).to_excel(w, index=False, header=False, sheet_name="Movimientos del Día")
        pd.DataFrame(
            {
                "Fecha": ["20-02-2026"],
                "Concepto": ["TRANSF CREDITO BANELCO"],
                "C3": [""],
                "C4": [""],
                "C5": [""],
                "C6": [""],
                "Crédito": ["200000"],
                "Detalle": ["CTE 000000200000"],
            }
        ).to_excel(w, index=False, startrow=6, sheet_name="Movimientos del Día")

        # Sheet 2: históricos (incluye un duplicado + uno nuevo)
        pd.DataFrame([["x"] for _ in range(6)]).to_excel(w, index=False, header=False, sheet_name="Movimientos Históricos")
        pd.DataFrame(
            {
                "Fecha": ["20-02-2026", "09-02-2026"],
                "Concepto": ["TRANSF CREDITO BANELCO", "TRANSF.BANEL 20301020304"],
                "C3": ["", ""],
                "C4": ["", ""],
                "C5": ["", ""],
                "C6": ["", ""],
                "Crédito": ["200000", "154682.13"],
                "Detalle": ["CTE 000000200000", "CTE 20301020304"],
            }
        ).to_excel(w, index=False, startrow=6, sheet_name="Movimientos Históricos")

    _, rows = parse_raw_bank_file(str(bbva_path))
    # Debe traer ambos movimientos únicos (sin duplicar el de 200000).
    assert len(rows) == 2
    amounts = sorted(round(r.importe, 2) for r in rows)
    assert amounts == [154682.13, 200000.00]


def test_bbva_gba_raw_keeps_all_positive_credit_rows(tmp_path):
    bbva_path = tmp_path / "bbva_gba_filter.xlsx"
    with pd.ExcelWriter(bbva_path, engine="openpyxl") as w:
        pd.DataFrame([["x"] for _ in range(6)]).to_excel(w, index=False, header=False, sheet_name="Movimientos Históricos")
        pd.DataFrame(
            {
                "Fecha": ["01-04-2026", "01-04-2026", "01-04-2026", "01-04-2026", "01-04-2026"],
                "Fecha Valor": ["01-04-2026"] * 5,
                "Concepto": [
                    "TRANSF.BANEL 20047049432",
                    "DNET CREDITO NE3117596",
                    "DEPOSITO AUT BUZON/01/15:12",
                    "TRANSF. CLIE 203407713375",
                    "EFECTIVO",
                ],
                "Número Documento": ["", "", "", "", ""],
                "Oficina": ["733 - N/A", "587 - DATANET", "195 - TORTUGAS OPEN MALL", "100 - BANCA ONLINE", "337 - MORON"],
                "Crédito": ["20086.19", "364386.04", "194000", "987258.52", "4050550"],
                "Débito": ["", "", "", "", ""],
                "Detalle": [
                    "CTE 20047049432",
                    "CTE 117596       007-002389330101       DELTA ISLAND ASES",
                    "CTE 000017917002",
                    "CTA.ORIGEN:  CC $ 340-771337/5",
                    "",
                ],
            }
        ).to_excel(w, index=False, startrow=6, sheet_name="Movimientos Históricos")

    _, rows = parse_raw_bank_file(str(bbva_path))
    assert len(rows) == 5
    detalles = [r.detalle for r in rows]
    assert any(d.endswith("CTE 20047049432") for d in detalles)
    assert any("DNET CREDITO NE3117596" in d for d in detalles)
    assert any("DEPOSITO AUT BUZON/01/15:12" in d for d in detalles)
    assert any("TRANSF. CLIE 203407713375" in d for d in detalles)
    assert any(d == "EFECTIVO" for d in detalles)


def test_build_runtime_workbook_from_raw_appends_new_rows_and_skips_duplicates(tmp_path):
    record_path = tmp_path / "record.xlsx"
    _build_record_workbook(str(record_path))

    bbva_raw = tmp_path / "bbva_salice.xlsx"
    with pd.ExcelWriter(bbva_raw, engine="openpyxl") as w:
        pd.DataFrame([["x"] for _ in range(6)]).to_excel(w, index=False, header=False)
        pd.DataFrame(
            {
                "Fecha": ["2026-02-01", "2026-02-02"],
                "Concepto": ["TRANSF.BANEL 20301020304", "TRANSF.BANEL 20999888776"],
                "Dummy1": ["", ""],
                "Dummy2": ["", ""],
                "Dummy3": ["", ""],
                "Dummy4": ["", ""],
                "Credito": ["1000.00", "9999.00"],
            }
        ).to_excel(w, index=False, startrow=6)

    gal_raw = tmp_path / "galicia.xlsx"
    pd.DataFrame(
        {
            "Fecha": ["2026-02-01", "2026-02-02"],
            "Descripción": ["Transferencia De Terceros", "Credito Transferencia Coelsa"],
            "Origen": ["A0D5", "A0D5"],
            "Créditos": ["2000.00", "7777.00"],
            "Número de Terminal": ["", ""],
            "Observaciones Cliente": ["", ""],
            "Número de Comprobante": ["", ""],
            "Leyendas Adicionales 1": ["CLIENTE A", "CLIENTE B"],
            "Leyendas Adicionales 2": ["20301020304", "20111222333"],
            "Leyendas Adicionales 3": ["", ""],
            "Leyendas Adicionales 4": ["", ""],
            "Tipo de Movimiento": ["Imputado", "Imputado"],
        }
    ).to_excel(gal_raw, index=False)

    mp_raw = tmp_path / "mp.xlsx"
    pd.DataFrame(
        {
            "NÚMERO DE IDENTIFICACIÓN": ["default", "default", "default"],
            "ID DE OPERACIÓN EN MERCADO PAGO": ["123456", "654321", "111111"],
            "VALOR DE LA COMPRA": ["3000.00", "8888.00", "-12.00"],
            "MONEDA": ["ARS", "ARS", "ARS"],
            "FECHA DE ORIGEN": ["2026-02-01T10:00:00.000-04:00", "2026-02-02T11:00:00.000-04:00", "2026-02-02T11:30:00.000-04:00"],
            "FECHA DE APROBACIÓN": ["", "", ""],
            "MONTO NETO DE LA OPERACIÓN": ["", "", ""],
            "TIPO DE IDENTIFICACIÓN DEL PAGADOR": ["CUIT", "CUIT", "CUIT"],
            "NÚMERO DE IDENTIFICACIÓN DEL PAGADOR": ["20301020304", "20111222333", "20999888776"],
            "PAGADOR": ["CLIENTE A", "CLIENTE C", "CLIENTE D"],
        }
    ).to_excel(mp_raw, index=False)

    out_path = tmp_path / "runtime.xlsx"
    meta = build_runtime_workbook_from_raw(
        record_excel_path=str(record_path),
        raw_bank_paths=[str(bbva_raw), str(gal_raw), str(mp_raw)],
        out_excel_path=str(out_path),
    )

    assert meta["raw_total_input_rows"] == 6
    assert meta["raw_total_appended_rows"] == 3
    assert meta["raw_ingestion_summary"]["BBVA"]["duplicates_skipped"] == 1
    assert meta["raw_ingestion_summary"]["GALICIA"]["duplicates_skipped"] == 1
    assert meta["raw_ingestion_summary"]["MERCADOPAGO"]["duplicates_skipped"] == 1

    wb = openpyxl.load_workbook(out_path, data_only=True)
    assert wb["SALICE BBVA"].max_row >= 5
    assert wb["SALICE GALICIA (ALARCON)"].max_row >= 3
    assert wb["MercadoPago "].max_row >= 3
    ws_mp = wb["MercadoPago "]
    header = [str(c.value or "").strip() for c in ws_mp[1]]
    assert "NÚMERO DE IDENTIFICACIÓN DEL PAGADOR" in header
    cuit_col = header.index("NÚMERO DE IDENTIFICACIÓN DEL PAGADOR") + 1
    cuit_values = [str(ws_mp.cell(r, cuit_col).value or "").strip() for r in range(2, ws_mp.max_row + 1)]
    assert any(v for v in cuit_values)


def test_build_runtime_workbook_from_raw_supports_gba_records(tmp_path):
    bank_record = tmp_path / "bank_record.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos gba"
    ws.append(["Fecha", "Fecha Valor", "Concepto", "Número Documento", "Oficina", "Crédito", "Débito", "Detalle", "Acreditado?"])
    ws.append(["20-03-2026", "20-03-2026", "TRANSF.BANEL 20302962112", "", "733 - N/A", 29000.0, "", "CTE 20302962112", ""])
    wb.save(bank_record)

    bbva_raw = tmp_path / "bbva_gba.xlsx"
    with pd.ExcelWriter(bbva_raw, engine="openpyxl") as w:
        pd.DataFrame([["x"] for _ in range(6)]).to_excel(w, index=False, header=False, sheet_name="Movimientos Históricos")
        pd.DataFrame(
            {
                "Fecha": ["20-03-2026", "21-03-2026"],
                "Fecha Valor": ["20-03-2026", "21-03-2026"],
                "Concepto": ["TRANSF.BANEL 20302962112", "TRF  IN COEL 30717691543"],
                "Codigo": ["136", "129"],
                "Número Documento": ["", ""],
                "Oficina": ["733 - N/A", "100 - BANCA ONLINE"],
                "Crédito": ["29000", "532600"],
                "Débito": ["", ""],
                "Detalle": ["CTE 20302962112", "CTE 000011110008"],
            }
        ).to_excel(w, index=False, startrow=6, sheet_name="Movimientos Históricos")

    out_bank = tmp_path / "bank_runtime.xlsx"
    meta_bank = build_runtime_workbook_from_raw(
        record_excel_path=str(bank_record),
        raw_bank_paths=[str(bbva_raw)],
        out_excel_path=str(out_bank),
    )

    assert meta_bank["raw_ingestion_summary"]["BBVA"]["appended"] == 1
    wb_bank = openpyxl.load_workbook(out_bank, data_only=True)
    ws_bank = wb_bank["Movimientos gba"]
    assert ws_bank.max_row == 3
    assert ws_bank.cell(3, 3).value == "TRF  IN COEL 30717691543"
    assert ws_bank.cell(3, 8).value == "CTE 000011110008"

    mp_record = tmp_path / "mp_record.xlsx"
    wb_mp = openpyxl.Workbook()
    ws_mp = wb_mp.active
    ws_mp.title = "MARZO"
    ws_mp.append(["Fecha de Pago", "Tipo de Operación", "Operación Relacionada", "Importe", "Control Logistica"])
    ws_mp.append(["2026-03-21T12:24:12.000-04:00", "151349770684", "Dinero disponible", "158792.00", ""])
    wb_mp.save(mp_record)

    mp_raw = tmp_path / "mp_gba.xlsx"
    pd.DataFrame(
        {
            "NÚMERO DE IDENTIFICACIÓN": ["default", "default"],
            "ID DE OPERACIÓN EN MERCADO PAGO": ["151349770684", "151348019862"],
            "MEDIO DE PAGO": ["Dinero disponible", "Dinero disponible"],
            "VALOR DE LA COMPRA": ["158792.00", "252373.00"],
            "FECHA DE ORIGEN": ["2026-03-21T12:24:12.000-04:00", "2026-03-21T12:12:32.000-04:00"],
            "NÚMERO DE IDENTIFICACIÓN DEL PAGADOR": ["20301020304", "20111222333"],
            "PAGADOR": ["Cliente A", "Cliente B"],
        }
    ).to_excel(mp_raw, index=False)

    out_mp = tmp_path / "mp_runtime.xlsx"
    meta_mp = build_runtime_workbook_from_raw(
        record_excel_path=str(mp_record),
        raw_bank_paths=[str(mp_raw)],
        out_excel_path=str(out_mp),
    )

    assert meta_mp["raw_ingestion_summary"]["MERCADOPAGO"]["appended"] == 1
    wb_mp_out = openpyxl.load_workbook(out_mp, data_only=True)
    ws_mp_out = wb_mp_out["MARZO"]
    assert ws_mp_out.max_row == 3
    assert ws_mp_out.cell(3, 2).value == "151348019862"
    assert ws_mp_out.cell(3, 3).value == "Dinero disponible"


def test_build_runtime_workbook_from_raw_accepts_combined_export_sheet_names(tmp_path):
    record = tmp_path / "combined_record.xlsx"
    wb = openpyxl.Workbook()
    ws_bank = wb.active
    ws_bank.title = "BBVA"
    ws_bank.append(["Fecha", "Fecha Valor", "Concepto", "Número Documento", "Oficina", "Crédito", "Débito", "Detalle", "Acreditado?", "ok", "cliente", "recibo"])
    ws_bank.append(["20-03-2026", "20-03-2026", "TRANSF.BANEL 20302962112", "", "733 - N/A", 29000.0, "", "CTE 20302962112", "", "", "", ""])
    ws_mp = wb.create_sheet("Mercado Pago")
    ws_mp.append(["Fecha de Pago", "Tipo de Operación", "Operación Relacionada", "Importe", "Control Logistica", "ok", "cliente", "recibo"])
    ws_mp.append(["2026-03-21T12:24:12.000-04:00", "151349770684", "Dinero disponible", "158792.00", "", "", "", ""])
    wb.save(record)

    bbva_raw = tmp_path / "bbva_gba.xlsx"
    with pd.ExcelWriter(bbva_raw, engine="openpyxl") as w:
        pd.DataFrame([["x"] for _ in range(6)]).to_excel(w, index=False, header=False, sheet_name="Movimientos Históricos")
        pd.DataFrame(
            {
                "Fecha": ["21-03-2026"],
                "Fecha Valor": ["21-03-2026"],
                "Concepto": ["TRF  IN COEL 30717691543"],
                "Número Documento": [""],
                "Oficina": ["100 - BANCA ONLINE"],
                "Crédito": ["532600"],
                "Débito": [""],
                "Detalle": ["CTE 000011110008"],
            }
        ).to_excel(w, index=False, startrow=6, sheet_name="Movimientos Históricos")

    mp_raw = tmp_path / "mp_gba.xlsx"
    pd.DataFrame(
        {
            "NÚMERO DE IDENTIFICACIÓN": ["default"],
            "ID DE OPERACIÓN EN MERCADO PAGO": ["151348019862"],
            "MEDIO DE PAGO": ["Dinero disponible"],
            "VALOR DE LA COMPRA": ["252373.00"],
            "FECHA DE ORIGEN": ["2026-03-21T12:12:32.000-04:00"],
            "NÚMERO DE IDENTIFICACIÓN DEL PAGADOR": ["20111222333"],
            "PAGADOR": ["Cliente B"],
        }
    ).to_excel(mp_raw, index=False)

    out = tmp_path / "runtime.xlsx"
    meta = build_runtime_workbook_from_raw(
        record_excel_path=str(record),
        raw_bank_paths=[str(bbva_raw), str(mp_raw)],
        out_excel_path=str(out),
    )

    assert meta["raw_ingestion_summary"]["BBVA"]["sheet"] == "BBVA"
    assert meta["raw_ingestion_summary"]["MERCADOPAGO"]["sheet"] == "Mercado Pago"
    wb_out = openpyxl.load_workbook(out, data_only=True)
    assert wb_out["BBVA"].max_row == 3
    assert wb_out["BBVA"].cell(3, 3).value == "TRF  IN COEL 30717691543"
    assert wb_out["Mercado Pago"].max_row == 3
    assert wb_out["Mercado Pago"].cell(3, 2).value == "151348019862"


def test_appended_rows_have_no_gap_and_are_oldest_to_newest(tmp_path):
    record_path = tmp_path / "record.xlsx"
    _build_record_workbook(str(record_path))

    bbva_raw = tmp_path / "bbva_order.xlsx"
    with pd.ExcelWriter(bbva_raw, engine="openpyxl") as w:
        pd.DataFrame([["x"] for _ in range(6)]).to_excel(w, index=False, header=False)
        # Vienen desordenados (nuevo -> viejo); el sistema debe escribir viejo -> nuevo.
        pd.DataFrame(
            {
                "Fecha": ["2026-02-03", "2026-02-02"],
                "Concepto": ["TRANSF.BANEL 20000000002", "TRANSF.BANEL 20000000001"],
                "D1": ["", ""],
                "D2": ["", ""],
                "D3": ["", ""],
                "D4": ["", ""],
                "Credito": ["3000.00", "2000.00"],
            }
        ).to_excel(w, index=False, startrow=6)

    out_path = tmp_path / "runtime.xlsx"
    build_runtime_workbook_from_raw(
        record_excel_path=str(record_path),
        raw_bank_paths=[str(bbva_raw)],
        out_excel_path=str(out_path),
    )

    wb = openpyxl.load_workbook(out_path, data_only=True)
    ws = wb["SALICE BBVA"]
    # En el record base la última fila con datos es la 4, por lo tanto no debe haber hueco en fila 5.
    assert str(ws.cell(5, 2).value or "").strip() != ""
    # Orden cronológico ascendente en las nuevas filas.
    assert str(ws.cell(5, 1).value).startswith("02-02-2026")
    assert str(ws.cell(6, 1).value).startswith("03-02-2026")


def test_build_runtime_workbook_from_raw_uses_bbva_daily_for_receipts_end_date(tmp_path):
    record_path = tmp_path / "record.xlsx"
    _build_record_workbook(str(record_path))

    bbva_raw = tmp_path / "bbva_mix.xlsx"
    with pd.ExcelWriter(bbva_raw, engine="openpyxl") as w:
        pd.DataFrame([["x"] for _ in range(6)]).to_excel(w, index=False, header=False, sheet_name="Movimientos del Día")
        pd.DataFrame(
            {
                "Fecha": ["04-03-2026"],
                "Concepto": ["TRANSF.BANEL 20999888776"],
                "C3": [""],
                "C4": [""],
                "C5": [""],
                "C6": [""],
                "Crédito": ["9999.00"],
                "Detalle": ["CTE 20999888776"],
            }
        ).to_excel(w, index=False, startrow=6, sheet_name="Movimientos del Día")

        pd.DataFrame([["x"] for _ in range(6)]).to_excel(w, index=False, header=False, sheet_name="Movimientos Históricos")
        pd.DataFrame(
            {
                "Fecha": ["03-03-2026"],
                "Concepto": ["TRANSF.BANEL 20301020304"],
                "C3": [""],
                "C4": [""],
                "C5": [""],
                "C6": [""],
                "Crédito": ["154682.13"],
                "Detalle": ["CTE 20301020304"],
            }
        ).to_excel(w, index=False, startrow=6, sheet_name="Movimientos Históricos")

    out_path = tmp_path / "runtime.xlsx"
    meta = build_runtime_workbook_from_raw(
        record_excel_path=str(record_path),
        raw_bank_paths=[str(bbva_raw)],
        out_excel_path=str(out_path),
    )

    assert meta["raw_max_date"] == "2026-03-04"


def test_build_runtime_workbook_from_raw_ignores_stale_file_for_receipts_start_date(tmp_path):
    record_path = tmp_path / "record.xlsx"
    _build_record_workbook(str(record_path))

    mp_raw = tmp_path / "mp_recent.xlsx"
    pd.DataFrame(
        {
            "NÚMERO DE IDENTIFICACIÓN": ["default", "default"],
            "ID DE OPERACIÓN EN MERCADO PAGO": ["123456", "654321"],
            "VALOR DE LA COMPRA": ["3000.00", "8888.00"],
            "MONEDA": ["ARS", "ARS"],
            "FECHA DE ORIGEN": ["2026-03-06T10:00:00.000-04:00", "2026-03-07T11:00:00.000-04:00"],
            "FECHA DE APROBACIÓN": ["", ""],
            "MONTO NETO DE LA OPERACIÓN": ["", ""],
            "TIPO DE IDENTIFICACIÓN DEL PAGADOR": ["CUIT", "CUIT"],
            "NÚMERO DE IDENTIFICACIÓN DEL PAGADOR": ["20301020304", "20111222333"],
            "PAGADOR": ["CLIENTE A", "CLIENTE B"],
        }
    ).to_excel(mp_raw, index=False)

    gal_raw = tmp_path / "galicia_old.xlsx"
    pd.DataFrame(
        {
            "Fecha": ["2026-02-06"],
            "Descripción": ["Transferencia De Terceros"],
            "Origen": ["A0D5"],
            "Créditos": ["4500.50"],
            "Número de Terminal": [""],
            "Observaciones Cliente": [""],
            "Número de Comprobante": [""],
            "Leyendas Adicionales 1": ["CLIENTE VIEJO"],
            "Leyendas Adicionales 2": ["20301020304"],
            "Leyendas Adicionales 3": [""],
            "Leyendas Adicionales 4": [""],
            "Tipo de Movimiento": ["Imputado"],
        }
    ).to_excel(gal_raw, index=False)

    bbva_raw = tmp_path / "bbva_recent.xlsx"
    with pd.ExcelWriter(bbva_raw, engine="openpyxl") as w:
        pd.DataFrame([["x"] for _ in range(6)]).to_excel(w, index=False, header=False, sheet_name="Movimientos Históricos")
        pd.DataFrame(
            {
                "Fecha": ["06-03-2026"],
                "Concepto": ["TRANSF.BANEL 20301020304"],
                "C3": [""],
                "C4": [""],
                "C5": [""],
                "C6": [""],
                "Crédito": ["154682.13"],
                "Detalle": ["CTE 20301020304"],
            }
        ).to_excel(w, index=False, startrow=6, sheet_name="Movimientos Históricos")

    out_path = tmp_path / "runtime.xlsx"
    meta = build_runtime_workbook_from_raw(
        record_excel_path=str(record_path),
        raw_bank_paths=[str(mp_raw), str(gal_raw), str(bbva_raw)],
        out_excel_path=str(out_path),
    )

    assert meta["raw_min_date_all"] == "2026-02-06"
    assert meta["raw_min_date"] == "2026-03-06"
    assert meta["raw_max_date"] == "2026-03-07"
    assert meta["raw_stale_files_ignored_for_receipts_start_date"] == ["galicia_old.xlsx"]
