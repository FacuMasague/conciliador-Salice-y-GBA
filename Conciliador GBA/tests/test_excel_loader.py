import datetime as dt

import openpyxl

from conciliador.excel_loader import load_bank_txns


def test_load_bank_txns_has_multiple_sources(excel_path):
    txns = load_bank_txns(excel_path)
    assert len(txns) > 100
    origins = {t.origen for t in txns}
    # deberÃÂ­amos ver al menos BBVA/GALICIA/MP en este archivo
    assert "BBVA" in origins
    assert "GALICIA" in origins
    assert "MERCADOPAGO" in origins


def test_load_bank_txns_parsing_flags(excel_path):
    txns = load_bank_txns(excel_path)
    assert any(t.parse_ok for t in txns)
    # No queremos que falle: los errores deben quedar marcados, no crashear
    if any((not t.parse_ok) for t in txns):
        assert all(t.parse_error for t in txns if not t.parse_ok)
    else:
        # dataset puede venir limpio; igual es válido
        assert True


def test_load_bank_txns_date_types(excel_path):
    txns = load_bank_txns(excel_path)
    ok = [t for t in txns if t.parse_ok]
    assert isinstance(ok[0].fecha, dt.date)
    assert isinstance(ok[0].importe, float)


def test_load_bank_txns_bbva_excludes_impuesto_ley(tmp_path):
    p = tmp_path / "bbva_impuesto.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SALICE BBVA"
    ws.append(["BBVA Frances"])
    ws.append(["Fecha", "Número Documento", "Oficina", "Débitos", "Importe", "ok", "cliente", "recibo"])
    ws.append(["23-02-2026", "IMPUESTO LEY 20/02/26 00002", "", "", "1,40", "", "", ""])
    ws.append(["23-02-2026", "TRANSF.BANEL 20302962112", "", "", "88200,00", "", "", ""])
    wb.save(str(p))

    txns = load_bank_txns(str(p))
    bbva = [t for t in txns if t.origen == "BBVA" and t.parse_ok]
    textos = [t.texto_ref for t in bbva]

    assert any("TRANSF.BANEL" in t for t in textos)
    assert not any("IMPUESTO LEY" in t.upper() for t in textos)


def test_load_bank_txns_bbva_keeps_transf_credito_banelco(tmp_path):
    p = tmp_path / "bbva_banelco.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SALICE BBVA"
    ws.append(["BBVA Frances"])
    ws.append(["Fecha", "Número Documento", "Oficina", "Débitos", "Importe", "ok", "cliente", "recibo"])
    ws.append(["23-02-2026", "TRANSF CREDITO BANELCO", "", "", "1.000,00", "", "", ""])
    ws.append(["23-02-2026", "TRANSF.BANEL 20302962112", "", "", "88200,00", "", "", ""])
    wb.save(str(p))

    txns = load_bank_txns(str(p))
    bbva = [t for t in txns if t.origen == "BBVA" and t.parse_ok]
    textos = [t.texto_ref for t in bbva]

    assert any("TRANSF.BANEL" in t for t in textos)
    assert any("TRANSF CREDITO BANELCO" in t.upper() for t in textos)


def test_load_bank_txns_gba_bbva_sheet_ignores_debits_and_keeps_sheet_name(tmp_path):
    p = tmp_path / "gba_bbva.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos gba"
    ws.append(["Fecha", "Fecha Valor", "Concepto", "Número Documento", "Oficina", "Crédito", "Débito", "Detalle", "Acreditado?"])
    ws.append(["20-03-2026", "20-03-2026", "TRANSF.BANEL 20302962112", "", "733 - N/A", 29000.0, "", "CTE 20302962112", ""])
    ws.append(["20-03-2026", "19-03-2026", "IMPUESTO LEY 19/03/26 00146", "", "481 - EMPRESA MAR DEL PLATA", "", -258316.58, "Saldo Disponible", ""])
    wb.save(str(p))

    txns = load_bank_txns(str(p), record_key="bank")
    assert len(txns) == 1
    assert txns[0].origen == "BBVA"
    assert txns[0].sheet_name == "Movimientos gba"
    assert txns[0].record_key == "bank"
    assert txns[0].importe == 29000.0


def test_load_bank_txns_gba_bbva_keeps_all_positive_credit_rows(tmp_path):
    p = tmp_path / "gba_bbva_filter.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos gba"
    ws.append(["Fecha", "Fecha Valor", "Concepto", "Número Documento", "Oficina", "Crédito", "Débito", "Detalle", "Acreditado?"])
    ws.append(["01-04-2026", "01-04-2026", "TRANSF.BANEL 20047049432", "", "733 - N/A", 20086.19, "", "CTE 20047049432", ""])
    ws.append(["01-04-2026", "01-04-2026", "DNET CREDITO NE3117596", "", "587 - DATANET", 364386.04, "", "CTE 117596       007-002389330101       DELTA ISLAND ASES", ""])
    ws.append(["01-04-2026", "01-04-2026", "DEPOSITO AUT BUZON/01/15:12", "", "195 - TORTUGAS OPEN MALL", 194000, "", "CTE 000017917002", ""])
    ws.append(["01-04-2026", "01-04-2026", "TRANSF. CLIE 203407713375", "", "100 - BANCA ONLINE", 987258.52, "", "CTA.ORIGEN:  CC $ 340-771337/5", ""])
    ws.append(["01-04-2026", "01-04-2026", "EFECTIVO", "", "337 - MORON", 4050550, "", "", ""])
    wb.save(str(p))

    txns = load_bank_txns(str(p), record_key="bank")
    assert len(txns) == 5
    textos = [t.texto_ref for t in txns]
    assert any(t.endswith("CTE 20047049432") for t in textos)
    assert any("DNET CREDITO NE3117596" in t for t in textos)
    assert any("DEPOSITO AUT BUZON/01/15:12" in t for t in textos)
    assert any("TRANSF. CLIE 203407713375" in t for t in textos)
    assert any(t == "EFECTIVO" for t in textos)


def test_load_bank_txns_gba_mp_reads_multiple_month_sheets(tmp_path):
    p = tmp_path / "gba_mp.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ENERO"
    ws.append(["Fecha de Pago", "Tipo de Operación", "Operación Relacionada", "Importe", "Control Logistica"])
    ws.append(["2026-01-05T10:00:00.000-04:00", "140447881780", "Dinero disponible", "530866.63", ""])
    ws2 = wb.create_sheet("MARZO")
    ws2.append(["Fecha de Pago", "Tipo de Operación", "Operación Relacionada", "Importe", "Control Logistica"])
    ws2.append(["2026-03-21T12:24:12.000-04:00", "151349770684", "Dinero disponible", "158792.00", ""])
    wb.save(str(p))

    txns = load_bank_txns(str(p), record_key="mp")
    assert len(txns) == 2
    assert {t.sheet_name for t in txns} == {"ENERO", "MARZO"}
    assert all(t.origen == "MERCADOPAGO" for t in txns)
    assert all(t.record_key == "mp" for t in txns)
