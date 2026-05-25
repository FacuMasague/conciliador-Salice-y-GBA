from __future__ import annotations

import openpyxl

from src.conciliador.exporter import export_dudosos_xlsx, export_no_encontrados_xlsx


def test_export_no_encontrados_xlsx_splits_into_3_operational_sheets_without_empty_side_columns(tmp_path):
    out = tmp_path / "no_encontrados.xlsx"
    result = {
        "no_encontrados": [
            {
                "Tipo no encontrado": "BANCO_SIN_RECIBO",
                "Origen": "BBVA",
                "Fecha movimiento": "2026-02-20",
                "Importe movimiento": 1000.0,
                "Detalle movimiento": "Detalle BBVA",
                "Fila Excel": 10,
            },
            {
                "Tipo no encontrado": "BANCO_SIN_RECIBO",
                "Origen": "MERCADOPAGO",
                "Fecha movimiento": "2026-02-21",
                "Importe movimiento": 2000.0,
                "Detalle movimiento": "Detalle MP",
                "Fila Excel": 20,
            },
            {
                "Tipo no encontrado": "BANCO_SIN_RECIBO",
                "Origen": "GALICIA",
                "Fecha movimiento": "2026-02-22",
                "Importe movimiento": 3000.0,
                "Detalle movimiento": "Detalle Galicia",
                "Fila Excel": 30,
            },
            {
                "Tipo no encontrado": "RECIBO_SIN_BANCO",
                "Empresa": "ALARCON",
                "Nro recibo": "68744",
                "Nro cliente": "30424",
                "Cliente": "Cliente Demo",
                "Medio de pago": "TRANSFERENCIA",
                "Fecha recibo": "2026-02-23",
                "Importe recibo": 575511.55,
                "Divisor": "",
            },
        ]
    }

    export_no_encontrados_xlsx(result, str(out))

    wb = openpyxl.load_workbook(str(out), data_only=False)
    assert wb.sheetnames == [
        "Recibos no encontrados",
        "Ingresos MP no encontrados",
        "Ingresos BBVA no encontrados",
    ]

    ws_rec = wb["Recibos no encontrados"]
    ws_mp = wb["Ingresos MP no encontrados"]
    ws_bbva = wb["Ingresos BBVA no encontrados"]

    headers_rec = [c.value for c in ws_rec[1]]
    headers_bbva = [c.value for c in ws_bbva[1]]
    headers_mp = [c.value for c in ws_mp[1]]

    for headers in (headers_rec, headers_bbva, headers_mp):
        assert not any(str(h).startswith("__") for h in headers)

    for headers in (headers_bbva, headers_mp):
        assert "Origen" in headers
        assert "Fecha movimiento" in headers
        assert "Importe movimiento" in headers
        assert "Nro recibo" not in headers
        assert "Nro cliente" not in headers
        assert "Importe recibo" not in headers

    assert "Nro recibo" in headers_rec
    assert "Nro cliente" in headers_rec
    assert "Cliente" in headers_rec
    assert "Importe recibo" in headers_rec
    assert "Origen" not in headers_rec
    assert "Fecha movimiento" not in headers_rec
    assert "Importe movimiento" not in headers_rec
    cliente_col = headers_rec.index("Cliente") + 1
    assert ws_rec.cell(2, cliente_col).value == "Cliente Demo"
    assert ws_bbva.max_row == 3  # BBVA + Galicia se revisan en la misma hoja operativa.


def test_export_no_encontrados_xlsx_formats_money_and_widens_columns(tmp_path):
    out = tmp_path / "no_encontrados_money.xlsx"
    result = {
        "no_encontrados": [
            {
                "Tipo no encontrado": "BANCO_SIN_RECIBO",
                "Origen": "BBVA",
                "Fecha movimiento": "2026-02-20",
                "Importe movimiento": "85500,00",
                "Detalle movimiento": "Transferencia de terceros con detalle largo para validar ancho",
                "Fila Excel": 1334,
            }
        ]
    }

    export_no_encontrados_xlsx(result, str(out))

    wb = openpyxl.load_workbook(str(out), data_only=False)
    ws = wb["Ingresos BBVA no encontrados"]
    headers = [c.value for c in ws[1]]
    importe_col = headers.index("Importe movimiento") + 1
    detalle_col = headers.index("Detalle movimiento") + 1
    importe_cell = ws.cell(2, importe_col)
    importe_col_letter = openpyxl.utils.get_column_letter(importe_col)
    detalle_col_letter = openpyxl.utils.get_column_letter(detalle_col)

    assert isinstance(importe_cell.value, (int, float))
    assert importe_cell.number_format == "#,##0.##"
    assert (ws.column_dimensions[importe_col_letter].width or 0) >= 18
    assert (ws.column_dimensions[detalle_col_letter].width or 0) >= 30


def test_export_no_encontrados_xlsx_formats_mp_dates_as_dd_mm_yyyy(tmp_path):
    out = tmp_path / "no_encontrados_mp_dates.xlsx"
    result = {
        "no_encontrados": [
            {
                "Tipo no encontrado": "BANCO_SIN_RECIBO",
                "Origen": "MERCADOPAGO",
                "Fecha movimiento": "2026-03-09",
                "Importe movimiento": 123.0,
                "Detalle movimiento": "Detalle MP",
                "Fila Excel": 20,
            }
        ]
    }

    export_no_encontrados_xlsx(result, str(out))

    wb = openpyxl.load_workbook(str(out), data_only=False)
    ws = wb["Ingresos MP no encontrados"]
    headers = [c.value for c in ws[1]]
    fecha_col = headers.index("Fecha movimiento") + 1
    assert ws.cell(2, fecha_col).value == "09/03/2026"


def test_export_dudosos_xlsx_includes_active_and_deleted_rows_by_origin(tmp_path):
    out = tmp_path / "dudosos.xlsx"
    result = {
        "dudosos": [
            {
                "Estado dudoso": "Activo",
                "Tipo fila": "PRINCIPAL",
                "Ranking": 1,
                "Nro recibo": "100",
                "Nro cliente": "10",
                "Cliente": "Cliente BBVA",
                "Medio de pago": "Transferencia",
                "Fecha recibo": "2026-04-06",
                "Importe recibo": 1000.0,
                "Origen": "BBVA",
                "Fecha movimiento": "2026-04-06",
                "Importe movimiento": 1000.0,
                "Detalle movimiento": "Banco",
                "Fila Excel": 12,
                "Peso": 10,
            }
        ],
        "dudosos_borrados": [
            {
                "Tipo fila": "PRINCIPAL",
                "Ranking": 1,
                "Nro recibo": "200",
                "Nro cliente": "20",
                "Cliente": "Cliente MP",
                "Medio de pago": "Mercado Pago",
                "Fecha recibo": "2026-04-07",
                "Importe recibo": 2000.0,
                "Origen": "MERCADOPAGO",
                "Fecha movimiento": "2026-04-07",
                "Importe movimiento": 2000.0,
                "Detalle movimiento": "MP",
                "Fila Excel": 22,
                "Peso": 20,
            }
        ],
    }

    export_dudosos_xlsx(result, str(out))

    wb = openpyxl.load_workbook(str(out), data_only=False)
    assert wb.sheetnames == ["Dudosos Mercado Pago", "Dudosos BBVA"]
    ws_mp = wb["Dudosos Mercado Pago"]
    ws_bbva = wb["Dudosos BBVA"]
    mp_headers = [c.value for c in ws_mp[1]]
    bbva_headers = [c.value for c in ws_bbva[1]]
    assert "Estado dudoso" not in mp_headers
    assert "Estado dudoso" not in bbva_headers
    assert "Ranking" not in mp_headers
    assert "Ranking" not in bbva_headers
    assert "Divisor" in mp_headers
    assert "Divisor" in bbva_headers
    assert ws_mp.row_dimensions[1].height == 24
    assert ws_mp.row_dimensions[2].height == 20
