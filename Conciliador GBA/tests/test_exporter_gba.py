from __future__ import annotations

import openpyxl

from src.conciliador.exporter import export_combined_records_excel, export_filled_generic_excel
from src.conciliador.excel_loader import load_bank_txns


def test_export_filled_generic_excel_writes_tracking_columns_into_gba_bank_record(tmp_path):
    p = tmp_path / "bank.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimientos gba"
    ws.append(["Fecha", "Fecha Valor", "Concepto", "Número Documento", "Oficina", "Crédito", "Débito", "Detalle", "Acreditado?"])
    ws.append(["20-03-2026", "20-03-2026", "DEPOSITO AUT BUZON/01/07:02", "", "331 - PANAMERICANA", 29000.0, "", "CTE 000000495002", ""])
    wb.save(p)

    result = {
        "validados": [
            {
                "Ranking": 1,
                "Origen": "BBVA",
                "Fila Excel": 2,
                "Nro cliente": "33119",
                "Cliente": "Cliente GBA",
                "Vendedor/Repartidor": "65 - Yanina Andrade",
                "Nro recibo": "272641",
                "Fecha recibo": "2026-03-20",
                "Medio de pago": "TRANSFERENCIA",
                "Importe recibo": 29000.0,
                "__sheet_name": "Movimientos gba",
                "__record_key": "bank",
            }
        ]
    }

    out = tmp_path / "bank_out.xlsx"
    export_filled_generic_excel(str(p), result, str(out), allowed_origins={"BBVA"}, record_key="bank")

    wb_out = openpyxl.load_workbook(out, data_only=True)
    ws_out = wb_out["Movimientos gba"]
    headers = [str(ws_out.cell(1, c).value or "").strip().lower() for c in range(1, ws_out.max_column + 1)]
    assert "ok" in headers
    assert "cliente nombre" in headers
    assert "recibo" in headers
    assert "vendedor/fletero" in headers
    ok_col = headers.index("ok") + 1
    cliente_nombre_col = headers.index("cliente nombre") + 1
    recibo_col = headers.index("recibo") + 1
    vendedor_col = headers.index("vendedor/fletero") + 1
    assert ws_out.cell(2, ok_col).value == "ok"
    assert ws_out.cell(2, cliente_nombre_col).value == "Cliente GBA"
    assert str(ws_out.cell(2, recibo_col).value) == "272641"
    assert ws_out.cell(2, vendedor_col).value == "65 - Yanina Andrade"

    txns = load_bank_txns(str(out), record_key="bank")
    assert len(txns) == 1
    assert txns[0].was_preconciled is True
    assert txns[0].preconciled_recibo == "272641"


def test_export_filled_generic_excel_writes_into_exact_mp_month_sheet(tmp_path):
    p = tmp_path / "mp.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ENERO"
    ws.append(["Fecha de Pago", "Tipo de Operación", "Operación Relacionada", "Importe", "Control Logistica"])
    ws.append(["2026-01-05T10:00:00.000-04:00", "140447881780", "Dinero disponible", "530866.63", ""])
    ws2 = wb.create_sheet("MARZO")
    ws2.append(["Fecha de Pago", "Tipo de Operación", "Operación Relacionada", "Importe", "Control Logistica"])
    ws2.append(["2026-03-21T12:24:12.000-04:00", "151349770684", "Dinero disponible", "158792.00", ""])
    wb.save(p)

    result = {
        "validados": [
            {
                "Ranking": 1,
                "Origen": "MERCADOPAGO",
                "Fila Excel": 2,
                "Nro cliente": "33119",
                "Cliente": "Cliente MP",
                "Nro recibo": "272642",
                "Fecha recibo": "2026-03-21",
                "Medio de pago": "MERCADOPAGO",
                "Importe recibo": 158792.0,
                "__sheet_name": "MARZO",
                "__record_key": "mp",
            }
        ]
    }

    out = tmp_path / "mp_out.xlsx"
    export_filled_generic_excel(str(p), result, str(out), allowed_origins={"MERCADOPAGO"}, record_key="mp")

    wb_out = openpyxl.load_workbook(out, data_only=True)
    ws_marzo = wb_out["MARZO"]
    ws_enero = wb_out["ENERO"]
    headers_marzo = [str(ws_marzo.cell(1, c).value or "").strip().lower() for c in range(1, ws_marzo.max_column + 1)]
    headers_enero = [str(ws_enero.cell(1, c).value or "").strip().lower() for c in range(1, ws_enero.max_column + 1)]
    assert "recibo" in headers_marzo
    assert "recibo" not in headers_enero
    recibo_col = headers_marzo.index("recibo") + 1
    assert str(ws_marzo.cell(2, recibo_col).value) == "272642"


def test_export_filled_generic_excel_forces_mp_operation_id_text(tmp_path):
    p = tmp_path / "mp_numeric_ids.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MAYO"
    ws.append(["Fecha de Pago", "Tipo de Operación", "Operación Relacionada", "Importe", "Control Logistica"])
    ws.append(["2026-05-02 09:47:06", 157000000000.0, "available_money", 13524.57, ""])
    ws.cell(2, 2).number_format = "0.00E+00"
    wb.save(p)

    result = {
        "validados": [
            {
                "Ranking": 1,
                "Origen": "MERCADOPAGO",
                "Fila Excel": 2,
                "Nro cliente": "10054",
                "Cliente": "Cliente MP",
                "Nro recibo": "92039",
                "Fecha recibo": "2026-05-02",
                "Medio de pago": "MERCADOPAGO",
                "Importe recibo": 13524.57,
                "__sheet_name": "MAYO",
                "__record_key": "mp",
            }
        ]
    }

    out = tmp_path / "mp_numeric_ids_out.xlsx"
    export_filled_generic_excel(str(p), result, str(out), allowed_origins={"MERCADOPAGO"}, record_key="mp")

    wb_out = openpyxl.load_workbook(out, data_only=False)
    cell = wb_out["MAYO"].cell(2, 2)
    assert cell.value == "157000000000"
    assert cell.data_type == "s"
    assert cell.number_format == "@"


def test_export_combined_records_prefers_touched_mp_month_sheet(tmp_path):
    mp_path = tmp_path / "mp_runtime.xlsx"
    wb = openpyxl.Workbook()
    ws_junio = wb.active
    ws_junio.title = "JUNIO"
    ws_junio.append(["Fecha de Pago", "Tipo de Operación", "Operación Relacionada", "Importe", "Control Logistica"])
    ws_junio.append(["2026-06-30T13:45:15.000-04:00", "166510534466", "available_money", 1338227.19, ""])
    ws_julio = wb.create_sheet("JULIO")
    ws_julio.append(["Fecha de Pago", "Tipo de Operación", "Operación Relacionada", "Importe", "Control Logistica"])
    ws_julio.append(["2026-07-01T08:27:10", "165790389871", "available_money", 1000.0, ""])
    wb.save(mp_path)

    out = tmp_path / "combined.xlsx"
    export_combined_records_excel(
        [
            {
                "key": "mp",
                "working_excel_path": str(mp_path),
                "base_excel_filename": "Mercado Pago.xlsx",
                "origins": ["MERCADOPAGO"],
                "export_mode": "generic",
            }
        ],
        {
            "validados": [],
            "meta": {
                "raw_ingestion_summary": {
                    "MERCADOPAGO": {"input": 1, "appended": 1, "duplicates_skipped": 0, "sheet": "JULIO"}
                }
            },
        },
        str(out),
    )

    wb_out = openpyxl.load_workbook(out, data_only=True)
    ws_out = wb_out["Mercado Pago"]
    assert ws_out.cell(2, 1).value == "2026-07-01T08:27:10"
    assert str(ws_out.cell(2, 2).value) == "165790389871"
