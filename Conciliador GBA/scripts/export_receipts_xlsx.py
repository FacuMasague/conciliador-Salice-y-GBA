#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.conciliador.external.receipts_api_client import fetch_receipts_payload
from src.conciliador.external.service import (
    _collect_text_fragments,
    _medio_from_forma_pago,
    _resolve_empresa,
)


def _compact_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def _clean_text(value: Any) -> str:
    s = str(value or "")
    s = s.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    return " ".join(s.split()).strip()


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = _clean_text(value).replace("$", "").replace(" ", "")
    if not s:
        return None
    if "." in s and "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def _formas_by_id(formas_rows: List[dict]) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for fr in formas_rows:
        if not isinstance(fr, dict):
            continue
        fid = str(fr.get("formaDePagoID") or "").strip()
        desc = " | ".join(
            [
                str(fr.get("descripcion") or "").strip(),
                str(fr.get("denominacion") or "").strip(),
                str(fr.get("nombre") or "").strip(),
                str(fr.get("codigo") or "").strip(),
            ]
        ).strip(" |")
        if fid:
            out[fid] = desc
    return out


def _comprobante_empresa_id(c: dict) -> str:
    for v in (
        c.get("sucursal_id"),
        c.get("sucursalID"),
        c.get("sucursalId"),
        c.get("empresaID"),
        c.get("empresa_id"),
        c.get("empresaId"),
    ):
        s = str(v or "").strip()
        if not s:
            continue
        digits = "".join(ch for ch in s if ch.isdigit())
        if digits:
            try:
                return str(int(digits))
            except Exception:
                return digits.lstrip("0") or "0"
        return s
    return ""


def _empresas_by_id(empresas_rows: List[dict]) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for er in empresas_rows:
        if not isinstance(er, dict):
            continue
        eid = str(
            er.get("empresaID")
            or er.get("id")
            or er.get("empresaId")
            or er.get("codigo")
            or ""
        ).strip()
        ename = " | ".join(
            [
                str(er.get("descripcion") or "").strip(),
                str(er.get("denominacion") or "").strip(),
                str(er.get("nombre") or "").strip(),
                str(er.get("razonSocial") or "").strip(),
            ]
        ).strip(" |")
        if eid:
            out[eid] = ename
    return out


def _format_header(ws, col_count: int) -> None:
    fill = PatternFill(fill_type="solid", fgColor="E8F0FE")
    for c in range(1, col_count + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws, max_width: int = 60) -> None:
    for col in ws.columns:
        width = 10
        col_letter = col[0].column_letter
        for cell in col:
            txt = _clean_text(cell.value)
            if txt:
                width = max(width, min(len(txt) + 2, max_width))
        ws.column_dimensions[col_letter].width = width


def _write_sheet(ws, headers: List[str], rows: List[List[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _format_header(ws, len(headers))
    _autosize(ws)


def build_workbook(days: int, empresa_filter: str | None, include_raw: bool = False) -> Workbook:
    resp = fetch_receipts_payload(days=days, empresa_filter=empresa_filter)
    payload = resp.payload if isinstance(resp.payload, dict) else {}

    comprobantes = payload.get("comprobantes") if isinstance(payload.get("comprobantes"), list) else []
    formas = payload.get("formasDePago") if isinstance(payload.get("formasDePago"), list) else []
    empresas = payload.get("empresas") if isinstance(payload.get("empresas"), list) else []
    formas_by_id = _formas_by_id(formas)
    empresas_by = _empresas_by_id(empresas)

    wb = Workbook()
    wb.remove(wb.active)

    meta_rows = [
        ["generated_at_utc", datetime.now(timezone.utc).isoformat()],
        ["request_id", str(resp.request_id or "")],
        ["days", int(days)],
        ["empresa_filter", str(empresa_filter or "")],
        ["warnings_count", len(list(resp.warnings or []))],
        ["warnings", " | ".join([_clean_text(w) for w in (resp.warnings or [])])],
        ["comprobantes_count", len(comprobantes)],
        ["formas_pago_count", len(formas)],
        ["empresas_count", len(empresas)],
    ]
    ws_meta = wb.create_sheet("meta")
    _write_sheet(ws_meta, ["key", "value"], meta_rows)

    ws_formas = wb.create_sheet("formas_pago")
    formas_rows: List[List[Any]] = []
    for i, fr in enumerate(formas, start=1):
        fid = str(fr.get("formaDePagoID") or "").strip()
        desc = " | ".join(
            [
                str(fr.get("descripcion") or "").strip(),
                str(fr.get("denominacion") or "").strip(),
                str(fr.get("nombre") or "").strip(),
                str(fr.get("codigo") or "").strip(),
            ]
        ).strip(" |")
        formas_rows.append([i, fid, _clean_text(desc)])
    _write_sheet(ws_formas, ["row_num", "forma_pago_id", "descripcion"], formas_rows)

    ws_emp = wb.create_sheet("empresas")
    emp_rows: List[List[Any]] = []
    for i, er in enumerate(empresas, start=1):
        eid = str(
            er.get("empresaID")
            or er.get("id")
            or er.get("empresaId")
            or er.get("codigo")
            or ""
        ).strip()
        desc = " | ".join(
            [
                str(er.get("descripcion") or "").strip(),
                str(er.get("denominacion") or "").strip(),
                str(er.get("nombre") or "").strip(),
                str(er.get("razonSocial") or "").strip(),
            ]
        ).strip(" |")
        emp_rows.append([i, eid, _clean_text(desc)])
    _write_sheet(ws_emp, ["row_num", "empresa_id", "descripcion"], emp_rows)

    ws_comp = wb.create_sheet("comprobantes_limpio")
    comp_rows: List[List[Any]] = []
    for i, c in enumerate(comprobantes, start=1):
        if not isinstance(c, dict):
            continue
        empresa_id = _comprobante_empresa_id(c)
        empresa_nombre_api = _clean_text(
            empresas_by.get(empresa_id)
            or c.get("empresa")
            or c.get("nombreEmpresa")
            or c.get("empresaNombre")
            or c.get("razonSocialEmpresa")
        )
        company_hint = " | ".join(
            [
                str(c.get("empresa") or ""),
                str(c.get("nombreEmpresa") or ""),
                str(c.get("empresaNombre") or ""),
                str(c.get("razonSocialEmpresa") or ""),
            ]
        )
        empresa_detectada = _resolve_empresa(
            empresa_id,
            empresa_filter=empresa_filter,
            empresas_by_id=empresas_by,
            company_hint=company_hint,
        )

        hint = " ".join(
            [
                str(c.get("notas") or ""),
                str(c.get("notas2") or ""),
                str(c.get("subtipo") or ""),
                str(c.get("serie") or ""),
                str(c.get("codigoDeImportacion") or ""),
                " ".join(_collect_text_fragments(c)),
            ]
        )
        medio, warning = _medio_from_forma_pago(
            c.get("formaDePagoID"),
            formas_by_id,
            desc_candidates=[
                str(c.get("formaDePago") or ""),
                str(c.get("formaPago") or ""),
                str(c.get("medioPago") or ""),
                str(c.get("metodoPago") or ""),
                str(c.get("canalDeCobro") or ""),
                str(c.get("tipoDeCobro") or ""),
                str(c.get("origenDelCobro") or ""),
                str(c.get("descripcionFormaDePago") or ""),
                str(c.get("denominacionFormaDePago") or ""),
            ],
            hint_text=hint,
        )

        fid = str(c.get("formaDePagoID") or "").strip()
        importe_num = _to_float(c.get("importeTotal"))
        codigo_import = _clean_text(c.get("codigoDeImportacion"))
        nro_pm = "".join(ch for ch in codigo_import if ch.isdigit()) if codigo_import else ""
        nro_pm = str(int(nro_pm)) if nro_pm.isdigit() else nro_pm
        comp_rows.append(
            [
                i,
                empresa_id,
                empresa_nombre_api,
                empresa_detectada,
                fid,
                _clean_text(formas_by_id.get(fid, "")),
                _clean_text(c.get("formaDePago") or c.get("formaPago") or c.get("medioPago") or c.get("metodoPago")),
                _clean_text(c.get("canalDeCobro") or c.get("tipoDeCobro") or c.get("origenDelCobro")),
                _clean_text(medio),
                _clean_text(warning),
                nro_pm,
                c.get("numero"),
                codigo_import,
                c.get("clienteID"),
                _clean_text(c.get("razonSocial")),
                _clean_text(c.get("fechaDeEmision") or c.get("fechaDePrimerVencimiento")),
                importe_num,
                _clean_text(c.get("notas")),
                _clean_text(c.get("notas2")),
            ]
        )
    comp_headers = [
        "row_num",
        "empresa_id",
        "empresa_api_nombre",
        "empresa_detectada",
        "forma_pago_id",
        "forma_pago_desc",
        "forma_pago_comprobante",
        "canal_cobro_comprobante",
        "medio_detectado",
        "warning_medio",
        "nro_recibo_pm",
        "nro_comprobante_esi",
        "codigo_importacion",
        "cliente_id",
        "razon_social",
        "fecha_emision",
        "importe_total",
        "notas",
        "notas2",
    ]
    _write_sheet(ws_comp, comp_headers, comp_rows)
    ws_comp.column_dimensions["Q"].width = 16
    for r in range(2, ws_comp.max_row + 1):
        ws_comp.cell(row=r, column=17).number_format = '#,##0.00'

    if include_raw:
        ws_raw = wb.create_sheet("comprobantes_raw")
        raw_rows: List[List[Any]] = []
        for i, c in enumerate(comprobantes, start=1):
            if isinstance(c, dict):
                raw_rows.append([i, _compact_json(c)])
        _write_sheet(ws_raw, ["row_num", "raw_json"], raw_rows)
    return wb


def main() -> int:
    parser = argparse.ArgumentParser(description="Exporta recibos API GESI a Excel prolijo para diagnóstico.")
    parser.add_argument("--days", type=int, default=15, help="Ventana de días (default: 15)")
    parser.add_argument("--empresa-filter", default=None, help="Filtro opcional de empresa")
    parser.add_argument("--out", default="receipts_last_15_days.xlsx", help="Ruta del archivo XLSX de salida")
    parser.add_argument("--include-raw", action="store_true", help="Incluye hoja extra con JSON crudo de comprobantes")
    args = parser.parse_args()

    wb = build_workbook(days=int(args.days), empresa_filter=args.empresa_filter, include_raw=bool(args.include_raw))
    out_path = Path(args.out).resolve()
    wb.save(out_path)
    print(str(out_path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
