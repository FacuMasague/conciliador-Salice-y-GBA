#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List
from urllib.parse import urlencode

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.conciliador.external.receipts_api_client import (
    _base_url,
    _build_auth_headers_for_empresa,
    _empresa_name_from_id,
    _extract_paginacion,
    _extract_rows,
    _headers_base,
    _http_json,
    _resolve_empresa_targets,
)
from src.conciliador.external.errors import ExternalProviderError


def _compact_json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def _clean_text(value: Any) -> str:
    s = str(value or "")
    s = s.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    return " ".join(s.split()).strip()


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = _clean_text(value).replace("$", "").replace(" ", "")
    if not s:
        return None
    if "." in s and "," in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def _format_header(ws, col_count: int) -> None:
    fill = PatternFill(fill_type="solid", fgColor="E8F0FE")
    for c in range(1, col_count + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws, max_width: int = 70) -> None:
    for col in ws.columns:
        width = 10
        col_letter = col[0].column_letter
        for cell in col:
            txt = _clean_text(cell.value)
            if txt:
                width = max(width, min(len(txt) + 2, max_width))
        ws.column_dimensions[col_letter].width = width


def _write_sheet(ws, headers: List[str], rows: List[List[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _format_header(ws, len(headers))
    _autosize(ws)


def _iter_keys(rows: Iterable[dict]) -> Dict[str, int]:
    keys: Dict[str, int] = {}
    for r in rows:
        if not isinstance(r, dict):
            continue
        for k, v in r.items():
            kk = str(k)
            if v is None or _clean_text(v) == "":
                keys.setdefault(kk, 0)
            else:
                keys[kk] = keys.get(kk, 0) + 1
    return keys


def _comprobante_empresa_id(c: dict) -> str:
    for v in (
        c.get("sucursal_id"),
        c.get("sucursalID"),
        c.get("sucursalId"),
        c.get("empresaID"),
        c.get("empresa_id"),
        c.get("empresaId"),
    ):
        s = _clean_text(v)
        if not s:
            continue
        digits = "".join(ch for ch in s if ch.isdigit())
        if digits:
            try:
                return str(int(digits))
            except Exception:
                return digits.lstrip("0") or "0"
        return s
    return ""


def _build_candidate_paths(endpoint_path: str | None) -> list[str]:
    requested = _clean_text(endpoint_path)
    candidates: list[str] = []
    if requested and requested.lower() != "auto":
        candidates.append(requested if requested.startswith("/") else f"/{requested}")

    defaults = [
        "/api/Ventas/Comprobantes/GetList",
        "/api/Ventas/Comprobantes/GetListComprobantes",
        "/api/Ventas/Comprobantes/List",
        "/api/Ventas/Comprobante/GetList",
        "/api/Maestros/Comprobantes/GetList",
    ]
    for d in defaults:
        if d not in candidates:
            candidates.append(d)
    return candidates


def fetch_comprobantes_payload(*, days: int, empresa_filter: str | None = None, page_size: int = 500, drop_sucursal_multi: bool = True, endpoint_path: str = "/api/Ventas/Comprobantes/GetList") -> dict[str, Any]:
    base = _base_url("RECEIPTS_API")
    headers_root = _headers_base("RECEIPTS_API")
    targets = _resolve_empresa_targets(empresa_filter)

    candidate_paths = _build_candidate_paths(endpoint_path)

    fecha_hasta = date.today() - timedelta(days=1)
    fecha_desde = fecha_hasta - timedelta(days=max(int(days) - 1, 0))

    base_body: Dict[str, Any] = {
        "datosOperacion": {
            "FechaDesde": fecha_desde.isoformat(),
            "FechaHasta": fecha_hasta.isoformat(),
        },
        "datosClientes": {},
    }

    comprobantes: List[dict] = []
    warnings: list[str] = []
    request_id_last: str | None = None
    counts_by_target: dict[str, int] = {}
    methods_used: dict[str, str] = {}

    selected_path: str | None = None

    for empresa_id in targets:
        empresa_name = _empresa_name_from_id(empresa_id)
        added_for_target = 0

        attempts = [
            ("post", drop_sucursal_multi and len(targets) > 1, True),
            ("post", False, True),
            ("post", False, False),
            ("get", False, False),
        ]
        for method, drop_sucursal, include_empresa_in_body in attempts:
            if added_for_target > 0:
                break

            headers_emp = _build_auth_headers_for_empresa(
                base=base,
                headers_root=headers_root,
                empresa_id=str(empresa_id),
                drop_sucursal=drop_sucursal,
            )

            body = json.loads(json.dumps(base_body))
            if include_empresa_in_body:
                body.setdefault("datosOperacion", {})
                body["datosOperacion"]["EmpresaID"] = int(str(empresa_id))
                body["datosOperacion"]["empresaID"] = int(str(empresa_id))
                body.setdefault("datosClientes", {})
                body["datosClientes"]["EmpresaID"] = int(str(empresa_id))
                body["empresaID"] = int(str(empresa_id))

            page = 1
            rows_added_this_attempt = 0
            for path in candidate_paths:
                path_not_found = False
                page = 1
                rows_added_this_path = 0
                while True:
                    q = {"pageNumber": str(page), "pageSize": str(page_size)}
                    # Endpoint de catálogo en Maestros (según Swagger):
                    # requiere estos query params para evitar 400 por contrato.
                    if path.lower() == "/api/maestros/comprobantes/getlist":
                        q["comprobanteID"] = "0"
                        q["claseDeComprobanteID"] = "0"
                    url = f"{base}{path}?{urlencode(q)}"
                    try:
                        payload, request_id = _http_json(
                            url,
                            method=method.upper(),
                            headers=headers_emp,
                            body=body if method == "post" else None,
                        )
                    except ExternalProviderError as e:
                        status_code = int(getattr(e, "status_code", 0) or 0)
                        if status_code in (404, 405):
                            path_not_found = True
                            warnings.append(
                                f"Comprobantes path descartado por HTTP {status_code} (empresaID={empresa_id}, method={method}): {path}"
                            )
                            break
                        raise
                    request_id_last = request_id or request_id_last

                    success = payload.get("success")
                    if success is False:
                        err = payload.get("error")
                        msg = f"GetList Comprobantes devolvió error (empresaID={empresa_id})"
                        if isinstance(err, dict):
                            msg = str(err.get("message") or msg)
                        raise ExternalProviderError("receipts", msg)

                    rows = _extract_rows(payload, "comprobantes")
                    if rows is None:
                        rows = _extract_rows(payload, "items")
                    if rows is None:
                        rows = _extract_rows(payload, "rows")
                    if rows is None:
                        warnings.append(
                            f"Comprobantes/GetList sin lista de filas (empresaID={empresa_id}, page={page}, method={method}, path={path}); se toma vacío."
                        )
                        rows = []

                    for r in rows:
                        rid = str(r.get("empresaID") or "").strip()
                        if not rid:
                            r["empresaID"] = empresa_id
                        if empresa_name:
                            r.setdefault("_empresa_target_name", empresa_name)
                        comprobantes.append(r)
                        rows_added_this_attempt += 1
                        rows_added_this_path += 1

                    pag = _extract_paginacion(payload)
                    if not isinstance(pag, dict):
                        if len(rows) >= int(page_size):
                            page += 1
                            if page > 500:
                                warnings.append(f"Corte de paginación de seguridad en Comprobantes/GetList (empresaID={empresa_id}, path={path})")
                                break
                            continue
                        break
                    try:
                        tp = int(
                            pag.get("totalPaginas")
                            or pag.get("totalpages")
                            or pag.get("totalPages")
                            or pag.get("pages")
                        )
                    except Exception:
                        break
                    if page >= tp:
                        break
                    page += 1
                    if page > 500:
                        warnings.append(f"Corte de paginación de seguridad en Comprobantes/GetList (empresaID={empresa_id}, path={path})")
                        break

                if path_not_found:
                    continue
                selected_path = path
                if rows_added_this_path == 0:
                    warnings.append(
                        f"Comprobantes/GetList path válido sin filas (empresaID={empresa_id}, method={method}): {path}"
                    )
                break

            if rows_added_this_attempt > 0:
                added_for_target += rows_added_this_attempt
                methods_used[str(empresa_id)] = method
            else:
                warnings.append(
                    f"Comprobantes/GetList empresaID={empresa_id}: 0 filas con intento method={method}, drop_sucursal={drop_sucursal}, empresa_in_body={include_empresa_in_body}."
                )

        counts_by_target[str(empresa_id)] = int(added_for_target)

    keys_stats = _iter_keys(comprobantes)

    return {
        "meta": {
            "request_id": request_id_last or "",
            "fecha_desde": fecha_desde.isoformat(),
            "fecha_hasta": fecha_hasta.isoformat(),
            "days": int(days),
            "empresa_filter": str(empresa_filter or ""),
            "endpoint_path": str(selected_path or ""),
            "endpoint_candidates": candidate_paths,
            "targets": targets,
            "counts_by_target": counts_by_target,
            "methods_used": methods_used,
            "warnings": warnings,
        },
        "comprobantes": comprobantes,
        "keys_stats": keys_stats,
    }


def build_workbook(days: int, empresa_filter: str | None, endpoint_path: str, include_raw: bool = True) -> Workbook:
    data = fetch_comprobantes_payload(
        days=days,
        empresa_filter=empresa_filter,
        endpoint_path=endpoint_path,
    )
    meta = data.get("meta") if isinstance(data.get("meta"), dict) else {}
    comprobantes = data.get("comprobantes") if isinstance(data.get("comprobantes"), list) else []
    keys_stats = data.get("keys_stats") if isinstance(data.get("keys_stats"), dict) else {}

    wb = Workbook()
    wb.remove(wb.active)

    ws_meta = wb.create_sheet("meta")
    meta_rows = [
        ["generated_at_utc", datetime.now(timezone.utc).isoformat()],
        ["request_id", str(meta.get("request_id") or "")],
        ["endpoint_path", str(meta.get("endpoint_path") or "")],
        ["days", int(meta.get("days") or days)],
        ["fecha_desde", str(meta.get("fecha_desde") or "")],
        ["fecha_hasta", str(meta.get("fecha_hasta") or "")],
        ["empresa_filter", str(meta.get("empresa_filter") or "")],
        ["targets", _clean_text(", ".join([str(x) for x in (meta.get("targets") or [])]))],
        ["counts_by_target", _compact_json(meta.get("counts_by_target") or {})],
        ["methods_used", _compact_json(meta.get("methods_used") or {})],
        ["warnings_count", len(list(meta.get("warnings") or []))],
        ["warnings", " | ".join([_clean_text(w) for w in (meta.get("warnings") or [])])],
        ["comprobantes_count", len(comprobantes)],
    ]
    _write_sheet(ws_meta, ["key", "value"], meta_rows)

    ws_keys = wb.create_sheet("keys_stats")
    ks_rows = [[k, int(v)] for k, v in sorted(keys_stats.items(), key=lambda x: (-x[1], x[0]))]
    _write_sheet(ws_keys, ["field", "non_empty_count"], ks_rows)

    ws_comp = wb.create_sheet("comprobantes_limpio")
    comp_headers = [
        "row_num",
        "empresa_id",
        "empresa_target",
        "nro_comprobante",
        "nro_recibo_pm_extraido",
        "codigo_importacion",
        "cliente_id",
        "razon_social",
        "forma_pago_id",
        "forma_pago_texto",
        "fecha_emision",
        "fecha_venc_1",
        "importe_total",
        "estado",
        "tipo",
        "subtipo",
        "serie",
        "sucursal_id",
        "notas",
        "notas2",
    ]
    comp_rows: List[List[Any]] = []
    for i, c in enumerate(comprobantes, start=1):
        if not isinstance(c, dict):
            continue
        codigo_import = _clean_text(c.get("codigoDeImportacion"))
        nro_pm = "".join(ch for ch in codigo_import if ch.isdigit()) if codigo_import else ""
        nro_pm = str(int(nro_pm)) if nro_pm.isdigit() else ""
        comp_rows.append(
            [
                i,
                _comprobante_empresa_id(c),
                _clean_text(c.get("_empresa_target_name")),
                _clean_text(c.get("numero")),
                nro_pm,
                codigo_import,
                _clean_text(c.get("clienteID")),
                _clean_text(c.get("razonSocial")),
                _clean_text(c.get("formaDePagoID")),
                _clean_text(c.get("formaDePago") or c.get("formaPago") or c.get("medioPago") or c.get("metodoPago") or c.get("descripcionFormaDePago")),
                _clean_text(c.get("fechaDeEmision")),
                _clean_text(c.get("fechaDePrimerVencimiento")),
                _to_float(c.get("importeTotal")),
                _clean_text(c.get("estado")),
                _clean_text(c.get("tipo")),
                _clean_text(c.get("subtipo")),
                _clean_text(c.get("serie")),
                _clean_text(c.get("sucursalID") or c.get("sucursalId")),
                _clean_text(c.get("notas")),
                _clean_text(c.get("notas2")),
            ]
        )
    _write_sheet(ws_comp, comp_headers, comp_rows)
    for r in range(2, ws_comp.max_row + 1):
        ws_comp.cell(row=r, column=13).number_format = '#,##0.00'

    if include_raw:
        ws_raw = wb.create_sheet("comprobantes_raw")
        raw_rows = []
        for i, c in enumerate(comprobantes, start=1):
            if isinstance(c, dict):
                raw_rows.append([i, _compact_json(c)])
        _write_sheet(ws_raw, ["row_num", "raw_json"], raw_rows)

    return wb


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Exporta Comprobantes API GESI (no Cobros) a un XLSX de diagnóstico."
    )
    parser.add_argument("--days", type=int, default=15, help="Ventana de días (default: 15)")
    parser.add_argument("--empresa-filter", default=None, help="Filtro opcional de empresa (SALICE, ALARCON, 3, 6)")
    parser.add_argument("--endpoint-path", default="auto", help="Path del endpoint de comprobantes (default: auto)")
    parser.add_argument("--out", default="comprobantes_last_15_days.xlsx", help="Ruta del XLSX de salida")
    parser.add_argument("--no-raw", action="store_true", help="No incluir hoja de JSON crudo")
    args = parser.parse_args()

    wb = build_workbook(
        days=int(args.days),
        empresa_filter=args.empresa_filter,
        endpoint_path=args.endpoint_path,
        include_raw=not bool(args.no_raw),
    )
    out_path = Path(args.out).resolve()
    wb.save(out_path)
    print(str(out_path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
